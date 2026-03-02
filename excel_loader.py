# excel_loader.py
# Загрузка данных из Excel (products.xlsx):
# Листы:
#   - "продукт"    : товары
#   - "изготовил"  : сотрудники, которые "Изготовил"
#   - "проверил"   : сотрудники, которые "Проверил"
#
# Требования к колонкам (по-русски):
#   Лист "продукт":
#     Код | Наименование | Срок годности (ч) | Ед. измер. | Активен | Комментарий
#
#   Лист "изготовил" / "проверил":
#     ФИО | Активен
#
# Примечание по единицам:
#   В Excel можно писать: "кг,шт" / "кг;шт" / "kg,pcs" / "kg;pcs"
#   На выходе будет список: ["кг","шт"]

from __future__ import annotations

from typing import Any, Dict, List
from openpyxl import load_workbook


def _norm_header(s: Any) -> str:
    """Нормализуем заголовок колонки для поиска (регистр, пробелы)."""
    if s is None:
        return ""
    return str(s).strip().lower()


def _get(row: List[Any], idx: int, default: Any = "") -> Any:
    """Безопасно получить значение из row по индексу."""
    if idx is None:
        return default
    if idx < 0 or idx >= len(row):
        return default
    v = row[idx]
    return default if v is None else v


def _parse_yes_no(v: Any) -> int:
    """Активен: поддержка 1/0, Да/Нет, True/False."""
    if v is None:
        return 0
    if isinstance(v, bool):
        return 1 if v else 0
    s = str(v).strip().lower()
    if s in ("1", "да", "true", "yes", "y", "on"):
        return 1
    if s in ("0", "нет", "false", "no", "n", "off", ""):
        return 0
    # Если в Excel кто-то написал что-то странное — считаем активным, только если похоже на "1/да"
    return 1 if s.startswith("д") or s == "активен" else 0


def _parse_units(raw_units: Any) -> List[str]:
    """Парсим единицы измерения из Excel в список: ["кг","шт"] (или другие)."""
    raw = "" if raw_units is None else str(raw_units).strip()
    if not raw:
        return []

    parts = [u.strip().lower() for u in raw.replace(";", ",").split(",") if u.strip()]
    norm_map = {
        "kg": "кг",
        "kgs": "кг",
        "кг": "кг",
        "килограмм": "кг",
        "килограммы": "кг",
        "pcs": "шт",
        "pc": "шт",
        "piece": "шт",
        "pieces": "шт",
        "шт": "шт",
        "штука": "шт",
        "штуки": "шт",
    }

    out: List[str] = []
    for p in parts:
        out.append(norm_map.get(p, p))
    # убираем дубли, сохраняя порядок
    uniq: List[str] = []
    for u in out:
        if u not in uniq:
            uniq.append(u)
    return uniq


def _header_index(header_row: List[Any]) -> Dict[str, int]:
    """Собираем словарь {нормализованный_заголовок: индекс}."""
    idx: Dict[str, int] = {}
    for i, cell in enumerate(header_row):
        key = _norm_header(cell)
        if key and key not in idx:
            idx[key] = i
    return idx


def load_products(excel_path: str, sheet_name: str = "продукт") -> List[Dict[str, Any]]:
    wb = load_workbook(excel_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        # Для совместимости: если у кого-то лист называется "products"
        if "products" in wb.sheetnames:
            sheet_name = "products"
        else:
            raise KeyError(f"Worksheet {sheet_name} does not exist.")

    ws = wb[sheet_name]

    # читаем все строки (values_only)
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header = list(rows[0])
    h = _header_index(header)

    # ищем индексы колонок по русским названиям
    idx_code = h.get("код")
    idx_name = h.get("наименование")
    idx_life = h.get("срок годности (ч)") or h.get("срок годности(ч)") or h.get("срок годности")
    idx_units = h.get("ед. измер.") or h.get("ед.измер.") or h.get("ед измер") or h.get("ед. измер")
    idx_active = h.get("активен")
    idx_comment = h.get("комментарий")

    products: List[Dict[str, Any]] = []
    for r in rows[1:]:
        r = list(r)
        code = _get(r, idx_code, "")
        name = str(_get(r, idx_name, "")).strip()
        if not name:
            continue  # пропускаем пустые строки

        # срок годности (ч)
        raw_life = _get(r, idx_life, 0)
        try:
            shelf_life_hours = int(float(str(raw_life).replace(",", ".")))
        except Exception:
            shelf_life_hours = 0

        allowed_units = _parse_units(_get(r, idx_units, ""))

        active = _parse_yes_no(_get(r, idx_active, 1))

        comment = str(_get(r, idx_comment, "")).strip()

        products.append(
            {
                "product_id": str(code).strip() if code is not None else "",
                "name": name,
                "shelf_life_hours": shelf_life_hours,
                "allowed_units": allowed_units,  # список!
                "active": active,
                "comment": comment,
            }
        )

    return products


def load_staff(excel_path: str, sheet_name: str) -> List[Dict[str, Any]]:
    """
    Загружает сотрудников из листа (например "изготовил" или "проверил").
    Возвращает список словарей: {"name": "Иванов И.И.", "active": 1}
    """
    wb = load_workbook(excel_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise KeyError(f"Worksheet {sheet_name} does not exist.")
    ws = wb[sheet_name]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header = list(rows[0])
    h = _header_index(header)

    idx_name = h.get("фио")
    idx_active = h.get("активен")

    staff: List[Dict[str, Any]] = []
    for r in rows[1:]:
        r = list(r)
        name = str(_get(r, idx_name, "")).strip()
        if not name:
            continue
        active = _parse_yes_no(_get(r, idx_active, 1))
        staff.append({"name": name, "active": active})

    return staff
