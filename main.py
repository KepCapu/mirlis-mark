# main.py
# Mirlis Mark — Система маркировки
# UI: "как на картинке" + редактор предпросмотра + ВИДИМЫЕ стрелочки в выпадающих списках
#
# ВАЖНО:
# - excel_loader.py / label_logic.py / printer.py НЕ ТРОГАЕМ
# - логотип берём по пути: D:\mirlis_mark\Mirlis software logo.png
#
# === PYQT5-СОВМЕСТИМАЯ ВЕРСИЯ ===
# Изменения относительно PyQt6:
# - Все импорты из PyQt5
# - Enum-стиль без вложенных флагов (Qt.AlignCenter вместо Qt.AlignmentFlag.AlignCenter)
# - QMediaPlayer / QVideoWidget из PyQt5.QtMultimedia / QtMultimediaWidgets
# - InsertPolicy, ScrollBarPolicy и т.п. — без вложенных enum-классов


import sys
import os
import shutil
import time
import math
import json
import calendar
from datetime import datetime
import ctypes
from ctypes import wintypes

from statistics_page import StatisticsPage
from resources import resource_path as _resource_path

from PyQt5.QtWidgets import (
    QApplication,
    QWidget,
    QVBoxLayout,
    QLabel,
    QComboBox,
    QLineEdit,
    QPushButton,
    QHBoxLayout,
    QCheckBox,
    QMessageBox,
    QCompleter,
    QFrame,
    QScrollArea,
    QSizePolicy,
    QTextEdit,
    QToolButton,
    QFontComboBox,
    QSpacerItem,
    QFileDialog,
    QDateTimeEdit,
    QDateEdit,
    QTimeEdit,
    QCalendarWidget,
    QGridLayout,
    QDialog,
    QListWidget,
    QListWidgetItem,
    QListView,
    QScrollBar,
    QScroller,
    QScrollerProperties,
    QStyledItemDelegate,
    QStackedWidget,
    QGraphicsView,
    QGraphicsScene,
    QGraphicsProxyWidget,
    QGraphicsDropShadowEffect,
)
from PyQt5.QtCore import QTimer, Qt, QUrl, QSize, QDateTime, QDate, QTime, pyqtSignal, QPoint, QLocale, QEvent, QSizeF, QRectF, QEasingCurve
from PyQt5.QtCore import QObject
from PyQt5.QtCore import QStringListModel
from PyQt5.QtGui import (
    QDesktopServices,
    QIcon,
    QPixmap,
    QFont,
    QIntValidator,
    QTextCharFormat,
    QTextBlockFormat,
    QTextCursor,
    QTextDocument,
    QSurfaceFormat,
    QImage,
    QPainter,
    QColor,
    QBrush,
    QPen,
    QRegion,
)

# PyQt5: мультимедиа
from PyQt5.QtMultimedia import QMediaPlayer, QMediaContent
from PyQt5.QtMultimediaWidgets import QVideoWidget

from excel_loader import load_products, load_staff
from label_logic import build_label, format_dt
from printer import print_text_as_bitmap_tspl, print_raw
from scale_reader import ScaleReader, SERIAL_AVAILABLE
from openpyxl import load_workbook as _peek_workbook
import win32print
from stats_store import append_entry as _append_stats_entry

# --- Компенсация preview/печати под high-DPI Windows (эталон ≈ 96 logical DPI, DPR 1). Подкрутка вручную: ---
SCREEN_COMP_REF_LOGICAL_DPI = 96.0
SCREEN_COMP_LDPI_BLEND = 0.55
SCREEN_COMP_MIN = 0.78
SCREEN_COMP_MANUAL_MULTIPLIER = 1.0
SCREEN_COMP_DPR_EXTRA = 0.12


def _is_foreground_window_tabtip():
    """
    Возвращает True, если активное окно Windows — сенсорная клавиатура
    (TabTip / TouchKeyboard / IME-pad). Используется чтобы игнорировать
    кратковременную деактивацию нашего окна при печати на OSK.
    """
    try:
        user32 = ctypes.windll.user32
        hwnd = user32.GetForegroundWindow()
        if not hwnd:
            return False
        # Имя класса окна
        buf = ctypes.create_unicode_buffer(256)
        user32.GetClassNameW(hwnd, buf, 256)
        cls = buf.value or ""
        # Известные классы окон сенсорной клавиатуры Windows
        known = (
            "IPTip_Main_Window",        # TabTip (Windows 10)
            "Windows.UI.Core.CoreWindow",  # Touch Keyboard (Windows 11)
            "ApplicationFrameWindow",   # обёртка некоторых UWP touch input
            "InputApp",                 # Windows 11 IME
        )
        return any(k in cls for k in known)
    except Exception:
        return False


# -------------------- ПУТИ: ресурсы и пользовательские данные --------------------
def _setup_auto_repeat(btn, callback, delay=400, interval=80):
    """Auto-repeat через мышиные события + debounce (устойчив к тач-экранам)."""
    btn.setAutoRepeat(False)
    btn.setFocusPolicy(Qt.NoFocus)
    btn.setContextMenuPolicy(Qt.NoContextMenu)

    btn._ar_repeat = QTimer(btn)
    btn._ar_repeat.setInterval(interval)
    btn._ar_repeat.timeout.connect(callback)

    btn._ar_delay = QTimer(btn)
    btn._ar_delay.setSingleShot(True)
    btn._ar_delay.setInterval(delay)
    btn._ar_delay.timeout.connect(btn._ar_repeat.start)

    # Debounce: при отпускании ждём 200мс, прежде чем реально остановить.
    # Если за это время придёт новый press (тач-экран часто так делает) —
    # отменяем остановку и продолжаем.
    btn._ar_stop_debounce = QTimer(btn)
    btn._ar_stop_debounce.setSingleShot(True)
    btn._ar_stop_debounce.setInterval(200)

    def _real_stop():
        btn._ar_delay.stop()
        btn._ar_repeat.stop()

    btn._ar_stop_debounce.timeout.connect(_real_stop)

    def _start():
        btn._ar_stop_debounce.stop()  # отменяем отложенную остановку
        btn._ar_delay.stop()
        btn._ar_repeat.stop()
        callback()
        btn._ar_delay.start()

    def _request_stop():
        btn._ar_stop_debounce.start()

    btn.pressed.connect(_start)
    btn.released.connect(_request_stop)


def resource_path(relative_path: str) -> str:
    """Путь к встроенному ресурсу: из исходников — от корня проекта, из exe — из sys._MEIPASS."""
    return _resource_path(relative_path)


def app_data_dir() -> str:
    """Папка приложения в профиле пользователя Windows (%LOCALAPPDATA%\\MirlisMark). Создаётся при первом вызове."""
    if sys.platform == "win32":
        root = os.environ.get("LOCALAPPDATA", os.path.expanduser("~"))
    else:
        root = os.path.expanduser("~")
    path = os.path.join(root, "MirlisMark")
    os.makedirs(path, exist_ok=True)
    return path


def get_config_path() -> str:
    """Путь к settings.json в пользовательской папке приложения."""
    return os.path.join(app_data_dir(), "settings.json")


def ensure_products_file() -> str:
    """
    Путь к products.xlsx для работы приложения.
    Файл лежит в пользовательской папке; при отсутствии копируется шаблон из ресурсов сборки.
    При отсутствии шаблона или ошибке копирования выбрасывает RuntimeError с понятным текстом.
    """
    app_dir = app_data_dir()
    target = os.path.join(app_dir, "products.xlsx")
    if os.path.isfile(target):
        return target
    template = resource_path("data_sources/products.xlsx")
    if not os.path.isfile(template):
        raise RuntimeError(
            "Встроенный шаблон products.xlsx не найден. "
            "Убедитесь, что при сборке добавлена папка data_sources (--add-data \"data_sources;data_sources\")."
        )
    try:
        shutil.copy2(template, target)
    except Exception as e:
        raise RuntimeError(
            f"Не удалось создать файл products.xlsx в папке приложения:\n{app_dir}\nОшибка: {e}"
        ) from e
    return target


# Константы листов Excel (не пути)
SHEET_PRODUCTS = "продукт"
SHEET_MADE = "изготовил"
SHEET_CHECKED = "цех"

# Ресурсы интерфейса (через resource_path для работы из исходников и из exe)
LOGO_PATH = resource_path("assets/logo.png")
SPLASH_VIDEO_PATH = resource_path("assets/loadingscreen.mp4")
HELP_IMG_PRODUCT = resource_path("assets/help_sheet_product.png")
HELP_IMG_MADE = resource_path("assets/help_sheet_made.png")
HELP_IMG_WORKSHOP = resource_path("assets/help_sheet_workshop.png")
HELP_IMG_TABS = resource_path("assets/help_sheet_tabs.png")
HAPPY_HERO_PATH = resource_path("assets/happy_hero.png")

APP_TITLE = "Mirlis Mark — Система маркировки"
APP_MARK = "Mark"
APP_VERSION = "1.0"
APP_SUBTITLE = "Система маркировки"

# фон предпросмотра в ручном режиме: мягкий светло-жёлтый (чуть насыщеннее бледного)
MANUAL_PREVIEW_BG = "#fef3c7"


# -------------------- HELPERS --------------------
def _load_settings() -> dict:
    """Загрузить настройки из settings.json (в папке приложения пользователя)."""
    try:
        path = get_config_path()
        if os.path.isfile(path):
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
    except (OSError, json.JSONDecodeError) as e:
        sys.stderr.write(f"[MirlisMark] Не удалось загрузить настройки: {e}\n")
    except Exception as e:
        sys.stderr.write(f"[MirlisMark] Ошибка при чтении settings.json: {e}\n")
    return {}


def _save_settings(data: dict):
    """Сохранить настройки в settings.json (в папке приложения пользователя)."""
    try:
        with open(get_config_path(), "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except OSError as e:
        sys.stderr.write(f"[MirlisMark] Не удалось сохранить настройки: {e}\n")
    except Exception as e:
        sys.stderr.write(f"[MirlisMark] Ошибка при записи settings.json: {e}\n")


def _fmt_dt_local(ts: float) -> str:
    return datetime.fromtimestamp(ts).strftime("%d.%m.%Y %H:%M:%S")


def _safe_int(v, default=0):
    try:
        return int(v)
    except Exception:
        return default


def _repair_mojibake_utf8_as_cp1251(s: str) -> str:
    """
    Best-effort repair for the typical case:
    UTF-8 bytes were mistakenly decoded as cp1251, producing strings like "РЁР°...".
    We fix it at the data-ingest stage (Excel -> app) so downstream storage remains clean.
    """
    try:
        src = "" if s is None else str(s)
    except Exception:
        return ""
    if not src:
        return src
    # Heuristic trigger: common mojibake sequences for Cyrillic in this project.
    if ("Р" not in src and "С" not in src) or src.count("Р") + src.count("С") < 2:
        return src
    try:
        repaired = src.encode("cp1251", errors="strict").decode("utf-8", errors="strict")
    except Exception:
        return src
    # Accept only if it looks like a real Cyrillic string and is different.
    if repaired and repaired != src:
        return repaired
    return src


# -------------------- UI Building Blocks --------------------
class Card(QFrame):
    """Скруглённая карточка с лёгкой рамкой."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setObjectName("Card")
        self.setFrameShape(QFrame.NoFrame)
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)


class Pill(QLabel):
    """Плашка-'pill' (серый бэйдж)."""

    def __init__(self, text="", parent=None):
        super().__init__(text, parent)
        self.setObjectName("Pill")
        self.setSizePolicy(QSizePolicy.Maximum, QSizePolicy.Fixed)


class HeaderLabel(QLabel):
    """Заголовок секции по центру с короткой тенью-плашкой."""

    def __init__(self, text="", parent=None):
        super().__init__(text, parent)
        self.setObjectName("SectionTitle")
        self.setAlignment(Qt.AlignCenter)


class PreviewHeaderLabel(HeaderLabel):
    """Заголовок «Предпросмотр»: двойной клик переключает ручной режим. Без hover/кнопки."""

    doubleClicked = pyqtSignal()

    def mouseDoubleClickEvent(self, event):
        if event.button() == Qt.LeftButton:
            self.doubleClicked.emit()
        super().mouseDoubleClickEvent(event)


class _StyledScrollBar(QScrollBar):
    """
    Кастомный QScrollBar с закруглённым жёлтым ползунком.
    Устанавливается на verticalScrollBar() popup QComboBox.
    """

    TRACK        = QColor("#cbd5e1")
    HANDLE       = QColor("#f9b233")
    HANDLE_HOVER = QColor("#e6a020")

    WIDGET_W = 52
    TRACK_W  = 42
    HANDLE_W = 32
    HANDLE_H = 68
    R        = 12

    def __init__(self, parent=None):
        super().__init__(Qt.Vertical, parent)
        self.setFixedWidth(self.WIDGET_W)

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing, True)
        painter.setPen(Qt.NoPen)

        w = self.width()
        h = self.height()

        # Сначала заливаем весь виджет белым, чтобы перекрыть дефолтный серый фон Qt по бокам трека
        painter.fillRect(self.rect(), QColor("#ffffff"))

        # трек — тонкий контур, без заливки
        tx = (w - self.TRACK_W) // 2
        painter.setBrush(Qt.NoBrush)
        painter.setPen(QPen(self.TRACK, 1))
        painter.drawRoundedRect(tx, 0, self.TRACK_W, h, self.R, self.R)
        painter.setPen(Qt.NoPen)     # возвращаем NoPen, чтобы ползунок ниже рисовался без обводки

        # ползунок — фиксированная высота, скользит по треку
        total = self.maximum() - self.minimum()
        if total <= 0:
            return
        travel = max(1, h - self.HANDLE_H)
        y0 = int((self.value() - self.minimum()) / total * travel)
        y0 = max(0, min(y0, travel))
        hx = tx + (self.TRACK_W - self.HANDLE_W) // 2
        hovered = self.underMouse()
        painter.setBrush(self.HANDLE_HOVER if hovered else self.HANDLE)
        painter.drawRoundedRect(hx, y0, self.HANDLE_W, self.HANDLE_H, self.R, self.R)


class ComboBoxFixedArrow(QComboBox):
    """
    Комбобокс с гарантированно видимой стрелкой.
    """

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setObjectName("ComboWithArrow")
        QTimer.singleShot(0, self._wire_completer_explicit_close)

    def _wire_completer_explicit_close(self):
        try:
            comp = self.completer()
            if comp is not None:
                comp.activated.connect(lambda *_: setattr(self, "_explicit_close", True))
        except Exception:
            pass

    def showPopup(self):
        # Editable combo с completer: ВСЕГДА работаем через completer-popup,
        # никогда не показываем super().showPopup() (полный список).
        # Полный список допустим только если поле ввода пустое.
        try:
            if self.isEditable() and self.completer() is not None:
                le = self.lineEdit()
                text = le.text().strip() if le is not None else ""
                if text:
                    comp = self.completer()
                    cp = comp.popup()

                    # 1) Completer-popup сейчас видим — клик по стрелке закрывает.
                    if cp is not None and cp.isVisible():
                        self._explicit_close = True
                        cp.hide()
                        return

                    # 2) Completer-popup только что закрылся — toggle close: не открывать.
                    import time as _t
                    last_hide = getattr(self, "_completer_popup_last_hide_ts", 0)
                    now = _t.time()
                    if now - last_hide < 0.3:
                        self._completer_popup_last_hide_ts = 0
                        return

                    # 3) Открываем completer с актуальным префиксом.
                    # Жёстко переустанавливаем prefix и фильтрацию.
                    try:
                        comp.setCompletionPrefix(text)
                    except Exception:
                        pass
                    comp.complete()
                    return
                # Поле ПУСТОЕ → разрешаем стандартный показ полного списка ниже.
        except Exception:
            pass

        super().showPopup()
        view = self.view()
        if view:
            # Включаем тач-скролл для popup
            scroller = QScroller.scroller(view.viewport())
            if scroller:
                props = scroller.scrollerProperties()
                props.setScrollMetric(QScrollerProperties.ScrollingCurve, QEasingCurve.OutQuad)
                props.setScrollMetric(QScrollerProperties.AxisLockThreshold, 0.0)
                scroller.setScrollerProperties(props)
            QScroller.grabGesture(view.viewport(), QScroller.LeftMouseButtonGesture)
            view.viewport().setAttribute(Qt.WA_AcceptTouchEvents, True)

        # Важно: кастомный скроллбар и маска — только для tablet
        if getattr(self, "_is_tablet_mode", False):
            lv = self.view()
            if lv is not None:
                try:
                    sb = _StyledScrollBar(lv)
                    lv.setVerticalScrollBar(sb)
                except Exception:
                    pass
                try:
                    popup = lv.window()
                    if popup is not None:
                        rect = popup.rect()
                        rounded = QRegion(rect, QRegion.Rectangle)
                        popup.setMask(rounded)
                except Exception:
                    pass


class ComboBoxPopupDown(ComboBoxFixedArrow):
    """
    Комбобокс, у которого выпадающий список всегда раскрывается вниз.
    """

    def showPopup(self):
        super().showPopup()
        view = self.view()
        if view and view.window():
            popup = view.window()
            bottom_left = self.mapToGlobal(self.rect().bottomLeft())
            popup.move(bottom_left.x(), bottom_left.y())


class _PopupUpKeeper(QObject):
    """Держит popup комбобокса НАД его полем (раскрытие вверх).
    Перехватывает Show/Resize/Move контейнера: как только Qt применяет
    финальную высоту (Resize) или пытается поставить popup вниз (Move/Show),
    keeper немедленно возвращает его наверх. Это убирает «прыжок вниз→вверх»,
    т.к. промежуточная нижняя позиция не успевает отрисоваться."""

    def __init__(self, combo, popup):
        super().__init__(popup)
        self._combo = combo
        self._popup = popup
        self._busy = False

    def _reposition(self):
        if self._busy:
            return
        self._busy = True
        try:
            ph = self._popup.height()
            tl = self._combo.mapToGlobal(self._combo.rect().topLeft())
            target_x = tl.x()
            target_y = tl.y() - ph
            if self._popup.x() != target_x or self._popup.y() != target_y:
                self._popup.move(target_x, target_y)
        except Exception:
            pass
        finally:
            self._busy = False

    def eventFilter(self, obj, event):
        if obj is self._popup and event.type() in (
            QEvent.Show, QEvent.Resize, QEvent.Move
        ):
            self._reposition()
        return False


class ComboBoxPopupUp(ComboBoxFixedArrow):
    """
    Комбобокс, у которого выпадающий список всегда раскрывается вверх.
    Используется для «Цех» — расположен ближе к низу левой колонки.
    """

    def showPopup(self):
        # КЛЮЧЕВОЕ: вешаем keeper ДО super().showPopup().
        # self.view() форсирует создание контейнера popup ещё до показа,
        # поэтому keeper успевает перехватить самые первые Show/Move события,
        # которые иначе (при установке после super) проходят мимо на ПЕРВОМ
        # открытии — отсюда был баг «первый раз вниз, потом вверх».
        view = self.view()
        if view and view.window():
            popup = view.window()
            keeper = getattr(self, "_popup_up_keeper", None)
            if keeper is None or keeper._popup is not popup:
                keeper = _PopupUpKeeper(self, popup)
                popup.installEventFilter(keeper)
                self._popup_up_keeper = keeper

        super().showPopup()

        # Повторный перенос после показа — на случай если финальная высота
        # применилась в самом super (keeper уже её поймает, но дублируем для надёжности).
        v2 = self.view()
        if v2 and v2.window():
            keeper = getattr(self, "_popup_up_keeper", None)
            if keeper is not None:
                keeper._reposition()


class PlainFontNameDelegate(QStyledItemDelegate):
    """Рендерит пункты как обычный текст (без предпросмотра шрифтом)."""

    def __init__(self, parent=None, min_h: int | None = None):
        super().__init__(parent)
        self._min_h = int(min_h) if min_h is not None else None

    def sizeHint(self, option, index):
        s = super().sizeHint(option, index)
        if self._min_h is None:
            return s
        return QSize(s.width(), max(int(s.height()), self._min_h))

    def paint(self, painter, option, index):
        opt = option
        opt.font = QFont("Segoe UI", opt.font.pointSize())
        super().paint(painter, opt, index)


class ToolBtn(QToolButton):
    def __init__(self, text="", parent=None):
        super().__init__(parent)
        self.setText(text)
        self.setCursor(Qt.PointingHandCursor)
        self.setObjectName("ToolBtn")
        self.setCheckable(True)
        self.setAutoRaise(False)


class ActionBtn(QPushButton):
    def __init__(self, text="", kind="default", parent=None):
        super().__init__(text, parent)
        self.setCursor(Qt.PointingHandCursor)
        self.setObjectName(f"Btn_{kind}")


# -------------------- SPLASH VIDEO --------------------
class SplashVideo(QWidget):
    """Окно загрузки с воспроизведением видео перед запуском основного приложения."""

    def __init__(self, video_path: str, on_finished_callback=None, parent=None):
        super().__init__(parent)
        self._on_finished = on_finished_callback
        self._finished_called = False
        self.setWindowFlags(Qt.FramelessWindowHint)
        self.setStyleSheet("background-color: #000000;")
        self.setAttribute(Qt.WA_TranslucentBackground, False)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)

        self._video_widget = QVideoWidget()
        self._video_widget.setStyleSheet("background-color: #000000;")
        self._video_widget.setAspectRatioMode(Qt.KeepAspectRatioByExpanding)
        layout.addWidget(self._video_widget)

        self._player = QMediaPlayer()
        self._player.setVideoOutput(self._video_widget)
        self._player.setMedia(QMediaContent(QUrl.fromLocalFile(video_path)))
        self._player.mediaStatusChanged.connect(self._on_media_status_changed)
        self._player.error.connect(self._on_error)

        # Таймаут-подстраховка: если видео не завершилось за 15 сек — запускаем приложение
        self._timeout = QTimer(self)
        self._timeout.setSingleShot(True)
        self._timeout.timeout.connect(self._finish)
        self._timeout.start(15000)

    def _finish(self):
        """Единоразовый запуск основного окна + закрытие splash."""
        if self._finished_called:
            return
        self._finished_called = True
        self._timeout.stop()
        self._player.stop()
        if callable(self._on_finished):
            self._on_finished()
        self.close()

    def _on_media_status_changed(self, status):
        if status == QMediaPlayer.EndOfMedia:
            self._finish()
        elif status == QMediaPlayer.InvalidMedia:
            self._finish()

    def _on_error(self):
        self._finish()

    def play(self):
        self._player.play()

    def showEvent(self, event):
        super().showEvent(event)
        self._player.play()


# -------------------- CUSTOM DATE-TIME PICKER --------------------
_MONTH_NAMES_RU = [
    "", "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
    "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь",
]
_MONTH_NAMES_RU_GENITIVE = [
    "", "января", "февраля", "марта", "апреля", "мая", "июня",
    "июля", "августа", "сентября", "октября", "ноября", "декабря",
]
_WEEKDAY_HEADERS = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]


class CustomDateTimePicker(QWidget):
    """Кастомный picker даты и времени с popup-панелью (календарь + время)."""

    dateTimeChanged = pyqtSignal()

    def __init__(self, parent=None):
        super().__init__(parent)
        # Ensure eventFilter can safely run during early init.
        self._popup = None
        self._date = QDate.currentDate()
        self._time = QTime.currentTime()
        self._cal_view = "day"  # day | month | year
        self._popup_hidden_at = 0.0
        self._closing_by_trigger = False
        self._trigger_pressed = False

        # --- триггер как у QComboBox ---
        combo_btn_path = resource_path("assets/combo-btn.svg").replace("\\", "/")
        self._trigger_frame = QFrame()
        self._trigger_frame.setObjectName("DateTimeTrigger")
        self._trigger_frame.setMinimumHeight(40)
        self._trigger_frame.setCursor(Qt.PointingHandCursor)
        self._trigger_frame.setStyleSheet(
            "#DateTimeTrigger { background: #ffffff; border: 1px solid #cfd6e0; border-radius: 12px; }"
            "#DateTimeTrigger:hover { border-color: #94a3b8; }"
        )
        self._trigger_frame.installEventFilter(self)
        inner = QHBoxLayout(self._trigger_frame)
        inner.setContentsMargins(14, 0, 4, 0)
        inner.setSpacing(0)
        self._text_label = QLabel()
        self._text_label.setStyleSheet(
            "background: transparent; border: none; font-size: 14px; color: #111827;"
        )
        self._text_label.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        inner.addWidget(self._text_label, 1)
        self._drop_icon = QLabel()
        self._drop_icon.setFixedSize(36, 36)
        self._drop_icon.setStyleSheet("background: transparent; border: none;")
        try:
            icon = QIcon(combo_btn_path)
            if not icon.isNull():
                self._drop_icon.setPixmap(icon.pixmap(QSize(36, 36)))
        except Exception:
            pass
        inner.addWidget(self._drop_icon, 0, Qt.AlignRight | Qt.AlignVCenter)
        self._text_label.installEventFilter(self)
        self._drop_icon.installEventFilter(self)

        lay = QHBoxLayout(self)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.addWidget(self._trigger_frame)

        # --- popup ---
        self._popup = QFrame(self, Qt.Popup)
        # Needed to detect Qt.Popup auto-hide and prevent immediate reopen on the same click.
        self._popup.installEventFilter(self)
        self._popup.setStyleSheet(
            "QFrame { background: #ffffff; border: 1px solid #e2e8f0; border-radius: 16px; }"
        )
        self._popup.setFixedWidth(625)
        self._build_popup()
        self._update_btn_text()

    # ---------- btn text ----------
    def _update_btn_text(self):
        d = self._date
        month_gen = _MONTH_NAMES_RU_GENITIVE[d.month()]
        txt = f"  📅   {d.day()} {month_gen} {d.year()}  {self._time.toString('HH:mm')}"
        self._text_label.setText(txt)

    # ---------- popup build ----------
    def _build_popup(self):
        popup_lay = QVBoxLayout(self._popup)
        popup_lay.setContentsMargins(16, 10, 16, 10)
        popup_lay.setSpacing(8)

        # --- навигация месяца ---
        nav = QHBoxLayout()
        nav.setSpacing(0)
        self._prev_btn = QPushButton("‹")
        self._prev_btn.setFixedSize(36, 36)
        self._prev_btn.setCursor(Qt.PointingHandCursor)
        self._prev_btn.setStyleSheet(
            "QPushButton { background: transparent; border: none; font-size: 22px; font-weight: 700; color: #64748b; }"
            "QPushButton:hover { color: #111827; }"
        )
        self._prev_btn.clicked.connect(lambda: self._change_month(-1))

        self._month_btn = QPushButton()
        self._month_btn.setCursor(Qt.PointingHandCursor)
        self._month_btn.setFlat(True)
        self._month_btn.setStyleSheet(
            "QPushButton { background: transparent; border: none; font-size: 16px; font-weight: 700; color: #111827; }"
            "QPushButton:hover { color: #0f172a; }"
        )
        self._month_btn.clicked.connect(self._on_month_label_clicked)

        self._next_btn = QPushButton("›")
        self._next_btn.setFixedSize(36, 36)
        self._next_btn.setCursor(Qt.PointingHandCursor)
        self._next_btn.setStyleSheet(
            "QPushButton { background: transparent; border: none; font-size: 22px; font-weight: 700; color: #64748b; }"
            "QPushButton:hover { color: #111827; }"
        )
        self._next_btn.clicked.connect(lambda: self._change_month(1))

        nav.addWidget(self._prev_btn)
        nav.addWidget(self._month_btn, 1)
        nav.addWidget(self._next_btn)
        popup_lay.addLayout(nav)

        # --- разделитель ---
        sep = QFrame()
        sep.setFrameShape(QFrame.HLine)
        sep.setStyleSheet("color: #e2e8f0; background: #e2e8f0; border: none; max-height: 1px;")
        popup_lay.addWidget(sep)

        # --- календарь: 3 режима (day/month/year) ---
        self._cal_container = QWidget()
        cal_wrap = QVBoxLayout(self._cal_container)
        cal_wrap.setContentsMargins(0, 0, 0, 0)
        cal_wrap.setSpacing(0)

        # day page
        self._day_page = QWidget()
        self._cal_grid = QGridLayout(self._day_page)
        self._cal_grid.setSpacing(2)
        self._cal_grid.setContentsMargins(0, 0, 0, 0)
        for c in range(7):
            self._cal_grid.setColumnMinimumWidth(c, 56)
        for col, name in enumerate(_WEEKDAY_HEADERS):
            lbl = QLabel(name)
            lbl.setAlignment(Qt.AlignCenter)
            color = "#ef6c00" if col >= 5 else "#9ca3af"
            lbl.setStyleSheet(
                f"font-size: 12px; font-weight: 600; color: {color}; "
                "background: transparent; padding: 4px 0;"
            )
            self._cal_grid.addWidget(lbl, 0, col)

        # month page
        self._month_page = QWidget()
        self._month_grid = QGridLayout(self._month_page)
        self._month_grid.setSpacing(6)
        self._month_grid.setContentsMargins(0, 6, 0, 0)

        # year page
        self._year_page = QWidget()
        self._year_grid = QGridLayout(self._year_page)
        self._year_grid.setSpacing(6)
        self._year_grid.setContentsMargins(0, 6, 0, 0)

        cal_wrap.addWidget(self._day_page)
        cal_wrap.addWidget(self._month_page)
        cal_wrap.addWidget(self._year_page)
        popup_lay.addWidget(self._cal_container)

        # --- разделитель ---
        sep2 = QFrame()
        sep2.setFrameShape(QFrame.HLine)
        sep2.setStyleSheet("color: #e2e8f0; background: #e2e8f0; border: none; max-height: 1px;")
        popup_lay.addWidget(sep2)

        # --- блок времени ---
        time_header = QLabel("Время")
        time_header.setAlignment(Qt.AlignCenter)
        time_header.setStyleSheet(
            "font-size: 21px; font-weight: 700; color: #111827; background: transparent; padding: 2px 0;"
        )
        time_header_row = QHBoxLayout()
        time_header_row.setContentsMargins(0, 0, 0, 0)
        time_header_row.addStretch(1)
        time_header_row.addWidget(time_header)
        time_header_row.addStretch(1)
        popup_lay.addLayout(time_header_row)

        arrow_style = (
            "QPushButton { background: #f9b233; border: none; border-radius: 10px; "
            "color: #ffffff; font-size: 16px; font-weight: 800; min-width: 34px; min-height: 28px; }"
            "QPushButton:hover { background: #e5a020; }"
            "QPushButton:pressed { background: #cc8c10; }"
        )
        btn_h_up = QPushButton("▲")
        btn_h_up.setStyleSheet(arrow_style)
        btn_h_up.setCursor(Qt.PointingHandCursor)
        btn_h_up.setFocusPolicy(Qt.NoFocus)
        _setup_auto_repeat(btn_h_up, lambda: self._step_time("h", 1))

        btn_h_down = QPushButton("▼")
        btn_h_down.setStyleSheet(arrow_style)
        btn_h_down.setCursor(Qt.PointingHandCursor)
        btn_h_down.setFocusPolicy(Qt.NoFocus)
        _setup_auto_repeat(btn_h_down, lambda: self._step_time("h", -1))

        btn_m_up = QPushButton("▲")
        btn_m_up.setStyleSheet(arrow_style)
        btn_m_up.setCursor(Qt.PointingHandCursor)
        btn_m_up.setFocusPolicy(Qt.NoFocus)
        _setup_auto_repeat(btn_m_up, lambda: self._step_time("m", 1))

        btn_m_down = QPushButton("▼")
        btn_m_down.setStyleSheet(arrow_style)
        btn_m_down.setCursor(Qt.PointingHandCursor)
        btn_m_down.setFocusPolicy(Qt.NoFocus)
        _setup_auto_repeat(btn_m_down, lambda: self._step_time("m", -1))

        # Время: оставляем тот же внешний вид, но добавляем ручной ввод (клавиатура).
        time_input_style = (
            "font-size: 28px; font-weight: 700; color: #111827; "
            "background: #f8fafc; border: 1px solid #e2e8f0; border-radius: 12px; "
            "padding: 0px;"
        )
        self._hour_label = QLineEdit()
        self._hour_label.setAlignment(Qt.AlignCenter)
        self._hour_label.setFixedSize(72, 40)
        self._hour_label.setMaxLength(2)
        self._hour_label.setValidator(QIntValidator(0, 23, self._hour_label))
        self._hour_label.setStyleSheet(time_input_style)
        # Мгновенное обновление предпросмотра при валидном вводе (без автоформатирования во время набора).
        self._hour_label.textEdited.connect(lambda txt: self._on_time_text_live("h", txt))
        self._hour_label.editingFinished.connect(lambda: self._apply_time_input("h"))

        self._min_label = QLineEdit()
        self._min_label.setAlignment(Qt.AlignCenter)
        self._min_label.setFixedSize(72, 40)
        self._min_label.setMaxLength(2)
        self._min_label.setValidator(QIntValidator(0, 59, self._min_label))
        self._min_label.setStyleSheet(time_input_style)
        self._min_label.textEdited.connect(lambda txt: self._on_time_text_live("m", txt))
        self._min_label.editingFinished.connect(lambda: self._apply_time_input("m"))

        colon = QLabel(":")
        colon.setAlignment(Qt.AlignCenter)
        colon.setFixedSize(24, 40)
        colon.setStyleSheet("font-size: 28px; font-weight: 700; color: #111827; background: transparent;")

        time_grid = QGridLayout()
        time_grid.setSpacing(4)
        time_grid.setContentsMargins(0, 2, 0, 0)
        time_grid.addWidget(self._hour_label, 0, 0, 1, 2, Qt.AlignCenter)
        time_grid.addWidget(colon, 0, 2, Qt.AlignCenter)
        time_grid.addWidget(self._min_label, 0, 3, 1, 2, Qt.AlignCenter)
        time_grid.addWidget(btn_h_up, 1, 0, Qt.AlignHCenter)
        time_grid.addWidget(btn_h_down, 1, 1, Qt.AlignHCenter)
        time_grid.addWidget(btn_m_up, 1, 3, Qt.AlignHCenter)
        time_grid.addWidget(btn_m_down, 1, 4, Qt.AlignHCenter)
        time_grid.setColumnMinimumWidth(0, 40)
        time_grid.setColumnMinimumWidth(1, 40)
        time_grid.setColumnMinimumWidth(2, 24)
        time_grid.setColumnMinimumWidth(3, 40)
        time_grid.setColumnMinimumWidth(4, 40)
        popup_lay.addLayout(time_grid)

        self._refresh_calendar()
        self._refresh_time_labels()

    # ---------- calendar grid ----------
    def _on_month_label_clicked(self):
        # day -> month -> year (циклический переход по клику на заголовок)
        if getattr(self, "_cal_view", "day") == "day":
            self._cal_view = "month"
            self._day_page.setVisible(False)
            self._month_page.setVisible(True)
            self._year_page.setVisible(False)
            self._refresh_month_grid()
            return

        if self._cal_view == "month":
            self._cal_view = "year"
            self._day_page.setVisible(False)
            self._month_page.setVisible(False)
            self._year_page.setVisible(True)
            self._refresh_year_grid()
            return

        # если уже year — остаёмся в year (без неожиданного сброса)
        self._cal_view = "year"
        self._day_page.setVisible(False)
        self._month_page.setVisible(False)
        self._year_page.setVisible(True)
        self._refresh_year_grid()

    def _refresh_month_grid(self):
        # очистка
        while self._month_grid.count():
            it = self._month_grid.takeAt(0)
            w = it.widget()
            if w:
                w.deleteLater()

        # 12 месяцев (3x4)
        current_m = self._date.month()
        for idx in range(1, 13):
            r = (idx - 1) // 4
            c = (idx - 1) % 4
            btn = QPushButton(_MONTH_NAMES_RU[idx])
            btn.setCursor(Qt.PointingHandCursor)
            btn.setFixedSize(132, 38)
            if idx == current_m:
                btn.setStyleSheet(
                    "QPushButton { background: #f9b233; color: #ffffff; border: none; border-radius: 10px; "
                    "font-size: 14px; font-weight: 700; }"
                    "QPushButton:hover { background: #e5a020; }"
                )
            else:
                btn.setStyleSheet(
                    "QPushButton { background: transparent; border: 1px solid #e2e8f0; border-radius: 10px; "
                    "font-size: 14px; color: #374151; }"
                    "QPushButton:hover { background: #f1f5f9; }"
                )
            btn.clicked.connect(lambda checked, m=idx: self._select_month(m))
            self._month_grid.addWidget(btn, r, c)

    def _select_month(self, month: int):
        y = self._date.year()
        max_day = calendar.monthrange(y, month)[1]
        d = min(self._date.day(), max_day)
        self._date = QDate(y, month, d)
        self._cal_view = "day"
        self._refresh_calendar()
        self._update_btn_text()
        self.dateTimeChanged.emit()

    def _refresh_year_grid(self):
        while self._year_grid.count():
            it = self._year_grid.takeAt(0)
            w = it.widget()
            if w:
                w.deleteLater()

        cur_y = self._date.year()
        start_y = cur_y - 7
        years = [start_y + i for i in range(16)]  # 4x4
        for i, y in enumerate(years):
            r = i // 4
            c = i % 4
            btn = QPushButton(str(y))
            btn.setCursor(Qt.PointingHandCursor)
            btn.setFixedSize(132, 38)
            if y == cur_y:
                btn.setStyleSheet(
                    "QPushButton { background: #f9b233; color: #ffffff; border: none; border-radius: 10px; "
                    "font-size: 14px; font-weight: 700; }"
                    "QPushButton:hover { background: #e5a020; }"
                )
            else:
                btn.setStyleSheet(
                    "QPushButton { background: transparent; border: 1px solid #e2e8f0; border-radius: 10px; "
                    "font-size: 14px; color: #374151; }"
                    "QPushButton:hover { background: #f1f5f9; }"
                )
            btn.clicked.connect(lambda checked, yy=y: self._select_year(yy))
            self._year_grid.addWidget(btn, r, c)

    def _select_year(self, year: int):
        m = self._date.month()
        max_day = calendar.monthrange(year, m)[1]
        d = min(self._date.day(), max_day)
        self._date = QDate(year, m, d)
        self._cal_view = "month"
        self._day_page.setVisible(False)
        self._month_page.setVisible(True)
        self._year_page.setVisible(False)
        self._refresh_month_grid()
        self._month_btn.setText(f"{_MONTH_NAMES_RU[self._date.month()]} {self._date.year()} ▾")
        self._update_btn_text()
        self.dateTimeChanged.emit()

    def _refresh_calendar(self):
        # всегда возвращаемся к выбору дней
        self._cal_view = "day"
        self._day_page.setVisible(True)
        self._month_page.setVisible(False)
        self._year_page.setVisible(False)

        for r in range(7, 0, -1):
            for c in range(7):
                item = self._cal_grid.itemAtPosition(r, c)
                if item and item.widget():
                    item.widget().deleteLater()

        self._month_btn.setText(f"{_MONTH_NAMES_RU[self._date.month()]} {self._date.year()} ▾")

        year = self._date.year()
        month = self._date.month()
        cal = calendar.Calendar(firstweekday=0)
        weeks = cal.monthdayscalendar(year, month)

        for row_idx, week in enumerate(weeks):
            for col_idx, day in enumerate(week):
                btn = QPushButton("" if day == 0 else str(day))
                btn.setFixedSize(56, 36)
                btn.setCursor(Qt.PointingHandCursor if day else Qt.ArrowCursor)

                if day == 0:
                    btn.setEnabled(False)
                    btn.setStyleSheet(
                        "QPushButton { background: transparent; border: none; }"
                    )
                elif day == self._date.day():
                    btn.setStyleSheet(
                        "QPushButton { background: #f9b233; color: #ffffff; border: none; "
                        "border-radius: 10px; font-size: 14px; font-weight: 700; }"
                        "QPushButton:hover { background: #e5a020; }"
                    )
                elif col_idx >= 5:
                    btn.setStyleSheet(
                        "QPushButton { background: transparent; border: none; border-radius: 10px; "
                        "font-size: 14px; color: #ef6c00; }"
                        "QPushButton:hover { background: #fff3e0; }"
                    )
                else:
                    btn.setStyleSheet(
                        "QPushButton { background: transparent; border: none; border-radius: 10px; "
                        "font-size: 14px; color: #374151; }"
                        "QPushButton:hover { background: #f1f5f9; }"
                    )

                if day > 0:
                    btn.clicked.connect(lambda checked, d=day: self._select_day(d))

                self._cal_grid.addWidget(btn, row_idx + 1, col_idx)

    def _change_month(self, delta):
        m = self._date.month() + delta
        y = self._date.year()
        if m < 1:
            m = 12
            y -= 1
        elif m > 12:
            m = 1
            y += 1
        max_day = calendar.monthrange(y, m)[1]
        d = min(self._date.day(), max_day)
        self._date = QDate(y, m, d)

        if getattr(self, "_cal_view", "day") == "day":
            self._refresh_calendar()
        elif self._cal_view == "month":
            self._day_page.setVisible(False)
            self._month_page.setVisible(True)
            self._year_page.setVisible(False)
            self._month_btn.setText(f"{_MONTH_NAMES_RU[self._date.month()]} {self._date.year()} ▾")
            self._refresh_month_grid()
        else:  # year
            self._day_page.setVisible(False)
            self._month_page.setVisible(False)
            self._year_page.setVisible(True)
            self._month_btn.setText(f"{_MONTH_NAMES_RU[self._date.month()]} {self._date.year()} ▾")
            self._refresh_year_grid()

        self._update_btn_text()
        self.dateTimeChanged.emit()

    def _select_day(self, day):
        self._date = QDate(self._date.year(), self._date.month(), day)
        self._refresh_calendar()
        self._update_btn_text()
        self.dateTimeChanged.emit()

    # ---------- time ----------
    def _refresh_time_labels(self):
        self._hour_label.setText(f"{self._time.hour():02d}")
        self._min_label.setText(f"{self._time.minute():02d}")

    def _apply_time_input(self, part: str):
        """
        Применяет ручной ввод часа/минуты:
        - диапазоны: 00–23 и 00–59
        - формат: всегда 2 цифры
        - при некорректном вводе: возвращаем последнее корректное значение
        """
        h = self._time.hour()
        m = self._time.minute()

        if part == "h":
            raw = self._hour_label.text().strip()
            try:
                v = int(raw) if raw != "" else None
            except Exception:
                v = None
            if v is None:
                self._refresh_time_labels()
                return
            v = max(0, min(23, v))
            h = v
        else:
            raw = self._min_label.text().strip()
            try:
                v = int(raw) if raw != "" else None
            except Exception:
                v = None
            if v is None:
                self._refresh_time_labels()
                return
            v = max(0, min(59, v))
            m = v

        self._time = QTime(h, m)
        self._refresh_time_labels()
        self._update_btn_text()
        self.dateTimeChanged.emit()

    def _on_time_text_live(self, part: str, txt: str):
        """
        Живое обновление предпросмотра при наборе:
        - обновляем итоговую дату/время сразу, как только ввод можно интерпретировать
        - НЕ форматируем поле (00/01/...) во время набора, чтобы не мешать сценарию '1' -> '10'
        """
        s = (txt or "").strip()
        if s == "":
            return
        # Разрешаем только цифры (валидатор обычно не пропускает другое, но на всякий случай).
        if not s.isdigit():
            return
        try:
            v = int(s)
        except Exception:
            return

        h = self._time.hour()
        m = self._time.minute()
        if part == "h":
            if 0 <= v <= 23:
                h = v
            else:
                return
        else:
            if 0 <= v <= 59:
                m = v
            else:
                return

        # Чтобы не спамить лишними сигналами — обновляем только при реальном изменении.
        if h == self._time.hour() and m == self._time.minute():
            return

        self._time = QTime(h, m)
        self._update_btn_text()
        self.dateTimeChanged.emit()

    def _step_time(self, part, delta):
        h = self._time.hour()
        m = self._time.minute()
        if part == "h":
            h = (h + delta) % 24
        else:
            m = (m + delta) % 60
        self._time = QTime(h, m)
        self._refresh_time_labels()
        self._update_btn_text()
        self.dateTimeChanged.emit()

    # ---------- event filter ----------
    def eventFilter(self, obj, event):
        pop = getattr(self, "_popup", None)
        if pop is None:
            return super().eventFilter(obj, event)

        # Qt.Popup может сам закрывать popup на клике "вне".
        # Фиксируем момент скрытия, чтобы не допустить переоткрытия тем же кликом.
        if obj is pop and event.type() == QEvent.Hide:
            # Если мы закрыли popup сами по клику на триггер — не включаем guard,
            # иначе следующий клик "сразу открыть" может быть ошибочно заблокирован.
            if getattr(self, "_closing_by_trigger", False):
                self._closing_by_trigger = False
            else:
                self._popup_hidden_at = time.monotonic()
            return super().eventFilter(obj, event)

        if event.type() == QEvent.MouseButtonPress and obj in (
            self._trigger_frame,
            self._text_label,
            self._drop_icon,
        ):
            # Строгий toggle как у остальных dropdown:
            # если popup открыт — повторный клик по триггеру всегда закрывает его.
            if pop.isVisible():
                self._closing_by_trigger = True
                pop.hide()
                return True

            # Если popup был скрыт совсем недавно (~250 мс назад),
            # значит этот же клик уже привёл к автозакрытию через Qt.Popup,
            # и НЕ нужно открывать popup снова.
            if (time.monotonic() - float(getattr(self, "_popup_hidden_at", 0.0))) < 0.25:
                return True

            # Важно: Qt.Popup может закрываться по MouseButtonRelease, если popup показан между press/release.
            # Поэтому помечаем нажатие, а открываем ПОСЛЕ release (ниже).
            self._trigger_pressed = True
            return True

        if event.type() == QEvent.MouseButtonRelease and obj in (
            self._trigger_frame,
            self._text_label,
            self._drop_icon,
        ):
            if not getattr(self, "_trigger_pressed", False):
                return super().eventFilter(obj, event)
            self._trigger_pressed = False

            # Если popup был скрыт совсем недавно (~250 мс назад), не переоткрываем тем же кликом.
            if (time.monotonic() - float(getattr(self, "_popup_hidden_at", 0.0))) < 0.25:
                return True

            # Открываем после release (следующим тиком), чтобы не попасть под авто-close Qt.Popup.
            QTimer.singleShot(0, self._toggle_popup)
            return True

        return super().eventFilter(obj, event)

    # ---------- popup toggle ----------
    def _toggle_popup(self):
        pop = getattr(self, "_popup", None)
        if pop is None:
            return
        if pop.isVisible():
            pop.hide()
        else:
            pop.adjustSize()
            popup_h = pop.height() or pop.sizeHint().height() or 320
            win = self.window()
            win_rect = win.frameGeometry()
            f = self._trigger_frame
            btn_bottom_global = f.mapToGlobal(QPoint(0, f.height())).y()
            btn_top_global = f.mapToGlobal(QPoint(0, 0)).y()
            space_below = win_rect.bottom() - btn_bottom_global
            margin = 8
            popup_below_gap = 14

            if space_below - margin >= popup_h:
                pos = f.mapToGlobal(QPoint(0, f.height() + popup_below_gap))
            else:
                pos = f.mapToGlobal(QPoint(0, 0))
                pos.setY(btn_top_global - popup_h - margin)
            pop.move(pos)
            pop.show()

    # ---------- public API ----------
    def date(self) -> QDate:
        return self._date

    def time_(self) -> QTime:
        return self._time

    def setDate(self, d: QDate):
        self._date = d
        self._refresh_calendar()
        self._update_btn_text()

    def setTime(self, t: QTime):
        self._time = t
        self._refresh_time_labels()
        self._update_btn_text()


# -------------------- START MODE SELECT --------------------
class ModeSelectDialog(QDialog):
    """Стартовый выбор режима приложения (pc/tablet)."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.selected_mode = None

        self.setWindowTitle("Выберите режим")
        self.setModal(True)
        self.setFixedSize(360, 190)
        self.setStyleSheet(
            "QDialog { background: #ffffff; }"
            "QLabel { color: #111827; }"
            "QPushButton { background: #f8fafc; border: 1px solid #e2e8f0; border-radius: 12px; "
            "padding: 10px 14px; font-size: 15px; font-weight: 700; color: #111827; }"
            "QPushButton:hover { background: #f1f5f9; }"
            "QPushButton:pressed { background: #e2e8f0; }"
        )

        lay = QVBoxLayout(self)
        lay.setContentsMargins(18, 16, 18, 16)
        lay.setSpacing(12)

        title = QLabel("Выберите режим")
        title.setStyleSheet("font-size: 18px; font-weight: 800;")
        subtitle = QLabel("Как открыть приложение?")
        subtitle.setStyleSheet("font-size: 13px; color: #64748b;")
        lay.addWidget(title)
        lay.addWidget(subtitle)

        btn_row = QHBoxLayout()
        btn_row.setSpacing(10)
        self._btn_pc = QPushButton("Версия ПК")
        self._btn_tablet = QPushButton("Версия планшет")
        self._btn_pc.setCursor(Qt.PointingHandCursor)
        self._btn_tablet.setCursor(Qt.PointingHandCursor)
        self._btn_pc.clicked.connect(lambda: self._select("pc"))
        self._btn_tablet.clicked.connect(lambda: self._select("tablet"))
        btn_row.addWidget(self._btn_pc)
        btn_row.addWidget(self._btn_tablet)
        lay.addLayout(btn_row)

        lay.addStretch(1)

    def _select(self, mode: str):
        self.selected_mode = mode
        self.accept()


# -------------------- MAIN APP --------------------
class MirlisMarkApp(QWidget):
    def __init__(self, app_mode="pc"):
        super().__init__()
        self.app_mode = app_mode
        self.setWindowIcon(QIcon(resource_path("assets/mark_app.ico")))
        self.setWindowTitle(APP_TITLE)

        self.setMinimumSize(1100, 650)

        self.last_printed_preview_text = None
        self._last_printed_tspl_bytes = None
        self.last_history_entry = None

        self.label_w_mm = 70.0
        self.label_h_mm = 70.0

        self.products = []
        self.staff_made = []
        self.staff_checked = []
        self.loaded_at_str = "—"

        settings = _load_settings()
        try:
            default_excel = ensure_products_file()
        except RuntimeError as e:
            QMessageBox.critical(
                self,
                "Ошибка инициализации",
                str(e),
            )
            default_excel = os.path.join(app_data_dir(), "products.xlsx")
        self.excel_path = settings.get("excel_path", default_excel)

        self._updating_preview = False
        self._user_edited_preview = False
        self._preview_manual_mode = False
        self._base_font_size = 20
        self._preview_scale = 1.0
        self._screen_compensation_scale = 1.0
        self._last_screen_comp_sig = None

        self.history_entries = []
        self._history_filter_text = ""
        self.history_page = 0
        self.history_page_size = 5
        self._loading_from_history = False
        self._selected_history_id = None

        self._apply_global_style()
        self.init_ui()
        self.reload_excel(show_message=False)

        self.timer = QTimer(self)
        self.timer.timeout.connect(self.refresh_preview)
        self.timer.start(5000)

        # автообновление Excel раз в минуту (без всплывающих окон при фоновых ошибках)
        self._excel_autorefresh_timer = QTimer(self)
        self._excel_autorefresh_timer.timeout.connect(self._on_excel_autorefresh_tick)
        self._excel_autorefresh_timer.start(60_000)

        # архив напечатанных этикеток: миграция в новую структуру + очистка старше 365 дней при старте
        self._migrate_labels_archive_layout()
        self._cleanup_old_label_archives(days=365)

        self.refresh_preview()

    def event(self, ev):
        # Гасим WindowDeactivate когда активируется TabTip — иначе Qt каскадирует
        # FocusOut на lineEdit, completer закрывает popup, и при каждом нажатии
        # клавиши OSK видно моргание popup'а.
        try:
            if ev.type() == QEvent.WindowDeactivate:
                if _is_foreground_window_tabtip():
                    return True
        except Exception:
            pass
        return super().event(ev)

    # ---------------- STYLE ----------------
    def _apply_global_style(self):
        combo_btn_path = resource_path("assets/combo-btn.svg").replace("\\", "/")

        _qss = """
            QWidget {
                background: #f6f7f9;
                font-family: "Segoe UI";
                color: #111827;
            }

            #TopBar {
                background: #ffffff;
                border: 1px solid #e5e7eb;
                border-radius: 18px;
            }

            #Card {
                background: #ffffff;
                border: 1px solid #e5e7eb;
                border-radius: 18px;
            }

            #HistoryPanel {
                background: transparent;
            }

            #HistoryScroll {
                border: none;
                background: transparent;
            }

            #HistoryCard {
                background: #ffffff;
                border-radius: 14px;
                border: 1px solid #e5e7eb;
                padding: 10px 12px;
            }

            #HistoryCard:hover {
                background: #ffffff;
            }

            #HistoryCard[selected="true"] {
                border: 2px solid #4f46e5;
                background: #eef2ff;
            }

            #LabelWrap {
                background: transparent;
            }

            #LabelPreviewView {
                border: none;
                background: transparent;
            }

            #SectionTitle {
                background: #eef2f6;
                border-radius: 14px;
                padding: 10px 22px;
                font-family: "Inter","Segoe UI","Manrope","Arial",sans-serif;
                font-size: 22px;
                font-weight: 650;
                letter-spacing: 0.2px;
                color: #1E2F45;
            }

            #FieldLabel {
                background: #eef2f6;
                border-radius: 14px;
                padding: 10px 16px;
                font-size: 16px;
                font-weight: 500;
                color: #111827;
            }

            #Pill {
                background: #eef2f6;
                border-radius: 14px;
                padding: 8px 14px;
                font-size: 14px;
                font-weight: 600;
                color: #111827;
            }

            #ExcelPill {
                background: #eef2f6;
                border-radius: 18px;
                padding: 14px 22px;
                font-size: 14px;
                font-weight: 600;
                color: #0f172a;
            }

            QLineEdit, QTextEdit {
                background: #ffffff;
                border: 1px solid #cfd6e0;
                border-radius: 16px;
                padding: 10px 14px;
                font-size: 14px;
                selection-background-color: #cfe3ff;
            }
            QLineEdit:focus, QTextEdit:focus {
                border: 1px solid #6ea8fe;
            }

            QCheckBox {
                font-size: 14px;
                background: transparent;
            }

            QCheckBox:focus {
                outline: none;
            }

            QCheckBox::indicator {
                width: 18px;
                height: 18px;
            }

            QPushButton {
                border-radius: 16px;
                padding: 10px 18px;
                font-size: 14px;
                border: 1px solid #d0d7e2;
                background: #ffffff;
                color: #111827;
            }
            QPushButton:hover {
                background: #eef2ff;
            }
            QPushButton:pressed {
                background: #e5e7ff;
                border-color: #4f46e5;
            }
            QPushButton:disabled {
                background: #f3f4f6;
                color: #9ca3af;
                border-color: #e5e7eb;
            }

            #StepperBtn {
                background: #f9fafb;
                border-radius: 12px;
                border: 1px solid #d1d5db;
            }
            #StepperBtn:hover {
                background: #e5e7eb;
                border-color: #d1d5db;
            }
            #StepperBtn:pressed {
                background: #d1d5db;
            }

            #Btn_primary {
                background: #16a34a;
                border: 1px solid #15803d;
                color: #ffffff;
                font-weight: 800;
                font-size: 18px;
                letter-spacing: 0.8px;
                padding: 18px 18px;
                border-radius: 18px;
            }
            #Btn_primary:hover {
                background: #15803d;
                border-color: #166534;
            }
            #Btn_primary:pressed {
                background: #166534;
                border-color: #14532d;
            }
            #Btn_primary:disabled {
                background: #d1fae5;
                border: 1px solid #bbf7d0;
                color: rgba(255,255,255,0.8);
            }

            #Btn_secondary {
                background: #f9fafb;
                border: 1px solid #d1d5db;
                color: #24364D;
                font-weight: 600;
                font-size: 16px;
                padding: 18px 18px;
                border-radius: 18px;
            }
            #Btn_secondary:hover {
                background: #eef2ff;
                border-color: #4f46e5;
            }
            #Btn_secondary:pressed {
                background: #e0e7ff;
                border-color: #4338ca;
            }
            #Btn_secondary:disabled {
                color: #9ca3af;
                background: #f3f4f6;
                border-color: #e5e7eb;
            }

            /* Подпись «Количество»: тот же вид, что у Btn_secondary в покое, без hover/клика */
            #CopiesCaption {
                background: #f9fafb;
                border: 1px solid #d1d5db;
                color: #24364D;
                font-weight: 600;
                font-size: 16px;
                padding: 18px 18px;
                border-radius: 18px;
            }

            #Btn_danger {
                border: 1px solid #ef4444;
                color: #b91c1c;
                background: #fef2f2;
                font-weight: 600;
            }
            #Btn_danger:hover {
                background: #fee2e2;
                border-color: #dc2626;
            }
            #Btn_danger:pressed {
                background: #fecaca;
                border-color: #b91c1c;
            }
            #Btn_danger:disabled {
                background: #fef2f2;
                color: #fca5a5;
                border-color: #fecaca;
            }

            #StatsBtn {
                background: #fef3c7;
                border: 1px solid #f59e0b;
                color: #92400e;
                font-weight: 700;
            }
            #StatsBtn:hover {
                background: #fde68a;
            }

            #ToolBtn {
                border: 1px solid #d0d7e2;
                border-radius: 14px;
                padding: 8px 12px;
                background: #ffffff;
                min-width: 40px;
                font-weight: 800;
            }
            #ToolBtn:hover {
                background: #eef2ff;
                border-color: #4f46e5;
            }
            #ToolBtn:checked {
                background: #e0e7ff;
                border-color: #4f46e5;
            }

            QComboBox,
            QFontComboBox {
                min-height: 40px;
                padding: 0 44px 0 14px;
                border: 1px solid #cfd6e0;
                border-radius: 12px;
                background: #ffffff;
                font-size: 14px;
                color: #111827;
            }

            QComboBox:editable,
            QFontComboBox:editable {
                background: #ffffff;
            }

            QComboBox:focus,
            QFontComboBox:focus {
                border: 1px solid #94a3b8;
            }

            QComboBox::drop-down,
            QFontComboBox::drop-down {
                subcontrol-origin: border;
                subcontrol-position: center right;
                width: 36px;
                border: none;
                background: transparent;
                margin-right: 4px;
                image: url(assets/combo-btn.svg);
            }

            QComboBox::down-arrow,
            QFontComboBox::down-arrow {
                image: none;
                width: 0px;
                height: 0px;
            }

            QComboBox:disabled::down-arrow,
            QFontComboBox:disabled::down-arrow {
                image: none;
                width: 0px;
                height: 0px;
            }

            QComboBox QAbstractItemView,
            QFontComboBox QAbstractItemView {
                background: #ffffff;
                border: none;
                border-radius: 0px;
                selection-background-color: #eef2ff;
                selection-color: #3730a3;
                outline: none;
                padding: 0px;
                margin: 0px;
            }

            QComboBox QAbstractItemView::item,
            QFontComboBox QAbstractItemView::item {
                min-height: 36px;
                padding: 4px 10px;
                border-radius: 0px;
            }

            QComboBox QAbstractItemView::item:hover,
            QFontComboBox QAbstractItemView::item:hover {
                background: #f1f5f9;
                color: #111827;
            }

            QComboBox QAbstractItemView::item:selected,
            QFontComboBox QAbstractItemView::item:selected {
                background: #eef2ff;
                color: #3730a3;
            }

            /* Убираем внешний нативный прямоугольный контейнер popup (оставляем только styled view) */
            QComboBoxPrivateContainer,
            QFontComboBoxPrivateContainer {
                background: #ffffff;
                border: 1px solid #cbd5e1;
                border-radius: 12px;
                padding: 0px;
                margin: 0px;
            }
            QComboBoxPrivateContainer > QFrame,
            QFontComboBoxPrivateContainer > QFrame {
                background: transparent;
                border: none;
                margin: 0px;
                padding: 0px;
            }

            /* Скроллбар — глобально для всех виджетов */
            QScrollBar:vertical {
                background: #e5e7eb;
                border: none;
                border-radius: 5px;
                width: 10px;
                margin: 0px;
            }
            QScrollBar::handle:vertical {
                background: #f9b233;
                border-radius: 5px;
                min-height: 32px;
            }
            QScrollBar::handle:vertical:hover {
                background: #e6a020;
            }
            QScrollBar::add-line:vertical,
            QScrollBar::sub-line:vertical {
                height: 0px;
                subcontrol-origin: margin;
            }
            QScrollBar::add-page:vertical,
            QScrollBar::sub-page:vertical {
                background: transparent;
            }
            QScrollBar::up-arrow:vertical,
            QScrollBar::down-arrow:vertical {
                image: none;
                width: 0;
                height: 0;
            }
            """
        _qss = _qss.replace("url(assets/combo-btn.svg)", f"url({combo_btn_path})")
        # PC-only: popup списков — прямоугольный, скроллбар внутри с небольшим правым отступом.
        if getattr(self, "app_mode", "pc") != "tablet":
            _qss += """
            QComboBoxPrivateContainer,
            QFontComboBoxPrivateContainer {
                background: #ffffff;
                border: 1px solid #d0d7e2;
                border-radius: 0px;
            }
            QComboBoxPrivateContainer > QFrame,
            QFontComboBoxPrivateContainer > QFrame {
                background: transparent;
                border: none;
                margin: 0px;
                padding: 0px;
            }
            QComboBox QAbstractItemView,
            QFontComboBox QAbstractItemView {
                background: #ffffff;
                border: 1px solid #d0d7e2;
                border-radius: 0px;
                padding: 0px 14px 0px 0px;
            }
            QComboBox QAbstractItemView QScrollBar:vertical,
            QFontComboBox QAbstractItemView QScrollBar:vertical {
                width: 12px;
                margin: 2px 2px 2px 0px;
            }
            QComboBox QAbstractItemView QScrollBar::handle:vertical,
            QFontComboBox QAbstractItemView QScrollBar::handle:vertical {
                margin: 1px;
            }
            """
        self.setStyleSheet(_qss)

    # ---------------- Auto Excel refresh ----------------
    def _on_excel_autorefresh_tick(self):
        """Фоновое обновление Excel раз в минуту без MessageBox."""
        try:
            self.reload_excel(show_message=False, silent_errors=True)
        except TypeError:
            # на случай если сигнатура reload_excel не содержит silent_errors
            self.reload_excel(show_message=False)

    # ---------------- UI ----------------
    def init_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(18, 18, 18, 18)
        root.setSpacing(14)

        # -------- Top Bar --------
        top = QFrame()
        top.setObjectName("TopBar")
        top_layout = QHBoxLayout(top)
        top_layout.setContentsMargins(18, 14, 18, 14)
        top_layout.setSpacing(14)

        self.logo = QLabel()
        self.logo.setStyleSheet("background: transparent;")
        self.logo.setScaledContents(False)
        self._load_logo()
        top_layout.addWidget(self.logo, 0, Qt.AlignVCenter)

        title_block = QVBoxLayout()
        title_block.setSpacing(2)
        title_row = QHBoxLayout()
        title_row.setSpacing(10)

        self.title_mark = QLabel(APP_MARK)
        self.title_mark.setStyleSheet(
            'font-size: 32px; font-weight: 700; color: #1E2F45; '
            'font-family: "Inter","Segoe UI","Manrope","Arial",sans-serif; '
            "background: transparent;"
        )
        title_row.addWidget(self.title_mark)

        self.badge_ver = Pill(APP_VERSION)
        title_row.addWidget(self.badge_ver, 0, Qt.AlignVCenter)

        title_row.addStretch(1)
        title_block.addLayout(title_row)

        self.subtitle = QLabel(APP_SUBTITLE)
        self.subtitle.setStyleSheet("font-size: 16px; color: #64748b; padding-left: 2px; margin-top: 0px; background: transparent;")
        title_block.addWidget(self.subtitle)

        top_layout.addLayout(title_block, 0)

        self.excel_pill = QLabel("Excel: —")
        self.excel_pill.setObjectName("ExcelPill")
        self.excel_pill.setAlignment(Qt.AlignCenter)
        self.excel_pill.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        self.excel_pill.setMinimumWidth(360)
        self.excel_pill.setMaximumWidth(520)
        self.excel_pill.setMinimumHeight(48)

        top_layout.addStretch(3)

        # ===== Группа "Файлы данных" — обособленный фрейм с рамкой и тенью =====
        self.tools_frame = QFrame()
        self.tools_frame.setObjectName("ExcelToolsFrame")
        self.tools_frame.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.tools_frame.setMinimumWidth(900)
        self.tools_frame.setStyleSheet(
            "#ExcelToolsFrame { background: #ffffff; border: 1px solid #e5e7eb; border-radius: 18px; }"
        )
        _tools_shadow = QGraphicsDropShadowEffect(self.tools_frame)
        _tools_shadow.setBlurRadius(18)
        _tools_shadow.setOffset(0, 4)
        _tools_shadow.setColor(QColor(0, 0, 0, 35))
        self.tools_frame.setGraphicsEffect(_tools_shadow)

        tools_layout = QHBoxLayout(self.tools_frame)
        tools_layout.setContentsMargins(14, 8, 14, 8)
        tools_layout.setSpacing(10)
        tools_layout.addStretch(1)

        self.add_excel_btn = ActionBtn("Добавить", kind="default")
        self.add_excel_btn.setIcon(QIcon(resource_path("assets/addition.png")))
        self.add_excel_btn.setIconSize(QSize(28, 28))
        self.add_excel_btn.setStyleSheet(
            "#Btn_default { font-size: 16px; font-weight: 600; padding: 14px 22px; }"
        )
        self.add_excel_btn.setToolTip("Добавить новый Excel-файл с продуктами / цехами / сотрудниками")
        self.add_excel_btn.clicked.connect(self._on_add_excel)
        tools_layout.addWidget(self.add_excel_btn, 0, Qt.AlignVCenter)

        self.active_excel_btn = ActionBtn("Активные", kind="default")
        self.active_excel_btn.setIcon(QIcon(resource_path("assets/active.png")))
        self.active_excel_btn.setIconSize(QSize(28, 28))
        self.active_excel_btn.setStyleSheet(
            "#Btn_default { font-size: 16px; font-weight: 600; padding: 14px 22px; }"
        )
        self.active_excel_btn.setToolTip("Выбрать, какие из добавленных Excel-файлов используются в работе")
        self.active_excel_btn.clicked.connect(self._on_active_excel)
        tools_layout.addWidget(self.active_excel_btn, 0, Qt.AlignVCenter)

        # Двухстрочная подпись по центру группы
        self.files_caption = QWidget()
        self.files_caption.setObjectName("FilesCaption")
        self.files_caption.setStyleSheet(
            "#FilesCaption { background: transparent; }"
        )
        self.files_caption.setAttribute(Qt.WA_TranslucentBackground, False)
        _files_caption_lay = QVBoxLayout(self.files_caption)
        _files_caption_lay.setContentsMargins(18, 0, 18, 0)
        _files_caption_lay.setSpacing(2)
        _files_title = QLabel("Файлы данных")
        _files_title.setStyleSheet(
            "font-weight: 700; font-size: 16px; color: #111827; background: transparent;"
        )
        _files_title.setAlignment(Qt.AlignCenter)
        _files_sub = QLabel("Управление Excel-файлами приложения")
        _files_sub.setStyleSheet(
            "font-weight: 400; font-size: 12px; color: #6b7280; background: transparent;"
        )
        _files_sub.setAlignment(Qt.AlignCenter)
        _files_caption_lay.addWidget(_files_title)
        _files_caption_lay.addWidget(_files_sub)
        tools_layout.addWidget(self.files_caption, 0, Qt.AlignVCenter)

        self.delete_excel_btn = ActionBtn("Удалить", kind="default")
        self.delete_excel_btn.setIcon(QIcon(resource_path("assets/trash.png")))
        self.delete_excel_btn.setIconSize(QSize(28, 28))
        self.delete_excel_btn.setStyleSheet(
            "#Btn_default { font-size: 16px; font-weight: 600; padding: 14px 22px; }"
        )
        self.delete_excel_btn.setToolTip("Убрать Excel-файл из списка приложения (с диска НЕ удаляется)")
        self.delete_excel_btn.clicked.connect(self._on_delete_excel)
        tools_layout.addWidget(self.delete_excel_btn, 0, Qt.AlignVCenter)

        self.reload_btn = ActionBtn("Обновить", kind="default")
        self.reload_btn.setIcon(QIcon(resource_path("assets/update.png")))
        self.reload_btn.setIconSize(QSize(28, 28))
        self.reload_btn.setStyleSheet(
            "#Btn_default { font-size: 16px; font-weight: 600; padding: 14px 22px; }"
        )
        self.reload_btn.setToolTip("Перечитать активные Excel-файлы")
        self.reload_btn.clicked.connect(self.reload_excel)
        tools_layout.addWidget(self.reload_btn, 0, Qt.AlignVCenter)
        tools_layout.addStretch(1)

        top_layout.addWidget(self.tools_frame, 10)

        top_layout.addStretch(5)

        # ===== Кнопка статистики (справа) =====
        self.stats_btn = ActionBtn("Статистика", kind="default")
        self.stats_btn.setObjectName("StatsBtn")
        self.stats_btn.setIcon(self._make_stats_icon())
        self.stats_btn.setIconSize(QSize(28, 28))
        self.stats_btn.setStyleSheet(
            "#StatsBtn { background: #fef3c7; border: 1px solid #f59e0b; "
            "color: #92400e; font-weight: 700; font-size: 16px; padding: 14px 22px; }"
            "#StatsBtn:hover { background: #fde68a; }"
        )
        self.stats_btn.clicked.connect(self._open_statistics)
        top_layout.addWidget(self.stats_btn, 0, Qt.AlignVCenter)

        # excel_pill больше не показывается в верхней панели, но виджет создан
        # выше — оставляем его невидимым, чтобы существующие вызовы update_excel_status()
        # не падали. На следующих этапах он будет окончательно удалён.
        try:
            self.excel_pill.hide()
        except Exception:
            pass

        # --- Statistics mode top-bar controls (hidden by default) ---
        self.day_btn = ActionBtn("День", kind="default")
        self.week_btn = ActionBtn("7 дней", kind="default")
        self.month_btn = ActionBtn("30 дней", kind="default")
        self.period_btn = ActionBtn("Период", kind="default")
        self.back_to_stats_dashboard_btn = ActionBtn("Вернуться на главную статистики", kind="default")
        self.back_to_stats_dashboard_btn.setObjectName("BackToStatsDashboardBtn")
        self.back_to_stats_dashboard_btn.setStyleSheet(
            "#BackToStatsDashboardBtn {"
            "background: #FFF4CC;"  # light yellow fill
            "font-weight: 600;"
            "color: #9A670E;"       # darker text
            "border: 1px solid #B7791F;"  # stronger border
            "}"
            "#BackToStatsDashboardBtn:hover {"
            "background: #FDE68A;"  # slightly richer but still light
            "color: #8B5E0A;"
            "border-color: #A16207;"
            "}"
            "#BackToStatsDashboardBtn:pressed {"
            "background: #F5D76E;"  # a bit denser
            "color: #7C4A03;"
            "border-color: #92400E;"
            "}"
        )
        self.back_to_print_btn = ActionBtn("Вернуться в режим печати", kind="danger")

        self.day_btn.clicked.connect(lambda: self._set_statistics_period("day"))
        self.week_btn.clicked.connect(lambda: self._set_statistics_period("week"))
        self.month_btn.clicked.connect(lambda: self._set_statistics_period("month"))
        self.period_btn.clicked.connect(self._open_statistics_custom_period)
        self.back_to_stats_dashboard_btn.clicked.connect(self._return_to_statistics_dashboard)
        self.back_to_print_btn.clicked.connect(self._return_to_print_mode)

        for b in (self.day_btn, self.week_btn, self.month_btn, self.period_btn, self.back_to_stats_dashboard_btn, self.back_to_print_btn):
            b.setVisible(False)
            top_layout.addWidget(b, 0, Qt.AlignVCenter)

        root.addWidget(top)

        # -------- Content Stack (under top bar) --------
        self.content_stack = QStackedWidget()
        self.main_page = QWidget()
        self.stats_page = StatisticsPage()

        self.content_stack.addWidget(self.main_page)
        self.content_stack.addWidget(self.stats_page)
        self.content_stack.setCurrentWidget(self.main_page)
        root.addWidget(self.content_stack, 1)

        main_page_layout = QVBoxLayout(self.main_page)
        main_page_layout.setContentsMargins(0, 0, 0, 0)
        main_page_layout.setSpacing(0)

        self._stats_period = "day"
        self._statistics_mode = False
        self._set_statistics_mode(False)
        if hasattr(self, "stats_page") and hasattr(self.stats_page, "detail_mode_changed"):
            self.stats_page.detail_mode_changed.connect(self._on_statistics_detail_mode)

        # -------- Content Row --------
        row = QHBoxLayout()
        row.setSpacing(14)

        # left card (input)
        self.card_left = Card()
        left_layout = QVBoxLayout(self.card_left)
        left_layout.setContentsMargins(18, 18, 18, 18)
        left_layout.setSpacing(14)

        left_title = HeaderLabel("Ввод")
        left_layout.addWidget(left_title, 0, Qt.AlignHCenter)

        lab_prod = QLabel("Продукт")
        lab_prod.setObjectName("FieldLabel")
        lab_prod.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        left_layout.addWidget(lab_prod)

        self.product_combo = ComboBoxFixedArrow()
        self.product_combo.setEditable(True)
        self.product_combo.setInsertPolicy(QComboBox.NoInsert)
        self.product_combo.setPlaceholderText("Введите продукт или выберите из списка")
        self.product_combo.setMaxVisibleItems(16)
        # Для editable QComboBox setMaxVisibleItems иногда игнорируется стилем —
        # отключаем системный стиль popup, чтобы лимит работал на Windows.
        self.product_combo.setStyleSheet(
            self.product_combo.styleSheet() + " QComboBox { combobox-popup: 0; }"
        )
        self.product_combo.view().viewport().setAttribute(Qt.WA_AcceptTouchEvents, True)
        scroller = QScroller.scroller(self.product_combo.view().viewport())
        if scroller:
            props = scroller.scrollerProperties()
            props.setScrollMetric(QScrollerProperties.ScrollingCurve, QEasingCurve.OutQuad)
            props.setScrollMetric(QScrollerProperties.AxisLockThreshold, 0.0)
            scroller.setScrollerProperties(props)
        QScroller.grabGesture(self.product_combo.view().viewport(), QScroller.LeftMouseButtonGesture)
        left_layout.addWidget(self.product_combo)

        self.product_model = QStringListModel([])
        self.product_completer = QCompleter(self.product_model, self)
        self.product_completer.setCaseSensitivity(Qt.CaseInsensitive)
        self.product_completer.setFilterMode(Qt.MatchContains)
        self.product_completer.setCompletionMode(QCompleter.PopupCompletion)
        self.product_combo.setCompleter(self.product_completer)

        grid = QHBoxLayout()
        grid.setSpacing(12)

        col_units = QVBoxLayout()
        col_units.setSpacing(8)
        col_units.setContentsMargins(0, 0, 0, 0)
        col_units.setAlignment(Qt.AlignTop)

        lab_units = QLabel("Ед. изм.")
        lab_units.setObjectName("FieldLabel")
        lab_units.setAlignment(Qt.AlignCenter)
        lab_units.setFixedHeight(45)
        lab_units.setStyleSheet(
            "#FieldLabel { background: #eef2f6; border-radius: 14px; "
            "padding: 8px 16px; font-size: 16px; font-weight: 500; color: #1E2F45; }"
        )
        col_units.addWidget(lab_units)

        self.unit_combo = ComboBoxFixedArrow()
        self.unit_combo.setFixedHeight(45)
        self.unit_combo.setStyleSheet(
            "QComboBox { min-height: 45px; padding: 0 44px 0 14px; "
            "border: 1px solid #cfd6e0; border-radius: 12px; background: #ffffff; "
            "font-size: 14px; color: #111827; }"
            "QComboBox::drop-down { subcontrol-origin: border; subcontrol-position: center right; "
            "width: 36px; border: none; background: transparent; margin-right: 4px; }"
        )
        self.unit_combo.view().viewport().setAttribute(Qt.WA_AcceptTouchEvents, True)
        scroller = QScroller.scroller(self.unit_combo.view().viewport())
        if scroller:
            props = scroller.scrollerProperties()
            props.setScrollMetric(QScrollerProperties.ScrollingCurve, QEasingCurve.OutQuad)
            props.setScrollMetric(QScrollerProperties.AxisLockThreshold, 0.0)
            scroller.setScrollerProperties(props)
        QScroller.grabGesture(self.unit_combo.view().viewport(), QScroller.LeftMouseButtonGesture)

        # Кнопка "Взвесить" — считать вес с весов через COM-порт
        self.scale_btn = ActionBtn("Взвесить", kind="default")
        self.scale_btn.setToolTip("Считать вес с весов через COM-порт")
        self.scale_btn.setFocusPolicy(Qt.NoFocus)
        _scales_pix = QPixmap(resource_path("assets/scales.png"))
        if not _scales_pix.isNull():
            _scales_pix = _scales_pix.scaledToHeight(28, Qt.SmoothTransformation)
            self.scale_btn.setIcon(QIcon(_scales_pix))
            self.scale_btn.setIconSize(QSize(28, 28))
        self.scale_btn.setStyleSheet(
            "#Btn_default { font-size: 18px; font-weight: 700; padding: 8px 14px; }"
        )
        self.scale_btn.clicked.connect(self._read_scale_weight)

        col_units.addWidget(self.unit_combo)
        col_units.addSpacing(20)
        col_units.addWidget(self.scale_btn)
        col_units.addStretch(1)
        grid.addLayout(col_units, 2)

        col_qty = QVBoxLayout()
        col_qty.setSpacing(8)

        # --- Верхний ряд: поле ввода количества (на уровне лейбла "Ед. изм.") ---
        self.qty_input = QLineEdit()
        self.qty_input.setPlaceholderText("Введите количество")
        self.qty_input.setAlignment(Qt.AlignCenter)
        self.qty_input.setStyleSheet(
            "QLineEdit { font-size: 18px; font-weight: 700; padding: 8px 12px; }"
            "QLineEdit:focus { border: 1px solid #6ea8fe; }"
        )
        self.qty_input.setMinimumHeight(64)
        col_qty.addWidget(self.qty_input)

        # --- Нижний ряд: кнопки "−" и "+" (на уровне combobox единиц) ---
        qty_btn_row = QHBoxLayout()
        qty_btn_row.setSpacing(10)
        qty_btn_row.setContentsMargins(0, 0, 0, 0)
        # Высота кнопок −/+ должна совпадать с высотой unit_combo

        self.minus_btn = ActionBtn("−", kind="default")
        self.minus_btn.setAutoRepeat(True)
        self.minus_btn.setAutoRepeatDelay(400)
        self.minus_btn.setAutoRepeatInterval(80)
        self.minus_btn.setStyleSheet(
            "#Btn_default { font-size: 29px; font-weight: 900; padding: 0px; }"
        )

        self.plus_btn = ActionBtn("+", kind="default")
        self.plus_btn.setAutoRepeat(True)
        self.plus_btn.setAutoRepeatDelay(400)
        self.plus_btn.setAutoRepeatInterval(80)
        self.plus_btn.setStyleSheet(
            "#Btn_default { font-size: 29px; font-weight: 900; padding: 0px; }"
        )

        self.minus_btn.clicked.connect(self.decrease_qty)
        self.plus_btn.clicked.connect(self.increase_qty)
        self.minus_btn.setFocusPolicy(Qt.NoFocus)
        self.plus_btn.setFocusPolicy(Qt.NoFocus)

        qty_btn_row.addWidget(self.minus_btn, 1)
        qty_btn_row.addWidget(self.plus_btn, 1)
        col_qty.addSpacing(20)
        col_qty.addLayout(qty_btn_row)

        grid.addLayout(col_qty, 3)  # фактическое соотношение col_units:col_qty = 2:3

        # Синхронизация: поле ввода — крупнее, кнопки −/+ выровнены по высоте с unit_combo
        self.qty_input.setMinimumHeight(64)
        self.minus_btn.setMinimumHeight(64)
        self.plus_btn.setMinimumHeight(64)
        self.minus_btn.setMinimumWidth(70)
        self.plus_btn.setMinimumWidth(70)

        # Выравнивание кнопок −/+ по вертикали с unit_combo (под полем qty_input).
        # Поскольку qty_input выше unit_combo, qty_btn_row нужно прижать к низу — добавим stretch сверху qty_btn_row.
        qty_btn_row.setAlignment(Qt.AlignBottom)

        left_layout.addLayout(grid)

        # Опустить блоки "Цех" / "Изготовил" / "Дата и время" ниже — добавим вертикальный отступ
        left_layout.addSpacing(28)

        self.lab_chk = QLabel("Цех")
        self.lab_chk.setObjectName("FieldLabel")
        self.lab_chk.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        left_layout.addWidget(self.lab_chk)

        self.checked_combo = ComboBoxPopupUp()
        self.checked_combo.addItem("— не выбрано —")
        self.checked_combo.setMaxVisibleItems(8)
        self.checked_combo.view().viewport().setAttribute(Qt.WA_AcceptTouchEvents, True)
        scroller = QScroller.scroller(self.checked_combo.view().viewport())
        if scroller:
            props = scroller.scrollerProperties()
            props.setScrollMetric(QScrollerProperties.ScrollingCurve, QEasingCurve.OutQuad)
            props.setScrollMetric(QScrollerProperties.AxisLockThreshold, 0.0)
            scroller.setScrollerProperties(props)
        QScroller.grabGesture(self.checked_combo.view().viewport(), QScroller.LeftMouseButtonGesture)
        left_layout.addWidget(self.checked_combo)

        self.checked_manual = QCheckBox("Ручной ввод")
        left_layout.addWidget(self.checked_manual)

        self.checked_input = QLineEdit()
        self.checked_input.setPlaceholderText("Цех (можно оставить пустым)")
        self.checked_input.setVisible(False)
        left_layout.addWidget(self.checked_input)

        self.lab_made = QLabel("Изготовил")
        self.lab_made.setObjectName("FieldLabel")
        self.lab_made.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        left_layout.addWidget(self.lab_made)

        self.made_combo = ComboBoxFixedArrow()
        self.made_combo.addItem("— не выбрано —")
        self.made_combo.setMaxVisibleItems(8)
        self.made_combo.view().viewport().setAttribute(Qt.WA_AcceptTouchEvents, True)
        scroller = QScroller.scroller(self.made_combo.view().viewport())
        if scroller:
            props = scroller.scrollerProperties()
            props.setScrollMetric(QScrollerProperties.ScrollingCurve, QEasingCurve.OutQuad)
            props.setScrollMetric(QScrollerProperties.AxisLockThreshold, 0.0)
            scroller.setScrollerProperties(props)
        QScroller.grabGesture(self.made_combo.view().viewport(), QScroller.LeftMouseButtonGesture)
        left_layout.addWidget(self.made_combo)

        self.made_manual = QCheckBox("Ручной ввод")
        left_layout.addWidget(self.made_manual)

        self.made_input = QLineEdit()
        self.made_input.setPlaceholderText("ФИО (можно оставить пустым)")
        self.made_input.setVisible(False)
        left_layout.addWidget(self.made_input)

        lab_dt = QLabel("Дата и время")
        lab_dt.setObjectName("FieldLabel")
        lab_dt.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        self.manual_datetime_label = lab_dt

        self.manual_datetime_picker = CustomDateTimePicker()

        self.manual_datetime_label.setVisible(False)
        self.manual_datetime_picker.setVisible(False)
        left_layout.addWidget(self.manual_datetime_label)
        left_layout.addWidget(self.manual_datetime_picker)

        left_layout.addStretch(1)

        left_panel = QWidget()
        left_panel_layout = QVBoxLayout(left_panel)
        left_panel_layout.setContentsMargins(0, 0, 0, 0)
        left_panel_layout.setSpacing(0)
        left_panel_layout.addWidget(self.card_left)

        # right card (preview)
        self.card_right = Card()
        right_layout = QVBoxLayout(self.card_right)
        right_layout.setContentsMargins(18, 18, 18, 18)
        right_layout.setSpacing(14)

        self.preview_header = PreviewHeaderLabel("Предпросмотр")
        right_layout.addWidget(self.preview_header, 0, Qt.AlignHCenter)

        # toolbar
        tb = QHBoxLayout()
        tb.setSpacing(10)

        self.btn_font_minus = ActionBtn("A-", kind="default")
        self.btn_font_minus.setFixedWidth(60)
        self.btn_font_plus = ActionBtn("A+", kind="default")
        self.btn_font_plus.setFixedWidth(60)

        self.font_size_combo = ComboBoxFixedArrow()
        self.font_size_combo.setEditable(True)
        # Убираем completer: ComboBoxFixedArrow.showPopup при editable+completer+
        # непустом тексте уходит в completer-ветку и обычный выпадающий список
        # НЕ открывается. Для списка размеров шрифта автодополнение не нужно.
        # Должно стоять ПОСЛЕ setEditable(True) (он сам создаёт completer).
        self.font_size_combo.setCompleter(None)
        self.font_size_combo.setFixedWidth(130)
        self.font_size_combo.addItems([str(s) for s in [8, 9, 10, 11, 12, 14, 16, 18, 20, 22, 24, 26, 28, 36, 48, 72]])
        self.font_size_combo.setCurrentText(str(self._base_font_size))
        self.font_size_combo.view().viewport().setAttribute(Qt.WA_AcceptTouchEvents, True)
        scroller = QScroller.scroller(self.font_size_combo.view().viewport())
        if scroller:
            props = scroller.scrollerProperties()
            props.setScrollMetric(QScrollerProperties.ScrollingCurve, QEasingCurve.OutQuad)
            props.setScrollMetric(QScrollerProperties.AxisLockThreshold, 0.0)
            scroller.setScrollerProperties(props)
        QScroller.grabGesture(self.font_size_combo.view().viewport(), QScroller.LeftMouseButtonGesture)

        self.btn_bold = ToolBtn("Ж")
        self.btn_bold.setFont(QFont("Segoe UI", 11, QFont.Black))
        self.btn_italic = ToolBtn("К")
        f_it = QFont("Segoe UI", 11, QFont.Black)
        f_it.setItalic(True)
        self.btn_italic.setFont(f_it)
        self.btn_underline = ToolBtn("Ч")
        f_un = QFont("Segoe UI", 11, QFont.Black)
        f_un.setUnderline(True)
        self.btn_underline.setFont(f_un)

        self.btn_align_left = ToolBtn("")
        self.btn_align_center = ToolBtn("")
        self.btn_align_right = ToolBtn("")

        for btn, mode in (
            (self.btn_align_left, "left"),
            (self.btn_align_center, "center"),
            (self.btn_align_right, "right"),
        ):
            btn.setIcon(self._make_align_icon(mode))
            btn.setIconSize(QSize(26, 18))
            btn.setFixedWidth(62)

        self.font_combo = QFontComboBox()
        self.font_combo.setObjectName("ComboWithArrow")
        self.font_combo.setFixedWidth(180)
        # делаем список шрифтов чище (без агрессивного предпросмотра системным рендером)
        try:
            self.font_combo.setView(QListView())
            self.font_combo.view().setItemDelegate(
                PlainFontNameDelegate(self.font_combo.view(), min_h=36)
            )
        except Exception:
            pass
        self.font_combo.view().viewport().setAttribute(Qt.WA_AcceptTouchEvents, True)
        scroller = QScroller.scroller(self.font_combo.view().viewport())
        if scroller:
            props = scroller.scrollerProperties()
            props.setScrollMetric(QScrollerProperties.ScrollingCurve, QEasingCurve.OutQuad)
            props.setScrollMetric(QScrollerProperties.AxisLockThreshold, 0.0)
            scroller.setScrollerProperties(props)
        QScroller.grabGesture(self.font_combo.view().viewport(), QScroller.LeftMouseButtonGesture)

        tb.addWidget(self.btn_font_minus)
        tb.addWidget(self.btn_font_plus)
        tb.addWidget(self.font_size_combo)
        tb.addSpacing(8)
        tb.addWidget(self.btn_bold)
        tb.addWidget(self.btn_italic)
        tb.addWidget(self.btn_underline)
        tb.addSpacing(8)
        tb.addWidget(self.btn_align_left)
        tb.addWidget(self.btn_align_center)
        tb.addWidget(self.btn_align_right)
        tb.addStretch(1)
        tb.addWidget(self.font_combo)

        right_layout.addLayout(tb)

        # preview editor
        self.preview = QTextEdit()
        self.preview.setObjectName("PreviewEditor")
        self.preview.setAcceptRichText(True)
        self.preview.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.preview.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)

        lw0, lh0 = self._logical_preview_size()
        self.preview.setFixedSize(lw0, lh0)

        self.preview.setStyleSheet(
            """
            QTextEdit {
                background: #ffffff;
                border: 1px solid #cfd6e0;
                border-radius: 18px;
                padding: 18px 2px;
            }
            """
        )
        self.preview.setReadOnly(True)

        self.preview_wrap = QFrame()
        self.preview_wrap.setObjectName("LabelWrap")

        # Логический размер этикетки фиксирован (450 × пропорция мм); вписывание в окно — только
        # масштаб QGraphicsView, без изменения pt в QTextDocument (иначе «плывёт» между мониторами).
        self._preview_scene = QGraphicsScene(self)
        self._preview_proxy = QGraphicsProxyWidget()
        self._preview_proxy.setWidget(self.preview)
        self._preview_scene.addItem(self._preview_proxy)

        self.preview_view = QGraphicsView(self._preview_scene, self.preview_wrap)
        self.preview_view.setObjectName("LabelPreviewView")
        self.preview_view.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.preview_view.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.preview_view.setFrameShape(QFrame.NoFrame)
        self.preview_view.setBackgroundBrush(QBrush(Qt.transparent))
        self.preview_view.setAlignment(Qt.AlignCenter)
        self.preview_view.setDragMode(QGraphicsView.NoDrag)
        self.preview_view.setRenderHint(QPainter.Antialiasing, True)
        self.preview_view.setRenderHint(QPainter.TextAntialiasing, True)
        self.preview_view.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.preview_view.setMinimumSize(80, 80)

        wrap_lay = QGridLayout(self.preview_wrap)
        wrap_lay.setContentsMargins(12, 12, 12, 12)
        wrap_lay.setHorizontalSpacing(0)
        wrap_lay.setVerticalSpacing(0)
        wrap_lay.setRowStretch(0, 1)

        # Этикетка строго по центру области предпросмотра (масштаб view, не документа)
        wrap_lay.addWidget(self.preview_view, 0, 0, 1, 3, Qt.AlignHCenter | Qt.AlignTop)

        # Кнопка «Очистить» — правый нижний угол поверх preview_wrap
        self.clear_btn = ActionBtn("Очистить", kind="danger")
        self.clear_btn.clicked.connect(self.clear_fields)
        if getattr(self, "app_mode", "pc") == "tablet":
            self.clear_btn.setFixedWidth(198)
        else:
            self.clear_btn.setFixedWidth(138)
        if getattr(self, "app_mode", "pc") == "tablet":
            self.clear_btn.setFixedHeight(84)
        else:
            self.clear_btn.setFixedHeight(64)
        wrap_lay.addWidget(self.clear_btn, 1, 2, Qt.AlignRight | Qt.AlignBottom)

        right_layout.addWidget(self.preview_wrap, 1)

        center_panel = QWidget()
        center_panel_layout = QVBoxLayout(center_panel)
        center_panel_layout.setContentsMargins(0, 0, 0, 0)
        center_panel_layout.setSpacing(10)
        center_panel_layout.addWidget(self.card_right)

        label_size_row = QHBoxLayout()
        label_size_row.setSpacing(8)
        label_size_lab = QLabel("Размер этикетки")
        label_size_lab.setObjectName("FieldLabel")
        self.label_size_combo = ComboBoxFixedArrow()
        self.label_size_combo.addItem("58×80 мм")
        self.label_size_combo.addItem("58×60 мм")
        self.label_size_combo.addItem("70×70 мм")
        self.label_size_combo.addItem("Цветные")
        self.label_size_combo.setCurrentIndex(2)
        self.label_size_combo.setMinimumWidth(140)
        label_size_row.addWidget(label_size_lab)
        label_size_row.addWidget(self.label_size_combo, 1)
        center_panel_layout.addLayout(label_size_row)

        # print row
        pr = QHBoxLayout()
        pr.setSpacing(8)

        self.print_btn = ActionBtn("ПЕЧАТЬ", kind="primary")
        self.print_btn.setMinimumWidth(360)

        self.copies_caption = QLabel("Количество")
        self.copies_caption.setObjectName("CopiesCaption")
        self.copies_caption.setAlignment(Qt.AlignCenter)
        self.copies_caption.setFocusPolicy(Qt.NoFocus)
        self.copies_caption.setVisible(False)  # убран из UI, скрыт
        self.copies_minus = ActionBtn("−", kind="default")
        self.copies_minus.setFixedWidth(106)
        self.copies_input = QLineEdit("")
        self.copies_input.setMinimumWidth(80)
        self.copies_input.setMaximumWidth(160)   # не расползается — освобождает место под Печать
        self.copies_input.setAlignment(Qt.AlignCenter)
        self.copies_plus = ActionBtn("+", kind="default")
        self.copies_plus.setFixedWidth(106)
        for w in (self.minus_btn, self.plus_btn, self.copies_minus, self.copies_plus):
            w.setObjectName("StepperBtn")
        self.minus_btn.setStyleSheet(
            "#StepperBtn { font-size: 29px; font-weight: 900; padding: 0px; }"
        )
        self.plus_btn.setStyleSheet(
            "#StepperBtn { font-size: 29px; font-weight: 900; padding: 0px; }"
        )
        # Inline-стили на виджете; для кнопок — #StepperBtn (имя задаётся строкой выше)
        self.copies_minus.setStyleSheet(
            "#StepperBtn { font-size: 29px; font-weight: 900; padding: 0px; }"
        )
        self.copies_plus.setStyleSheet(
            "#StepperBtn { font-size: 29px; font-weight: 900; padding: 0px; }"
        )
        self.copies_input.setStyleSheet(
            "QLineEdit { font-size: 20px; font-weight: 700; padding: 4px 6px; }"
            "QLineEdit:focus { border: 1px solid #6ea8fe; }"
        )
        _setup_auto_repeat(self.copies_minus, self.decrease_copies)
        _setup_auto_repeat(self.copies_plus, self.increase_copies)
        self.copies_minus.setFocusPolicy(Qt.NoFocus)
        self.copies_plus.setFocusPolicy(Qt.NoFocus)

        copies_wrap = QWidget()
        cw = QHBoxLayout(copies_wrap)
        cw.setContentsMargins(0, 0, 0, 0)
        cw.setSpacing(6)
        # copies_caption убран — освобождает место под print_btn и copies_input
        cw.addWidget(self.copies_minus, 0)
        cw.addWidget(self.copies_input, 1)  # stretch=1: занимает всё свободное место
        cw.addWidget(self.copies_plus, 0)

        for w in (self.print_btn, self.copies_minus, self.copies_plus, self.copies_input):
            w.setMinimumHeight(68)

        pr.addWidget(self.print_btn, 2)
        pr.addWidget(copies_wrap, 1)

        center_panel_layout.addLayout(pr)

        # -------- History panel (right) --------
        self.history_panel = QWidget()
        self.history_panel.setObjectName("HistoryPanel")
        history_layout = QVBoxLayout(self.history_panel)
        history_layout.setContentsMargins(18, 18, 18, 18)
        history_layout.setSpacing(12)

        history_title = HeaderLabel("История")

        self.open_labels_folder_btn = QPushButton()
        self.open_labels_folder_btn.setObjectName("OpenLabelsFolderBtn")
        self.open_labels_folder_btn.setIcon(QIcon(resource_path("assets/folder.png")))
        self.open_labels_folder_btn.setIconSize(QSize(32, 32))
        self.open_labels_folder_btn.setFixedSize(66, 44)
        self.open_labels_folder_btn.setCursor(Qt.PointingHandCursor)
        self.open_labels_folder_btn.setToolTip("Открыть папку с напечатанными этикетками")
        self.open_labels_folder_btn.setStyleSheet(
            "#OpenLabelsFolderBtn { background: #ffffff; border: 1px solid #d0d7e2; "
            "border-radius: 12px; padding: 0px; }"
            "#OpenLabelsFolderBtn:hover { background: #eef2ff; border-color: #d0d7e2; }"
            "#OpenLabelsFolderBtn:pressed { background: #e5e7ff; border-color: #4f46e5; }"
        )
        self.open_labels_folder_btn.clicked.connect(self.open_labels_folder)

        history_title_row = QHBoxLayout()
        history_title_row.setContentsMargins(0, 0, 0, 0)
        history_title_row.setSpacing(0)

        # Невидимый спейсер слева такой же ширины как кнопка справа,
        # чтобы заголовок остался строго по центру панели
        _title_left_spacer = QWidget()
        _title_left_spacer.setFixedWidth(66)
        _title_left_spacer.setFixedHeight(1)
        _title_left_spacer.setStyleSheet("background: transparent;")
        history_title_row.addWidget(_title_left_spacer, 0)

        history_title_row.addStretch(1)
        history_title_row.addWidget(history_title, 0, Qt.AlignVCenter)
        history_title_row.addStretch(1)
        history_title_row.addWidget(self.open_labels_folder_btn, 0, Qt.AlignVCenter)
        history_layout.addLayout(history_title_row)

        self.history_search = QLineEdit()
        self.history_search.setPlaceholderText("Поиск по истории")
        history_layout.addWidget(self.history_search)

        self.history_scroll = QScrollArea()
        QScroller.grabGesture(self.history_scroll.viewport(), QScroller.TouchGesture)
        self.history_scroll.setObjectName("HistoryScroll")
        self.history_scroll.setWidgetResizable(True)
        self.history_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.history_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        history_layout.addWidget(self.history_scroll, 1)

        history_scroll_content = QWidget()
        self.history_list_layout = QVBoxLayout(history_scroll_content)
        self.history_list_layout.setContentsMargins(0, 0, 0, 0)
        self.history_list_layout.setSpacing(10)

        self.history_scroll.setWidget(history_scroll_content)

        self.history_prev_btn = ActionBtn("←", kind="default")
        self.history_prev_btn.setFixedSize(72, 56)
        self.history_prev_btn.setStyleSheet(
            "#Btn_default { font-size: 22px; font-weight: 900; padding: 0px; }"
        )
        self.history_next_btn = ActionBtn("→", kind="default")
        self.history_next_btn.setFixedSize(72, 56)
        self.history_next_btn.setStyleSheet(
            "#Btn_default { font-size: 22px; font-weight: 900; padding: 0px; }"
        )
        self.history_page_label = QLabel("Страница 1 из 1")
        self.history_page_label.setAlignment(Qt.AlignCenter)
        self.history_page_label.setStyleSheet(
            "color: #4b5563; font-size: 15px; font-weight: 700; background: transparent;"
        )

        # Контейнер-фрейм для всей панели пагинации
        pager_frame = QFrame()
        pager_frame.setObjectName("PagerFrame")
        pager_frame.setStyleSheet(
            "#PagerFrame { background: #ffffff; border: 1px solid #e5e7eb; border-radius: 14px; }"
        )
        pager_row = QHBoxLayout(pager_frame)
        pager_row.setContentsMargins(12, 10, 12, 10)
        pager_row.setSpacing(8)

        pager_row.addWidget(self.history_prev_btn, 0)
        pager_row.addWidget(self.history_page_label, 1)
        pager_row.addWidget(self.history_next_btn, 0)

        # pager_frame: единственное место добавления в history_layout
        pager_frame.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        history_layout.addWidget(pager_frame, 0)

        row.addWidget(left_panel, 3)
        row.addWidget(center_panel, 4)
        row.addWidget(self.history_panel, 3)
        main_page_layout.addLayout(row)

        # ---------------- Signals ----------------
        self.product_combo.currentTextChanged.connect(self.on_product_changed)
        self.unit_combo.currentTextChanged.connect(self.refresh_preview)
        self.qty_input.textChanged.connect(self.refresh_preview)

        self.made_manual.stateChanged.connect(self.toggle_made_mode)
        self.checked_manual.stateChanged.connect(self.toggle_checked_mode)

        self.made_combo.currentTextChanged.connect(self.refresh_preview)
        self.checked_combo.currentTextChanged.connect(self.refresh_preview)
        self.made_input.textChanged.connect(self.refresh_preview)
        self.checked_input.textChanged.connect(self.refresh_preview)

        self.print_btn.clicked.connect(self.print_label)
        self.copies_input.textChanged.connect(self._sanitize_copies)
        self.copies_input.textChanged.connect(lambda _: self._refresh_print_btn_state())
        self.label_size_combo.currentIndexChanged.connect(self._on_label_size_changed)

        self.preview.textChanged.connect(self._on_preview_text_changed)
        self.preview.cursorPositionChanged.connect(self._sync_format_toolbar_from_cursor)
        self.preview.selectionChanged.connect(self._sync_format_toolbar_from_cursor)

        self.btn_font_minus.clicked.connect(lambda: self._change_font_size(-1))
        self.btn_font_plus.clicked.connect(lambda: self._change_font_size(+1))
        self.font_size_combo.currentTextChanged.connect(self.on_font_size_combo_changed)

        self.btn_bold.clicked.connect(self._toggle_bold_on_selection)
        self.btn_italic.clicked.connect(self._toggle_italic_on_selection)
        self.btn_underline.clicked.connect(self._toggle_underline_on_selection)

        self.btn_align_left.clicked.connect(lambda: self._set_alignment(Qt.AlignLeft))
        self.btn_align_center.clicked.connect(lambda: self._set_alignment(Qt.AlignHCenter))
        self.btn_align_right.clicked.connect(lambda: self._set_alignment(Qt.AlignRight))

        self.font_combo.currentFontChanged.connect(self._set_font_family_on_selection)
        self.preview_header.doubleClicked.connect(self._toggle_preview_manual_mode)
        self.manual_datetime_picker.dateTimeChanged.connect(self.refresh_preview)

        self.history_search.textChanged.connect(self._on_history_search_text_changed)
        self.history_prev_btn.clicked.connect(lambda: self._change_history_page(-1))
        self.history_next_btn.clicked.connect(lambda: self._change_history_page(+1))

        for cb in (self.product_combo, self.unit_combo, self.made_combo, self.checked_combo, self.font_combo, self.label_size_combo):
            cb.setObjectName("ComboWithArrow")

        if getattr(self, "app_mode", "pc") == "tablet":
            self._apply_tablet_combobox_popups()
            self._apply_tablet_ui_tweaks()

        self._resize_label_preview()
        eff0 = self._effective_preview_font_scale()
        self.preview.setFont(QFont("Segoe UI", max(8, int(round(self._base_font_size * eff0)))))

        self._rebuild_history_view()

    def _apply_tablet_combobox_popups(self):
        """
        Tablet-only: делаем выпадающие списки QComboBox удобными для тач-ввода.
        Важно: влияет только на popup/list view, не увеличивает весь UI целиком.
        """

        # App-level фильтр: трекает timestamp последнего клика/тача ВНУТРИ окна
        # приложения. Используется чтобы отличать "пользователь кликнул в UI"
        # (popup должен закрыться навсегда) от "сенсорная клавиатура закрылась"
        # (popup должен переоткрыться, если lineEdit с текстом).
        if not hasattr(self, "_app_click_tracker_installed"):
            class _AppClickTracker(QObject):
                def __init__(self, win):
                    super().__init__(win)
                    self._win = win
                def eventFilter(self, obj, event):
                    et = event.type()
                    if et in (QEvent.MouseButtonPress, QEvent.TouchBegin):
                        try:
                            w = self._win
                            target_widget = None
                            in_window = False
                            if w is not None and hasattr(event, "globalPos"):
                                pos = event.globalPos()
                                tl = w.mapToGlobal(w.rect().topLeft())
                                br = w.mapToGlobal(w.rect().bottomRight())
                                in_window = (tl.x() <= pos.x() <= br.x() and tl.y() <= pos.y() <= br.y())
                                try:
                                    target_widget = QApplication.widgetAt(pos)
                                except Exception:
                                    target_widget = None
                            else:
                                in_window = True

                            if in_window:
                                import time as _t
                                self._win._last_app_user_click_ts = _t.time()
                                self._win._last_app_user_click_widget = target_widget
                                # Клик в чужой виджет → явное закрытие completer-popup.
                                try:
                                    if target_widget is not None:
                                        for combo in self._win.findChildren(QComboBox):
                                            if not isinstance(combo, ComboBoxFixedArrow):
                                                continue
                                            comp = combo.completer()
                                            if comp is None:
                                                continue
                                            cp = comp.popup()
                                            if cp is None or not cp.isVisible():
                                                continue
                                            le = combo.lineEdit()
                                            if le is not None:
                                                try:
                                                    if target_widget is le or le.isAncestorOf(target_widget):
                                                        continue
                                                except Exception:
                                                    pass
                                            try:
                                                if target_widget is cp or cp.isAncestorOf(target_widget):
                                                    continue
                                            except Exception:
                                                pass
                                            combo._explicit_close = True
                                except Exception:
                                    pass
                        except Exception:
                            pass
                    return False

            try:
                app = QApplication.instance()
                self._app_click_tracker = _AppClickTracker(self)
                if app is not None:
                    app.installEventFilter(self._app_click_tracker)
                self._app_click_tracker_installed = True
                self._last_app_user_click_ts = 0
            except Exception:
                pass

        class _TabletComboItemDelegate(QStyledItemDelegate):
            def __init__(self, parent=None, min_h: int = 54):
                super().__init__(parent)
                self._min_h = int(min_h)

            def sizeHint(self, option, index):
                s = super().sizeHint(option, index)
                return QSize(s.width(), max(int(s.height()), self._min_h))

        def _rows_for_combo(combo: QComboBox) -> int:
            # Требуемые лимиты (tablet-only)
            if hasattr(self, "product_combo") and combo is self.product_combo:
                return 8
            if hasattr(self, "made_combo") and combo is self.made_combo:
                return 4
            if hasattr(self, "checked_combo") and combo is self.checked_combo:
                return 4
            if hasattr(self, "font_size_combo") and combo is self.font_size_combo:
                return 6
            if hasattr(self, "font_combo") and combo is self.font_combo:
                return 8
            if hasattr(self, "label_size_combo") and combo is self.label_size_combo:
                return 4
            if hasattr(self, "unit_combo") and combo is self.unit_combo:
                return 2
            return 8

        def _popup_width_mult(combo: QComboBox) -> float:
            if hasattr(self, "font_size_combo") and combo is self.font_size_combo:
                return 1.5
            return 1.0

        def _apply_popup_geometry(combo: QComboBox, view: QListView):
            # Важно: с кастомным view Qt/PyQt5 не всегда ограничивает высоту popup сам.
            # Поэтому делаем 2 уровня: setMaxVisibleItems + ручная фиксация высоты view на Show.
            limit = int(_rows_for_combo(combo))
            count = int(combo.count())
            visible_rows = max(1, min(count, limit))

            row_h = 0
            try:
                if count > 0:
                    row_h = int(view.sizeHintForRow(0) or 0)
            except Exception:
                row_h = 0
            if row_h <= 0:
                row_h = 54  # fallback tablet

            try:
                spacing = int(view.spacing() or 0)
            except Exception:
                spacing = 0

            try:
                frame = int(view.frameWidth() or 0) * 2
            except Exception:
                frame = 0

            total_height = visible_rows * row_h + max(0, visible_rows - 1) * spacing + frame + 12

            # Детерминированная высота строки tablet (QSS min-height: 54px + item padding).
            # sizeHintForRow ненадёжен на момент расчёта, поэтому фиксируем явно.
            ROW_H_TABLET = 62   # было 56 — последняя строка с выделением обрезалась
            VIEW_VPAD = 24      # было 20 — чуть больше воздуха сверху/снизу
            FRAME_PAD = 4       # рамка контейнера

            if combo is self.product_combo:
                total_height = int(total_height * 0.75)
            elif hasattr(self, "font_combo") and combo is self.font_combo:
                total_height = int(total_height * 0.7)
            elif hasattr(self, "unit_combo") and combo is self.unit_combo:
                total_height = int(total_height * 0.9)
            elif (
                combo is self.made_combo
                or (hasattr(self, "checked_combo") and combo is self.checked_combo)
                or (hasattr(self, "label_size_combo") and combo is self.label_size_combo)
            ):
                # Высота строго по числу видимых строк (= min(count, limit)),
                # без пустого места и без обрезания. visible_rows уже посчитан выше.
                total_height = visible_rows * ROW_H_TABLET + VIEW_VPAD + FRAME_PAD
            view.setFixedHeight(max(60, int(total_height)))

            # product_combo / font_combo: ограничить и внешний popup (иначе остаётся «лишняя» высота)
            popup = view.window()
            if popup is not None and combo is self.product_combo:
                popup.setFixedHeight(int(total_height + 4))
            if popup is not None and hasattr(self, "font_combo") and combo is self.font_combo:
                popup.setFixedHeight(int(total_height + 4))
            if popup is not None and hasattr(self, "label_size_combo") and combo is self.label_size_combo:
                popup.setFixedHeight(int(total_height + 4))
            if popup is not None and hasattr(self, "unit_combo") and combo is self.unit_combo:
                popup.setFixedHeight(int(total_height + 4))
            if popup is not None and hasattr(self, "made_combo") and combo is self.made_combo:
                popup.setFixedHeight(int(total_height + 4))
            if popup is not None and hasattr(self, "checked_combo") and combo is self.checked_combo:
                popup.setFixedHeight(int(total_height + 4))
            if popup is not None and hasattr(self, "font_size_combo") and combo is self.font_size_combo:
                popup.setFixedHeight(int(total_height + 4))

            base_w = combo.width() if combo.width() > 0 else combo.sizeHint().width()
            view.setMinimumWidth(int(base_w * float(_popup_width_mult(combo))))

            # После финальной геометрии — синхронизируем rounded mask контейнера (tablet-only)
            try:
                if popup is not None:
                    w = int(popup.width() or 0)
                    h = int(popup.height() or 0)
                    if w > 0 and h > 0:
                        from PyQt5.QtGui import QRegion, QPainterPath

                        mask_radius = 13
                        path = QPainterPath()
                        path.addRoundedRect(0, 0, w, h, mask_radius, mask_radius)
                        region = QRegion(path.toFillPolygon().toPolygon())
                        popup.setMask(region)
            except Exception:
                pass

        class _TabletPopupSizer(QObject):
            def __init__(self, combo: QComboBox, view: QListView):
                super().__init__(view)
                self._combo = combo
                self._view = view

            def eventFilter(self, obj, event):
                if event.type() == QEvent.Show:
                    try:
                        obj.setStyleSheet(
                            "background: #ffffff; border: 1.5px solid #cbd5e1; "
                            "border-radius: 13px; padding: 0px;"
                        )
                        _apply_popup_geometry(self._combo, self._view)
                    except Exception:
                        pass
                return False

        class _TabletComboPreShow(QObject):
            """made_combo / checked_combo / font_combo / unit_combo: геометрия до первого show (QEvent.Show у popup — поздно)."""

            def __init__(self, apply_fn, combo: QComboBox, view: QListView):
                super().__init__(combo)
                self._apply_fn = apply_fn
                self._combo = combo
                self._view = view

            def eventFilter(self, obj, event):
                et = event.type()
                if et in (QEvent.MouseButtonPress, QEvent.TouchBegin):
                    try:
                        # Применяем стиль контейнера ДО показа, чтобы не было чёрного flash
                        popup = self._view.window()
                        if popup is not None:
                            popup.setStyleSheet(
                                "background: #ffffff; border: 1.5px solid #cbd5e1; "
                                "border-radius: 13px; padding: 0px;"
                            )
                        self._apply_fn(self._combo, self._view)
                    except Exception:
                        pass
                return False

        qss = """
            QAbstractItemView {
                background: #ffffff;
                border: none;
                border-radius: 13px;
                outline: none;
                padding: 10px 10px 10px 6px;
                margin: 0px;
            }

            QAbstractItemView::viewport {
                background: #ffffff;
                border: none;
                border-radius: 13px;
            }

            QAbstractItemView::item {
                min-height: 54px;
                padding: 12px 18px;
                border-radius: 10px;
                font-size: 17px;
            }

            QAbstractItemView::item:hover {
                background: #f1f5f9;
            }

            QAbstractItemView::item:selected {
                background: #eef2ff;
                color: #3730a3;
            }

            /* Современный скроллбар для tablet dropdown (как в PC стиле) */
            QScrollBar {
                background: transparent;
            }
            QScrollBar:vertical {
                background: transparent;
                border: none;
                border-radius: 6px;
                width: 52px;
                margin: 4px 10px 4px 0px;
            }
            QScrollBar::handle:vertical {
                background: #f9b233;
                border-radius: 10px;
                min-height: 60px;
                margin: 0px;
            }
            QScrollBar::handle:vertical:hover {
                background: #e6a020;
            }
            QScrollBar::add-line:vertical,
            QScrollBar::sub-line:vertical {
                height: 0px;
                subcontrol-origin: margin;
            }
            QScrollBar::add-page:vertical,
            QScrollBar::sub-page:vertical {
                background: transparent;
            }
            QScrollBar::up-arrow:vertical,
            QScrollBar::down-arrow:vertical {
                image: none;
                width: 0;
                height: 0;
            }
        """

        for combo in self.findChildren(QComboBox):
            try:
                limit = int(_rows_for_combo(combo))
                # Главный механизм ограничения высоты popup (tablet-only) — стандартный Qt.
                combo.setMaxVisibleItems(limit)

                # закрытый комбобокс — крупнее текст (~1.5×), без изменения popup (QListView ниже)
                combo.setMinimumHeight(48)
                combo.setStyleSheet(
                    "QComboBox { min-height: 48px; padding: 6px 12px; font-size: 21px; }"
                    "QComboBox::drop-down { width: 44px; }"
                    "QComboBox QLineEdit { font-size: 21px; }"
                )
                if combo is self.font_size_combo:
                    combo.setStyleSheet(
                        "QComboBox { min-height: 48px; padding-top: 6px; padding-bottom: 6px; "
                        "padding-left: 16px; padding-right: 12px; font-size: 21px; }"
                        "QComboBox::drop-down { width: 44px; }"
                        "QComboBox QLineEdit { font-size: 21px; }"
                    )

                # popup — свой QListView с принудительной высотой элементов
                lv = QListView()
                lv.setUniformItemSizes(True)
                lv.setSpacing(6)
                _lv_qss = qss
                # unit_combo (tablet): не расширяем popup, но увеличиваем полезную ширину item-area
                # за счёт меньшего правого padding у view (иначе короткие "КГ/ШТ" выглядят зажатыми).
                if hasattr(self, "unit_combo") and combo is self.unit_combo:
                    _lv_qss = _lv_qss.replace(
                        "padding: 10px 10px 10px 6px;", "padding: 10px 10px 10px 6px;"
                    )
                    _lv_qss = _lv_qss.replace(
                        "padding: 12px 18px;", "padding: 12px 14px;"
                    )
                lv.setStyleSheet(_lv_qss)
                lv.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
                lv.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)

                f = lv.font()
                f.setPointSize(max(f.pointSize(), 12))
                lv.setFont(f)

                if combo is self.font_combo:
                    lv.setItemDelegate(PlainFontNameDelegate(lv, min_h=54))
                else:
                    lv.setItemDelegate(_TabletComboItemDelegate(lv, min_h=54))
                combo.setView(lv)

                # Единый кастомный скроллбар на ВСЕХ tablet popup'ах (а не только у font_combo)
                try:
                    old_sb = lv.verticalScrollBar()
                    if old_sb is None or not isinstance(old_sb, _StyledScrollBar):
                        lv.setVerticalScrollBar(_StyledScrollBar(lv))
                except Exception:
                    pass

                # Tablet: стилизуем popup QCompleter (например, product_combo при вводе текста)
                try:
                    comp = combo.completer() if hasattr(combo, "completer") else None
                    if comp is not None and comp.completionMode() == QCompleter.PopupCompletion:
                        cp = comp.popup()
                        if cp is not None:
                            # Защита от "popup поверх других приложений":
                            # явно ставим флаги Qt.Popup + FramelessWindowHint.
                            # WindowStaysOnTopHint НЕ ставится.
                            try:
                                cp.setWindowFlags(
                                    Qt.Popup | Qt.FramelessWindowHint | Qt.NoDropShadowWindowHint
                                )
                            except Exception:
                                pass
                            # тот же QSS, что у основного view
                            cp.setStyleSheet(_lv_qss)
                            cp.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
                            cp.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
                            # шрифт того же размера
                            cf = cp.font()
                            cf.setPointSize(max(cf.pointSize(), 12))
                            cp.setFont(cf)
                            # тот же делегат с min_h=54, чтобы строки были крупными
                            cp.setItemDelegate(_TabletComboItemDelegate(cp, min_h=54))
                            # кастомный жёлтый скроллбар
                            old_sb = cp.verticalScrollBar()
                            if old_sb is not None and not isinstance(old_sb, _StyledScrollBar):
                                cp.setVerticalScrollBar(_StyledScrollBar(cp))
                            # Event filter: переустанавливаем _StyledScrollBar каждый раз
                            # при показе popup'а completer'а (Qt может пересоздавать scrollbar
                            # при пересчёте модели после фильтрации).
                            class _CompleterPopupFilter(QObject):
                                """
                                Отслеживает Show/Hide событий popup'а QCompleter.
                                - Show: переустанавливает кастомный _StyledScrollBar.
                                - Hide: пишет timestamp на combo.
                                """
                                def __init__(self, popup, combo, parent=None):
                                    super().__init__(parent)
                                    self._popup = popup
                                    self._combo = combo
                                def eventFilter(self, obj, event):
                                    if obj is self._popup:
                                        # Блокируем закрытие popup'а от focus loss / window deactivation
                                        # когда активным становится окно сенсорной клавиатуры Windows.
                                        if event.type() in (QEvent.FocusOut, QEvent.WindowDeactivate):
                                            try:
                                                if _is_foreground_window_tabtip():
                                                    return True  # eat — popup не реагирует на OSK
                                            except Exception:
                                                pass
                                            # FocusOut от обычной потери фокуса (другой виджет, USB-клава)
                                            # — пропускаем дальше, popup закроется естественно.
                                            if event.type() == QEvent.FocusOut:
                                                return True
                                        if event.type() == QEvent.Hide:
                                            try:
                                                import time as _t
                                                self._combo._completer_popup_last_hide_ts = _t.time()
                                            except Exception:
                                                pass
                                            try:
                                                if not getattr(self._combo, "_explicit_close", False):
                                                    # неявное закрытие — отменяем
                                                    combo_ref = self._combo
                                                    def _undo_hide():
                                                        try:
                                                            le = combo_ref.lineEdit()
                                                            if le is None:
                                                                return
                                                            text = le.text().strip()
                                                            if not text:
                                                                return
                                                            # Если за это время пользователь
                                                            # явно кликнул в другой виджет — не возвращаем.
                                                            if getattr(combo_ref, "_explicit_close", False):
                                                                combo_ref._explicit_close = False
                                                                return
                                                            comp = combo_ref.completer()
                                                            if comp is None:
                                                                return
                                                            comp.setCompletionPrefix(text)
                                                            comp.complete()
                                                        except Exception:
                                                            pass
                                                    QTimer.singleShot(0, _undo_hide)
                                            except Exception:
                                                pass
                                            # Не возвращаем True (Hide нельзя есть), просто планируем re-open сразу
                                        if event.type() == QEvent.Show:
                                            try:
                                                sb = self._popup.verticalScrollBar()
                                                if sb is not None and not isinstance(sb, _StyledScrollBar):
                                                    self._popup.setVerticalScrollBar(_StyledScrollBar(self._popup))
                                            except Exception:
                                                pass
                                            # Принудительно синхронизируем prefix completer'а
                                            # с текстом lineEdit'а — это страхует от случаев,
                                            # когда Qt по какой-то причине открывает completer
                                            # с пустым/устаревшим prefix'ом (отсюда "полный список").
                                            try:
                                                comp_here = self._combo.completer()
                                                le_here = self._combo.lineEdit()
                                                if comp_here is not None and le_here is not None:
                                                    cur_text = le_here.text().strip()
                                                    if cur_text and comp_here.completionPrefix() != cur_text:
                                                        comp_here.setCompletionPrefix(cur_text)
                                            except Exception:
                                                pass
                                    return False

                            if not hasattr(self, "_completer_popup_filters"):
                                self._completer_popup_filters = []
                            flt = _CompleterPopupFilter(cp, combo, cp)
                            cp.installEventFilter(flt)
                            self._completer_popup_filters.append(flt)
                            # тач-скролл для completer popup
                            cp.viewport().setAttribute(Qt.WA_AcceptTouchEvents, True)
                            scroller = QScroller.scroller(cp.viewport())
                            if scroller:
                                props = scroller.scrollerProperties()
                                props.setScrollMetric(QScrollerProperties.ScrollingCurve, QEasingCurve.OutQuad)
                                props.setScrollMetric(QScrollerProperties.AxisLockThreshold, 0.0)
                                scroller.setScrollerProperties(props)
                            QScroller.grabGesture(cp.viewport(), QScroller.LeftMouseButtonGesture)
                except Exception:
                    pass

                # Убираем системную рамку/тень контейнера popup (tablet)
                try:
                    popup = lv.window()
                    if popup is not None:
                        popup.setWindowFlags(
                            Qt.Popup | Qt.FramelessWindowHint | Qt.NoDropShadowWindowHint
                        )
                        # WA_TranslucentBackground на Windows может давать чёрный flash
                        popup.setStyleSheet(
                            "background: #ffffff; border: 1.5px solid #cbd5e1; "
                            "border-radius: 13px; padding: 0px;"
                        )
                except Exception:
                    pass

                if (
                    combo is self.made_combo
                    or combo is self.checked_combo
                    or combo is self.font_combo
                    or combo is self.unit_combo
                ):
                    if not hasattr(self, "_tablet_combo_preshow_filters"):
                        self._tablet_combo_preshow_filters = []
                    pre = _TabletComboPreShow(_apply_popup_geometry, combo, lv)
                    combo.installEventFilter(pre)
                    self._tablet_combo_preshow_filters.append(pre)

                # Подстрахуем высоту строки через gridSize (некоторые стили игнорируют min-height).
                try:
                    row_h = 54
                    lv.setGridSize(QSize(0, row_h))
                except Exception:
                    pass

                # Строго ограничиваем popup по числу видимых строк именно на показе popup
                # (важно: списки могут заполняться/обновляться после init_ui).
                try:
                    popup = lv.window()
                    if popup is not None:
                        if not hasattr(self, "_tablet_combo_popup_sizers"):
                            self._tablet_combo_popup_sizers = []
                        sizer = _TabletPopupSizer(combo, lv)
                        popup.installEventFilter(sizer)
                        self._tablet_combo_popup_sizers.append(sizer)
                        _apply_popup_geometry(combo, lv)
                except Exception:
                    pass
            except Exception:
                pass

    def _apply_tablet_ui_tweaks(self):
        """Tablet-only: точечные правки размеров кнопок без перестройки layout."""
        # font_size_combo: шире закрытое поле (было 90), чтобы 20/22/24 не обрезались у стрелки
        try:
            if hasattr(self, "font_size_combo"):
                self.font_size_combo.setFixedWidth(108)
        except Exception:
            pass

        # Кнопки +/- рядом с полем количества (возле "Ед. изм.") — крупнее для тач-ввода
        try:
            if hasattr(self, "minus_btn") and hasattr(self, "qty_input") and hasattr(self, "plus_btn"):
                h = 96      # было 72, делаем выше для тач-удобства
                w_btn = 120  # было 72, делаем шире (длиннее по горизонтали)
                self.minus_btn.setFixedWidth(w_btn)
                self.plus_btn.setFixedWidth(w_btn)
                self.minus_btn.setMinimumHeight(h)
                self.plus_btn.setMinimumHeight(h)

                # поле ввода количества: шире на ~20% и выше на ~10%
                self.qty_input.setMinimumWidth(168)
                self.qty_input.setMinimumHeight(int(round(h * 1.1)))
                self.qty_input.setAlignment(Qt.AlignCenter)
                f_qty = self.qty_input.font()
                # Явно задаём крупный размер для tablet (не от текущего, чтобы рост был заметным)
                f_qty.setPointSize(28)
                f_qty.setWeight(QFont.DemiBold)
                self.qty_input.setFont(f_qty)
                self.qty_input.setStyleSheet(
                    "QLineEdit { font-size: 28px; font-weight: 600; padding: 4px 10px; }"
                )
        except Exception:
            pass

        # «Взвесить» — уже компактнее по ширине, выше для тач-нажатия
        try:
            if hasattr(self, "scale_btn") and self.scale_btn is not None:
                self.scale_btn.setMinimumHeight(90)
                self.scale_btn.setMaximumHeight(90)
                self.scale_btn.setMinimumWidth(0)
                self.scale_btn.setMaximumWidth(220)
        except Exception:
            pass

        # Кнопки toolbar предпросмотра — выровнять по высоте
        toolbar_h = 52
        for w in (
            getattr(self, "btn_font_minus", None),
            getattr(self, "btn_font_plus", None),
            getattr(self, "btn_bold", None),
            getattr(self, "btn_italic", None),
            getattr(self, "btn_underline", None),
            getattr(self, "btn_align_left", None),
            getattr(self, "btn_align_center", None),
            getattr(self, "btn_align_right", None),
        ):
            if w is not None:
                try:
                    w.setMinimumHeight(toolbar_h)
                except Exception:
                    pass

        # Нижний блок: сделать "-" и "+" для количества ближе к квадратным,
        # за счёт ограничения ширины «Печать»
        try:
            if hasattr(self, "copies_minus") and hasattr(self, "copies_plus"):
                sq = 163  # удвоенная ширина относительно квадрата 68×68
                self.copies_minus.setFixedWidth(sq)
                self.copies_plus.setFixedWidth(sq)
        except Exception:
            pass

        try:
            if hasattr(self, "print_btn"):
                self.print_btn.setMaximumWidth(208)
        except Exception:
            pass

        # Чекбоксы "Ручной ввод" — чуть крупнее зона попадания
        try:
            for cb in (getattr(self, "made_manual", None), getattr(self, "checked_manual", None)):
                if cb is not None:
                    cb.setStyleSheet(
                        "QCheckBox { padding: 6px 2px; font-size: 15px; }"
                        "QCheckBox::indicator { width: 22px; height: 22px; }"
                    )
        except Exception:
            pass

        # Кнопки перелистывания истории — чуть крупнее
        try:
            for b in (getattr(self, "history_prev_btn", None), getattr(self, "history_next_btn", None)):
                if b is not None:
                    b.setMinimumHeight(52)
                    b.setFixedWidth(64)
        except Exception:
            pass

    def _make_align_icon(self, mode):
        pix = QPixmap(26, 18)
        pix.fill(Qt.transparent)

        painter = QPainter(pix)
        painter.setRenderHint(QPainter.Antialiasing, False)

        color = QColor("#111827")
        bar_h = 2
        gaps = [2, 8, 14]

        if mode == "left":
            widths = [16, 20, 14]
            xs = [3, 3, 3]
        elif mode == "center":
            widths = [16, 20, 14]
            xs = [(26 - w) // 2 for w in widths]
        elif mode == "right":
            widths = [16, 20, 14]
            xs = [26 - w - 3 for w in widths]
        else:
            widths = [18, 18, 18]
            xs = [4, 4, 4]

        for x, y, w in zip(xs, gaps, widths):
            painter.fillRect(x, y, w, bar_h, color)

        painter.end()
        return QIcon(pix)

    def _make_stats_icon(self) -> QIcon:
        size = 18
        pix = QPixmap(size, size)
        pix.fill(Qt.transparent)

        painter = QPainter(pix)
        painter.setRenderHint(QPainter.Antialiasing, True)
        painter.setPen(Qt.NoPen)

        colors = [
            QColor("#f97316"),  # красно-оранжевый
            QColor("#f59e0b"),  # оранжевый
            QColor("#facc15"),  # жёлтый
            QColor("#22c55e"),  # зелёный
        ]
        heights = [6, 9, 12, 15]

        bar_w = 3
        gap = 1
        left = 2
        bottom = size - 2
        radius = 1.2

        for i, (c, h) in enumerate(zip(colors, heights)):
            x = left + i * (bar_w + gap)
            y = bottom - h
            painter.setBrush(c)
            painter.drawRoundedRect(QRectF(x, y, bar_w, h), radius, radius)

        painter.end()
        return QIcon(pix)
    def _load_logo(self):
        if os.path.exists(LOGO_PATH):
            pix = QPixmap(LOGO_PATH)
            if not pix.isNull():
                target_h = 56
                dpr = self.devicePixelRatioF() if hasattr(self, 'devicePixelRatioF') else 2.0
                scaled = pix.scaledToHeight(
                    int(target_h * dpr),
                    Qt.SmoothTransformation,
                )
                scaled.setDevicePixelRatio(dpr)
                self.logo.setPixmap(scaled)
                return
        self.logo.setText("")

    def _lock_window_size_against_osk(self):
        """
        Запоминает фактический размер окна после showMaximized() и блокирует
        дальнейшие ресайзы. Нужно, чтобы Windows-сенсорная клавиатура не
        поджимала окно по высоте — иначе layout схлопывается.
        """
        try:
            geom = self.geometry()
            self._locked_window_geom = geom
            # Минимальный размер = текущий, чтобы Windows не мог сделать окно меньше.
            # Максимальный не ставим (вдруг юзер развернёт между мониторами).
            self.setMinimumSize(geom.width(), geom.height())
        except Exception:
            pass

    def resizeEvent(self, event):
        super().resizeEvent(event)
        try:
            locked = getattr(self, "_locked_window_geom", None)
            if locked is not None:
                cur = self.geometry()
                if cur.width() < locked.width() or cur.height() < locked.height():
                    QTimer.singleShot(0, lambda: self.setGeometry(locked))
        except Exception:
            pass
        self._resize_label_preview()
        if hasattr(self, "history_panel"):
            self.history_panel.setVisible(self.width() >= 1100)

    def showEvent(self, event):
        super().showEvent(event)
        try:
            self._update_screen_compensation_scale()
        except Exception:
            pass

    def _effective_preview_font_scale(self):
        """Множитель pt в документе: окно (_preview_scale) × компенсация экрана."""
        a = float(getattr(self, "_preview_scale", 1.0) or 1.0)
        b = float(getattr(self, "_screen_compensation_scale", 1.0) or 1.0)
        return max(0.4, min(1.5, a * b))

    def _get_screen_compensation_scale(self):
        """
        Практический коэффициент < 1 на high-DPI / масштабировании Windows; ≈1 на эталонном 96 DPI.
        Не использует диагональ — только logical/physical DPI и DPR экрана Qt.
        """
        screen = None
        try:
            wh = self.windowHandle()
            if wh is not None:
                screen = wh.screen()
        except Exception:
            pass
        if screen is None:
            app = QApplication.instance()
            if app is not None:
                screen = app.primaryScreen()

        if screen is None:
            self._screen_comp_last_meta = {}
            return 1.0

        ldpi = float(screen.logicalDotsPerInchX())
        pdpi = float(screen.physicalDotsPerInchX())
        dpr = float(screen.devicePixelRatio())
        name = screen.name()
        geom = screen.geometry()

        ref = float(SCREEN_COMP_REF_LOGICAL_DPI)
        ratio = ref / max(ldpi, 72.0)
        ratio = min(ratio, 1.0)
        comp = 1.0 + (ratio - 1.0) * float(SCREEN_COMP_LDPI_BLEND)
        if dpr > 1.01:
            comp *= max(0.88, 1.0 - (dpr - 1.0) * float(SCREEN_COMP_DPR_EXTRA))
        comp *= float(SCREEN_COMP_MANUAL_MULTIPLIER)
        comp = max(float(SCREEN_COMP_MIN), min(1.0, comp))

        self._screen_comp_last_meta = {
            "name": name,
            "logical_dpi_x": ldpi,
            "physical_dpi_x": pdpi,
            "device_pixel_ratio": dpr,
            "geometry": geom,
            "compensation": comp,
        }
        return comp

    def _update_screen_compensation_scale(self):
        new = float(self._get_screen_compensation_scale())
        old = float(getattr(self, "_screen_compensation_scale", 1.0) or 1.0)
        self._screen_compensation_scale = new

        meta = getattr(self, "_screen_comp_last_meta", {}) or {}
        sig = (
            meta.get("name"),
            round(float(meta.get("logical_dpi_x", 0)), 2),
            round(float(meta.get("physical_dpi_x", 0)), 2),
            round(float(meta.get("device_pixel_ratio", 0)), 3),
            meta.get("geometry"),
            round(new, 4),
        )
        if sig != getattr(self, "_last_screen_comp_sig", None):
            self._last_screen_comp_sig = sig
            print(
                "[screen_comp]",
                f"screen={meta.get('name')!r}",
                f"logicalDpiX={meta.get('logical_dpi_x')}",
                f"physicalDpiX={meta.get('physical_dpi_x')}",
                f"dpr={meta.get('device_pixel_ratio')}",
                f"geometry={meta.get('geometry')}",
                f"compensation={meta.get('compensation')}",
                sep=" | ",
            )

        if abs(new - old) > 1e-5:
            if not getattr(self, "_user_edited_preview", False) and not getattr(
                self, "_preview_manual_mode", False
            ):
                try:
                    self.refresh_preview(force=True)
                except Exception:
                    pass

    def _logical_preview_size(self):
        """Стабильный логический размер предпросмотра (тот же эталон, что и PRINT_VIRTUAL_W в печати)."""
        w_mm = float(getattr(self, "label_w_mm", 58.0) or 58.0)
        h_mm = float(getattr(self, "label_h_mm", 80.0) or 80.0)
        if w_mm <= 1e-9:
            w_mm = 58.0
        ref_w = 450
        ref_h = int(round(ref_w * (h_mm / w_mm)))
        return ref_w, max(180, ref_h)

    def _fit_label_preview_graphics(self):
        """Вписывает логическую этикетку в область view; масштаб ≤ 1, шрифты документа не трогаем."""
        if not getattr(self, "preview_view", None):
            return
        lw, lh = self._logical_preview_size()
        if hasattr(self, "_preview_scene"):
            self._preview_scene.setSceneRect(0.0, 0.0, float(lw), float(lh))
        if hasattr(self, "_preview_proxy"):
            self._preview_proxy.setPos(0.0, 0.0)

        vw = max(2, self.preview_view.viewport().width())
        vh = max(2, self.preview_view.viewport().height())
        s = min(vw / float(lw), vh / float(lh))
        s = min(float(s), 1.0)

        self.preview_view.resetTransform()
        if s > 0:
            self.preview_view.scale(s, s)
        self.preview_view.centerOn(lw * 0.5, lh * 0.5)

    def _resize_label_preview(self):
        if not hasattr(self, "preview_wrap"):
            return

        self._update_screen_compensation_scale()

        # Документ: фиксированный логический размер области; масштаб шрифта — _effective_preview_font_scale().
        self._preview_scale = 1.0

        lw, lh = self._logical_preview_size()
        self.preview.setFixedSize(lw, lh)

        if hasattr(self, "_preview_scene"):
            self._preview_scene.setSceneRect(0.0, 0.0, float(lw), float(lh))
        if hasattr(self, "_preview_proxy"):
            self._preview_proxy.setPos(0.0, 0.0)

        QTimer.singleShot(0, self._fit_label_preview_graphics)
    def _on_label_size_changed(self, index):
        sizes = {0: (58.0, 80.0), 1: (58.0, 60.0), 2: (70.0, 70.0), 3: (70.0, 70.0)}
        self.label_w_mm, self.label_h_mm = sizes.get(index, (58.0, 80.0))
        self._resize_label_preview()

        self._user_edited_preview = False
        text, can_print = self._build_label_plain_text()
        self._refresh_print_btn_state()
        self._set_preview_text_programmatically(text)
        QApplication.processEvents()

    def _toggle_preview_manual_mode(self):
        self._preview_manual_mode = not self._preview_manual_mode
        on = self._preview_manual_mode
        self.manual_datetime_label.setVisible(on)
        self.manual_datetime_picker.setVisible(on)
        self._apply_preview_header_style(on)
        self._set_made_checked_blocks_visible(not on)
        self.preview.setReadOnly(not on)
        self._user_edited_preview = False
        self.refresh_preview()

    def _set_made_checked_blocks_visible(self, visible: bool):
        """
        Скрывает или показывает блоки 'Изготовил' и 'Цех' целиком (лейблы +
        комбобоксы + чекбоксы 'Ручной ввод' + поля ручного ввода).
        При показе видимость combo/input восстанавливается по состоянию
        соответствующих чекбоксов 'Ручной ввод'.
        """
        if not visible:
            # Скрываем всё — режим редактирования preview
            for name in ("lab_made", "made_combo", "made_manual", "made_input",
                         "lab_chk", "checked_combo", "checked_manual", "checked_input"):
                w = getattr(self, name, None)
                if w is not None:
                    w.setVisible(False)
            return

        # Возвращаем блоки в обычный режим:
        # - лейблы и чекбоксы 'Ручной ввод' всегда видимы
        # - combo/input — по состоянию чекбокса
        if hasattr(self, "lab_made") and self.lab_made is not None:
            self.lab_made.setVisible(True)
        if hasattr(self, "made_manual") and self.made_manual is not None:
            self.made_manual.setVisible(True)
        made_manual_on = bool(self.made_manual.isChecked()) if hasattr(self, "made_manual") else False
        if hasattr(self, "made_combo") and self.made_combo is not None:
            self.made_combo.setVisible(not made_manual_on)
        if hasattr(self, "made_input") and self.made_input is not None:
            self.made_input.setVisible(made_manual_on)

        if hasattr(self, "lab_chk") and self.lab_chk is not None:
            self.lab_chk.setVisible(True)
        if hasattr(self, "checked_manual") and self.checked_manual is not None:
            self.checked_manual.setVisible(True)
        checked_manual_on = bool(self.checked_manual.isChecked()) if hasattr(self, "checked_manual") else False
        if hasattr(self, "checked_combo") and self.checked_combo is not None:
            self.checked_combo.setVisible(not checked_manual_on)
        if hasattr(self, "checked_input") and self.checked_input is not None:
            self.checked_input.setVisible(checked_manual_on)

    def _apply_preview_header_style(self, manual_mode: bool):
        """
        Переключает внешний вид заголовка над предпросмотром:
        - normal: серая плашка с текстом "Предпросмотр" (как обычно, стиль из глобального QSS)
        - manual_mode: фиолетовая плашка с текстом "Редактирование"
        """
        if manual_mode:
            self.preview_header.setText("Редактирование")
            self.preview_header.setStyleSheet(
                "#SectionTitle {"
                " background: #8b5cf6;"
                " border-radius: 14px;"
                " padding: 10px 22px;"
                " font-family: 'Inter','Segoe UI','Manrope','Arial',sans-serif;"
                " font-size: 22px;"
                " font-weight: 650;"
                " letter-spacing: 0.2px;"
                " color: #ffffff;"
                "}"
            )
        else:
            self.preview_header.setText("Предпросмотр")
            # Сброс локального QSS — возвращается стиль из глобального QSS (#SectionTitle)
            self.preview_header.setStyleSheet("")

    def _resolve_sheet_names(self, excel_path, silent: bool = False):
        """
        Проверяет, что в файле есть все 3 обязательных листа.
        Возвращает (sheet_products, sheet_made, sheet_checked) или None если файл не подходит.
        """
        wb = _peek_workbook(excel_path, read_only=True, data_only=True)
        available = wb.sheetnames
        wb.close()

        # проверяем наличие каждого листа
        sheet_products = SHEET_PRODUCTS if SHEET_PRODUCTS in available else ("products" if "products" in available else None)
        sheet_made = SHEET_MADE if SHEET_MADE in available else None
        sheet_checked = SHEET_CHECKED if SHEET_CHECKED in available else None

        missing = []
        if not sheet_products:
            missing.append(f"«{SHEET_PRODUCTS}»")
        if not sheet_made:
            missing.append(f"«{SHEET_MADE}»")
        if not sheet_checked:
            missing.append(f"«{SHEET_CHECKED}»")

        if missing:
            available_str = ", ".join([f"«{s}»" for s in available]) if available else "(пусто)"
            if not silent:
                QMessageBox.warning(
                    self,
                    "Неподходящий файл",
                    f"В выбранном файле отсутствуют обязательные листы:\n"
                    f"{', '.join(missing)}\n\n"
                    f"Листы в файле: {available_str}\n\n"
                    f"Файл Excel должен содержать 3 листа:\n\n"
                    f"1. «{SHEET_PRODUCTS}» — список продукции\n"
                    f"   Заголовки: Код | Наименование | Срок годности (ч) | Ед. измер. | Активен\n\n"
                    f"2. «{SHEET_MADE}» — сотрудники\n"
                    f"   Заголовки: ФИО | Активен\n\n"
                    f"3. «{SHEET_CHECKED}» — цехи\n"
                    f"   Заголовки: Цех | Активен\n\n"
                    f"Первая строка каждого листа — заголовки, данные со второй строки.",
                )
            return None

        return (sheet_products, sheet_made, sheet_checked)

    def reload_excel(self, show_message=True, silent_errors: bool = False):
        try:
            current_product = self.product_combo.currentText().strip()
            current_unit = self.unit_combo.currentText()
            current_qty = self.qty_input.text().strip()

            made_manual = self.made_manual.isChecked()
            checked_manual = self.checked_manual.isChecked()
            made_text = self.made_input.text().strip()
            checked_text = self.checked_input.text().strip()
            made_combo = self.made_combo.currentText()
            checked_combo = self.checked_combo.currentText()

            # === Этап C: чтение данных из всех активных Excel-файлов ===
            self._ensure_default_excel_source_present()

            sources = self._load_excel_sources()
            active_paths = [s["path"] for s in sources if s.get("active") and os.path.isfile(s["path"])]
            missing_files = [s["path"] for s in sources if s.get("active") and not os.path.isfile(s["path"])]

            # Fallback: если список активных пуст — используем self.excel_path как раньше
            if not active_paths:
                if os.path.isfile(self.excel_path):
                    active_paths = [self.excel_path]
                else:
                    self.products = []
                    self.staff_made = []
                    self.staff_checked = []
                    if not silent_errors:
                        QMessageBox.warning(
                            self,
                            "Нет активных файлов",
                            "В приложении нет активных Excel-файлов с данными.\n\n"
                            "Нажмите «Добавить», чтобы добавить файл, "
                            "затем включите его в «Активные».",
                        )
                    self.fill_products(current_product)
                    self.fill_staff()
                    return

            all_products = []
            all_made = []
            all_checked = []
            file_errors = []  # [(path, error_message), ...]

            for path in active_paths:
                try:
                    resolved = self._resolve_sheet_names(path, silent=True)
                    if resolved is None:
                        file_errors.append((path, "В файле нет обязательных листов"))
                        continue
                    sheet_products, sheet_made, sheet_checked = resolved

                    if sheet_products:
                        products_all = load_products(path, sheet_name=sheet_products)
                        active_products = [p for p in products_all if int(p.get("active", 0)) == 1]
                        for p in active_products:
                            if isinstance(p, dict):
                                if "name" in p:
                                    p["name"] = _repair_mojibake_utf8_as_cp1251(p.get("name") or "")
                                if "comment" in p:
                                    p["comment"] = _repair_mojibake_utf8_as_cp1251(p.get("comment") or "")
                        all_products.append(active_products)

                    if sheet_made:
                        made_list = [s for s in load_staff(path, sheet_made) if int(s.get("active", 0)) == 1]
                        all_made.append(made_list)

                    if sheet_checked:
                        checked_list = [s for s in load_staff(path, sheet_checked) if int(s.get("active", 0)) == 1]
                        all_checked.append(checked_list)
                except Exception as e:
                    file_errors.append((path, str(e)))

            # Слияние с дедупликацией
            self.products = self._merge_products(all_products)
            self.products.sort(key=lambda x: (x.get("name") or "").lower())

            self.staff_made = self._merge_staff(all_made)
            self.staff_checked = self._merge_staff(all_checked)

            # Сообщения о проблемных файлах (только если несколько активных
            # или если показ запрошен явно)
            if not silent_errors:
                problems = []
                for p in missing_files:
                    problems.append(f"• {p}\n  данные по файлу отсутствуют (файл не найден на диске)")
                for p, err in file_errors:
                    problems.append(f"• {p}\n  данные по файлу отсутствуют: {err}")
                if problems and (show_message or len(active_paths) > 1):
                    QMessageBox.warning(
                        self,
                        "Проблемы с некоторыми файлами",
                        "Часть Excel-файлов не удалось прочитать:\n\n" + "\n\n".join(problems),
                    )

            self.staff_made = [
                {
                    "fio": _repair_mojibake_utf8_as_cp1251((x.get("fio") or x.get("name") or "")).strip(),
                    "active": x.get("active", 1),
                }
                for x in self.staff_made
            ]
            self.staff_checked = [
                {
                    "fio": _repair_mojibake_utf8_as_cp1251((x.get("fio") or x.get("name") or "")).strip(),
                    "active": x.get("active", 1),
                }
                for x in self.staff_checked
            ]

            self.staff_made.sort(key=lambda x: (x.get("fio") or "").lower())
            self.staff_checked.sort(key=lambda x: (x.get("fio") or "").lower())

            self.fill_products(current_product)
            self.fill_staff()

            self.loaded_at_str = datetime.now().strftime("%d.%m.%Y %H:%M:%S")
            self.update_excel_status()

            if current_qty:
                self.qty_input.setText(current_qty)

            self.made_manual.setChecked(made_manual)
            self.checked_manual.setChecked(checked_manual)
            self.made_input.setText(made_text)
            self.checked_input.setText(checked_text)

            if made_combo and not made_combo.startswith("—"):
                idx = self.made_combo.findText(made_combo)
                if idx >= 0:
                    self.made_combo.setCurrentIndex(idx)
            if checked_combo and not checked_combo.startswith("—"):
                idx = self.checked_combo.findText(checked_combo)
                if idx >= 0:
                    self.checked_combo.setCurrentIndex(idx)

            if current_product:
                self.on_product_changed(current_product)
                if current_unit:
                    idxu = self.unit_combo.findText(current_unit)
                    if idxu >= 0:
                        self.unit_combo.setCurrentIndex(idxu)
            else:
                self.unit_combo.clear()
                # не добавляем пустой пункт — в списке должны быть только реальные единицы

            if show_message and not silent_errors:
                QMessageBox.information(self, "Готово", "Excel обновлён.")

        except Exception as e:
            if silent_errors:
                sys.stderr.write(f"[MirlisMark] Excel auto-refresh error: {e}\n")
                return
            QMessageBox.critical(
                self,
                "Ошибка Excel",
                f"Не удалось загрузить Excel.\n\nФайл: {self.excel_path}\nОшибка: {e}",
            )

    def fill_products(self, current_product=None):
        self.product_combo.blockSignals(True)
        self.product_combo.clear()

        names = [p["name"] for p in self.products if p.get("name")]
        for n in names:
            self.product_combo.addItem(n)

        self.product_model.setStringList(names)

        self.product_combo.setCurrentIndex(-1)
        self.product_combo.setEditText("")

        if current_product:
            idx = self.product_combo.findText(current_product)
            if idx >= 0:
                self.product_combo.setCurrentIndex(idx)
            else:
                self.product_combo.setEditText(current_product)

        self.product_combo.blockSignals(False)

    def fill_staff(self):
        self.made_combo.blockSignals(True)
        self.made_combo.clear()
        self.made_combo.addItem("— не выбрано —")
        for s in self.staff_made:
            fio = (s.get("fio") or "").strip()
            if fio:
                self.made_combo.addItem(fio)
        # если только один сотрудник — выбираем автоматически
        if self.made_combo.count() == 2:
            self.made_combo.setCurrentIndex(1)
        self.made_combo.blockSignals(False)

        self.checked_combo.blockSignals(True)
        self.checked_combo.clear()
        self.checked_combo.addItem("— не выбрано —")
        for s in self.staff_checked:
            fio = (s.get("fio") or "").strip()
            if fio:
                self.checked_combo.addItem(fio)
        # если только один цех — выбираем автоматически
        if self.checked_combo.count() == 2:
            self.checked_combo.setCurrentIndex(1)
        self.checked_combo.blockSignals(False)

    def update_excel_status(self):
        try:
            mtime = os.path.getmtime(self.excel_path)
            mtime_str = _fmt_dt_local(mtime)
        except Exception:
            mtime_str = "—"

        default_dir = app_data_dir()
        current_dir = os.path.dirname(self.excel_path)
        if os.path.normpath(current_dir) != os.path.normpath(default_dir):
            short_path = current_dir
            if len(short_path) > 40:
                short_path = "..." + short_path[-37:]
            self.excel_pill.setText(f"Excel: {short_path}\nобновлено {mtime_str}")
        else:
            self.excel_pill.setText(f"Excel: обновлено {mtime_str}")

    def open_excel_folder(self):
        folder = os.path.dirname(self.excel_path)
        QDesktopServices.openUrl(QUrl.fromLocalFile(folder))

    def open_labels_folder(self):
        """Открыть папку с архивом напечатанных этикеток (%LOCALAPPDATA%\\MirlisMark\\Готовые этикетки\\)."""
        folder = self._labels_archive_root()
        os.makedirs(folder, exist_ok=True)
        QDesktopServices.openUrl(QUrl.fromLocalFile(folder))

    # =================== Excel sources management (Этап B) ===================
    # Хранится в settings.json как массив:
    #   "excel_sources": [
    #       {"path": "...", "active": true, "added_at": 1715800000.0}, ...
    #   ]
    # Логика загрузки данных пока работает с одним первым активным файлом
    # (см. _current_active_excel_path); полноценное слияние будет в Этапе C.

    REQUIRED_EXCEL_SHEETS = ("продукт", "изготовил", "цех")

    def _load_excel_sources(self) -> list:
        """Прочитать список Excel-источников из settings.json."""
        settings = _load_settings()
        sources = settings.get("excel_sources")
        if not isinstance(sources, list):
            return []
        # Нормализация: оставляем только валидные записи
        result = []
        for s in sources:
            if not isinstance(s, dict):
                continue
            path = str(s.get("path") or "").strip()
            if not path:
                continue
            result.append({
                "path": path,
                "active": bool(s.get("active", True)),
                "added_at": float(s.get("added_at") or 0.0),
            })
        return result

    def _save_excel_sources(self, sources: list):
        """Сохранить список Excel-источников в settings.json."""
        settings = _load_settings()
        settings["excel_sources"] = sources
        _save_settings(settings)

    def _validate_excel_file(self, path: str) -> tuple:
        """Проверить, что файл существует и содержит все три обязательных листа.
        Возвращает (ok, error_message). Если ok=True, error_message=''."""
        if not path:
            return (False, "Путь к файлу пустой.")
        if not os.path.isfile(path):
            return (False, f"Файл не найден:\n{path}")
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            sheet_names = [s.strip().lower() for s in wb.sheetnames]
            wb.close()
        except Exception as e:
            return (False, f"Не удалось открыть Excel-файл:\n{e}")
        missing = []
        for required in self.REQUIRED_EXCEL_SHEETS:
            if required.lower() not in sheet_names:
                missing.append(f"«{required}»")
        if missing:
            return (False, "В выбранном файле нет обязательных листов: " + ", ".join(missing) + ".\nФайл не будет добавлен.")
        return (True, "")

    def _current_active_excel_path(self):
        """Вернуть путь первого активного Excel-файла из списка (для совместимости
        со старой логикой, которая работает с одним файлом). На Этапе C это будет
        заменено на слияние данных из всех активных файлов."""
        for s in self._load_excel_sources():
            if s.get("active") and os.path.isfile(s.get("path", "")):
                return s["path"]
        return None

    def _show_add_excel_help_dialog(self) -> bool:
        """Показать инструкцию по структуре Excel перед добавлением файла.
        Возвращает True если пользователь нажал «Выбрать файл», иначе False."""
        dlg = QDialog(self)
        dlg.setWindowTitle("Структура файла Excel")
        dlg.setMinimumSize(700, 500)
        dlg.resize(750, 600)

        dlg_layout = QVBoxLayout(dlg)
        dlg_layout.setContentsMargins(0, 0, 0, 12)
        dlg_layout.setSpacing(0)

        scroll = QScrollArea()
        QScroller.grabGesture(scroll.viewport(), QScroller.TouchGesture)
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("QScrollArea { border: none; background: #ffffff; }")

        content = QWidget()
        content.setStyleSheet("background: #ffffff;")
        lay = QVBoxLayout(content)
        lay.setContentsMargins(28, 24, 28, 24)
        lay.setSpacing(16)

        def add_title(text):
            lbl = QLabel(text)
            lbl.setStyleSheet("font-size: 20px; font-weight: 800; color: #111827; background: transparent;")
            lbl.setWordWrap(True)
            lay.addWidget(lbl)

        def add_text(text):
            lbl = QLabel(text)
            lbl.setStyleSheet("font-size: 14px; color: #374151; background: transparent;")
            lbl.setWordWrap(True)
            lay.addWidget(lbl)

        def add_image(path, max_width=640):
            if not os.path.isfile(path):
                return
            lbl = QLabel()
            lbl.setStyleSheet("background: transparent;")
            pix = QPixmap(path)
            if not pix.isNull():
                if pix.width() > max_width:
                    pix = pix.scaledToWidth(max_width, Qt.SmoothTransformation)
                lbl.setPixmap(pix)
            lay.addWidget(lbl)

        def add_separator():
            sep = QFrame()
            sep.setFrameShape(QFrame.HLine)
            sep.setStyleSheet("color: #e5e7eb; background: #e5e7eb; border: none; max-height: 1px;")
            lay.addWidget(sep)

        # Hero
        hero_lbl = QLabel()
        hero_lbl.setStyleSheet("background: transparent;")
        hero_lbl.setAlignment(Qt.AlignHCenter)
        hero_pix = QPixmap(HAPPY_HERO_PATH)
        if not hero_pix.isNull():
            hero_pix = hero_pix.scaled(220, 220, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            hero_lbl.setPixmap(hero_pix)
            lay.addWidget(hero_lbl)

        add_title("Структура файла Excel")
        add_text(
            "Файл должен содержать 3 листа с точными названиями: «продукт», «изготовил», «цех».\n"
            "Первая строка каждого листа — заголовки колонок, данные начинаются со второй строки."
        )

        add_separator()
        add_text("Внизу файла должны быть видны 3 вкладки:")
        add_image(HELP_IMG_TABS)

        add_separator()
        add_title("1. Лист «продукт»")
        add_text(
            "Заголовки: Код | Наименование | Срок годности (ч) | Ед. измер. | Активен | Комментарий\n\n"
            "• «Ед. измер.» — через запятую: кг, шт или кг,шт\n"
            "• «Активен» — Да или 1 (показывать), Нет или 0 (скрыть)"
        )
        add_image(HELP_IMG_PRODUCT)

        add_separator()
        add_title("2. Лист «изготовил»")
        add_text("Заголовки: ФИО | Активен")
        add_image(HELP_IMG_MADE)

        add_separator()
        add_title("3. Лист «цех»")
        add_text("Заголовки: Цех | Активен")
        add_image(HELP_IMG_WORKSHOP)

        add_separator()
        add_text(
            "Если в листе «изготовил» или «цех» только одна запись — "
            "она подставится автоматически, и работнику не нужно будет её выбирать каждый раз."
        )

        scroll.setWidget(content)
        dlg_layout.addWidget(scroll, 1)

        # Кнопки внизу
        btn_row = QHBoxLayout()
        btn_row.setContentsMargins(20, 0, 20, 0)
        btn_row.setSpacing(12)
        btn_row.addStretch(1)

        btn_cancel = QPushButton("Отмена")
        btn_cancel.setMinimumHeight(40)
        btn_cancel.setCursor(Qt.PointingHandCursor)
        btn_cancel.clicked.connect(dlg.reject)

        btn_ok = QPushButton("Выбрать файл")
        btn_ok.setMinimumHeight(40)
        btn_ok.setCursor(Qt.PointingHandCursor)
        btn_ok.setStyleSheet(
            "QPushButton { background: #16a34a; color: #ffffff; font-weight: 700; "
            "border: 1px solid #15803d; border-radius: 12px; padding: 8px 24px; }"
            "QPushButton:hover { background: #15803d; }"
        )
        btn_ok.clicked.connect(dlg.accept)

        btn_row.addWidget(btn_cancel)
        btn_row.addWidget(btn_ok)
        dlg_layout.addLayout(btn_row)

        return dlg.exec_() == QDialog.Accepted

    def _on_add_excel(self):
        """Кнопка «Добавить»: выбрать Excel-файл, проверить листы, добавить в список."""
        # Сначала показываем инструкцию о структуре файла
        if not self._show_add_excel_help_dialog():
            return

        start_dir = os.path.dirname(self.excel_path) if getattr(self, "excel_path", "") else ""
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Выберите Excel-файл",
            start_dir,
            "Excel-файлы (*.xlsx *.xlsm);;Все файлы (*.*)",
        )
        if not path:
            return
        path = os.path.normpath(path)

        # Проверка на дубликат
        sources = self._load_excel_sources()
        for s in sources:
            if os.path.normcase(os.path.normpath(s["path"])) == os.path.normcase(path):
                QMessageBox.information(
                    self,
                    "Файл уже добавлен",
                    f"Этот файл уже есть в списке приложения:\n{path}",
                )
                return

        # Валидация листов
        ok, err = self._validate_excel_file(path)
        if not ok:
            QMessageBox.warning(self, "Файл не подходит", err)
            return

        # Добавить в список (активным по умолчанию)
        sources.append({
            "path": path,
            "active": True,
            "added_at": time.time(),
        })
        self._save_excel_sources(sources)

        QMessageBox.information(
            self,
            "Файл добавлен",
            f"Файл добавлен в список приложения и помечен активным:\n{path}\n\n"
            f"Нажмите «Обновить», чтобы перечитать данные.",
        )

    def _on_delete_excel(self):
        """Кнопка «Удалить»: показать список добавленных файлов с возможностью удаления.
        С диска НЕ удаляется."""
        sources = self._load_excel_sources()
        if not sources:
            QMessageBox.information(
                self,
                "Список пуст",
                "В приложении пока нет добавленных Excel-файлов.\n\nНажмите «Добавить», чтобы добавить файл.",
            )
            return

        dlg = QDialog(self)
        dlg.setWindowTitle("Удалить Excel-файл из списка")
        dlg.setMinimumSize(640, 380)

        lay = QVBoxLayout(dlg)
        lay.setContentsMargins(20, 20, 20, 16)
        lay.setSpacing(12)

        info = QLabel(
            "Выберите файл, который нужно убрать из списка приложения.\n"
            "Сам файл на диске НЕ удаляется."
        )
        info.setStyleSheet("font-size: 13px; color: #374151; background: transparent;")
        info.setWordWrap(True)
        lay.addWidget(info)

        list_widget = QListWidget()
        list_widget.setStyleSheet(
            "QListWidget { border: 1px solid #d1d5db; border-radius: 10px; background: #ffffff; }"
            "QListWidget::item { padding: 10px 12px; border-bottom: 1px solid #f3f4f6; }"
            "QListWidget::item:selected { background: #fee2e2; color: #991b1b; }"
        )
        for s in sources:
            text = s["path"]
            if not os.path.isfile(s["path"]):
                text = text + "   (файл не найден на диске)"
            item = QListWidgetItem(text)
            item.setData(Qt.UserRole, s["path"])
            list_widget.addItem(item)
        lay.addWidget(list_widget, 1)

        btn_row = QHBoxLayout()
        btn_row.addStretch(1)
        cancel_btn = QPushButton("Отмена")
        cancel_btn.setMinimumHeight(40)
        cancel_btn.clicked.connect(dlg.reject)
        btn_row.addWidget(cancel_btn)

        delete_btn = ActionBtn("Удалить из списка", kind="danger")
        delete_btn.setMinimumHeight(40)
        btn_row.addWidget(delete_btn)
        lay.addLayout(btn_row)

        def _do_delete():
            cur = list_widget.currentItem()
            if cur is None:
                QMessageBox.information(dlg, "Не выбран файл", "Выберите файл в списке.")
                return
            target_path = cur.data(Qt.UserRole)
            reply = QMessageBox.question(
                dlg,
                "Подтверждение",
                f"Убрать файл из списка приложения?\n\n{target_path}\n\nСам файл на диске НЕ удаляется.",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No,
            )
            if reply != QMessageBox.Yes:
                return
            new_sources = [s for s in self._load_excel_sources()
                           if os.path.normcase(os.path.normpath(s["path"])) !=
                              os.path.normcase(os.path.normpath(target_path))]
            self._save_excel_sources(new_sources)
            row = list_widget.row(cur)
            list_widget.takeItem(row)
            if list_widget.count() == 0:
                dlg.accept()

        delete_btn.clicked.connect(_do_delete)
        dlg.exec_()

    def _on_active_excel(self):
        """Кнопка «Активные»: чек-лист всех добавленных файлов с галочками."""
        sources = self._load_excel_sources()
        if not sources:
            QMessageBox.information(
                self,
                "Список пуст",
                "В приложении пока нет добавленных Excel-файлов.\n\nНажмите «Добавить», чтобы добавить файл.",
            )
            return

        dlg = QDialog(self)
        dlg.setWindowTitle("Активные Excel-файлы")
        dlg.setMinimumSize(640, 420)

        lay = QVBoxLayout(dlg)
        lay.setContentsMargins(20, 20, 20, 16)
        lay.setSpacing(12)

        info = QLabel(
            "Отметьте галочкой файлы, которые приложение должно использовать.\n"
            "После сохранения нажмите «Обновить» в главном окне, чтобы перечитать данные."
        )
        info.setStyleSheet("font-size: 13px; color: #374151; background: transparent;")
        info.setWordWrap(True)
        lay.addWidget(info)

        list_widget = QListWidget()
        list_widget.setStyleSheet(
            "QListWidget { border: 1px solid #d1d5db; border-radius: 10px; background: #ffffff; }"
            "QListWidget::item { padding: 8px 10px; border-bottom: 1px solid #f3f4f6; }"
        )
        for s in sources:
            text = s["path"]
            if not os.path.isfile(s["path"]):
                text = text + "   (файл не найден на диске)"
            item = QListWidgetItem(text)
            item.setFlags(item.flags() | Qt.ItemIsUserCheckable)
            item.setCheckState(Qt.Checked if s.get("active") else Qt.Unchecked)
            item.setData(Qt.UserRole, s["path"])
            list_widget.addItem(item)
        lay.addWidget(list_widget, 1)

        btn_row = QHBoxLayout()
        btn_row.addStretch(1)
        cancel_btn = QPushButton("Отмена")
        cancel_btn.setMinimumHeight(40)
        cancel_btn.clicked.connect(dlg.reject)
        btn_row.addWidget(cancel_btn)

        save_btn = ActionBtn("Сохранить", kind="primary")
        save_btn.setMinimumHeight(40)
        btn_row.addWidget(save_btn)
        lay.addLayout(btn_row)

        def _do_save():
            current_sources = self._load_excel_sources()
            by_path = {os.path.normcase(os.path.normpath(s["path"])): s
                       for s in current_sources}
            for i in range(list_widget.count()):
                it = list_widget.item(i)
                p = it.data(Qt.UserRole)
                key = os.path.normcase(os.path.normpath(p))
                if key in by_path:
                    by_path[key]["active"] = (it.checkState() == Qt.Checked)
            new_sources = list(by_path.values())
            self._save_excel_sources(new_sources)
            dlg.accept()

        save_btn.clicked.connect(_do_save)
        dlg.exec_()

    def _ensure_default_excel_source_present(self):
        """Однократная миграция: если текущий self.excel_path не присутствует
        в excel_sources, добавить его как активный. Это гарантирует, что после
        обновления приложения исходный файл не «пропадёт» из загрузки."""
        try:
            cur_path = getattr(self, "excel_path", "") or ""
            if not cur_path or not os.path.isfile(cur_path):
                return
            sources = self._load_excel_sources()
            cur_norm = os.path.normcase(os.path.normpath(cur_path))
            for s in sources:
                if os.path.normcase(os.path.normpath(s["path"])) == cur_norm:
                    return  # уже есть — ничего не делаем
            sources.append({
                "path": cur_path,
                "active": True,
                "added_at": time.time(),
            })
            self._save_excel_sources(sources)
        except Exception as e:
            sys.stderr.write(f"[MirlisMark] Default excel source migration error: {e}\n")

    def _merge_products(self, product_lists: list) -> list:
        """Слить продукты из нескольких файлов с дедупликацией.
        Дубликат = совпадение (name, shelf_life_hours, allowed_units).
        Любое различие хотя бы в одном поле делает продукты разными — оба остаются."""
        seen = set()
        result = []
        for products in product_lists:
            for p in products:
                name = (p.get("name") or "").strip()
                if not name:
                    continue
                life = int(p.get("shelf_life_hours") or 0)
                units = tuple(sorted([str(u).strip().lower() for u in (p.get("allowed_units") or [])]))
                key = (name.strip().lower(), life, units)
                if key in seen:
                    continue
                seen.add(key)
                result.append(p)
        return result

    def _merge_staff(self, staff_lists: list) -> list:
        """Слить сотрудников/цехи из нескольких файлов с дедупликацией по имени.
        Дубликат = полное совпадение значения name/fio (case-insensitive после strip)."""
        seen = set()
        result = []
        for staff in staff_lists:
            for s in staff:
                fio = (s.get("name") or s.get("fio") or "").strip()
                if not fio:
                    continue
                key = fio.lower()
                if key in seen:
                    continue
                seen.add(key)
                result.append(s)
        return result

    def _open_statistics(self):
        """Открыть страницу статистики."""
        if hasattr(self, "content_stack") and hasattr(self, "stats_page"):
            self.content_stack.setCurrentWidget(self.stats_page)
        self._set_statistics_mode(True)
        if hasattr(self, "stats_page") and hasattr(self.stats_page, "set_archive_root"):
            self.stats_page.set_archive_root(self._labels_archive_root())
        if hasattr(self, "stats_page") and hasattr(self.stats_page, "refresh_from_archive"):
            self.stats_page.refresh_from_archive()
        self._set_statistics_period(getattr(self, "_stats_period", "day"))

    def _return_to_print_mode(self):
        """Вернуться на основной экран печати."""
        if hasattr(self, "stats_page") and hasattr(self.stats_page, "leave_statistics_detail"):
            self.stats_page.leave_statistics_detail()
        if hasattr(self, "content_stack") and hasattr(self, "main_page"):
            self.content_stack.setCurrentWidget(self.main_page)
        self._set_statistics_mode(False)

    def _return_to_statistics_dashboard(self):
        """Выйти из drill-down detail и показать главный dashboard статистики."""
        if hasattr(self, "stats_page") and hasattr(self.stats_page, "leave_statistics_detail"):
            self.stats_page.leave_statistics_detail()

    def _on_statistics_detail_mode(self, in_detail: bool):
        b = getattr(self, "back_to_stats_dashboard_btn", None)
        if b is None:
            return
        b.setVisible(bool(in_detail) and getattr(self, "_statistics_mode", False))

    def _set_statistics_mode(self, enabled: bool):
        """Переключение видимости кнопок top bar для режима статистики."""
        enabled = bool(enabled)
        self._statistics_mode = enabled

        for b in (
            getattr(self, "tools_frame", None),
            getattr(self, "stats_btn", None),
        ):
            if b is not None:
                b.setVisible(not enabled)

        for b in (
            getattr(self, "day_btn", None),
            getattr(self, "week_btn", None),
            getattr(self, "month_btn", None),
            getattr(self, "period_btn", None),
            getattr(self, "back_to_print_btn", None),
        ):
            if b is not None:
                b.setVisible(enabled)

        bd = getattr(self, "back_to_stats_dashboard_btn", None)
        if bd is not None:
            bd.setVisible(enabled and getattr(self.stats_page, "_detail_mode", False) if hasattr(self, "stats_page") else False)

    def _set_statistics_period(self, period: str):
        """Выбор периода статистики: day/week/month/custom."""
        if period not in ("day", "week", "month", "custom"):
            return
        self._stats_period = period
        if hasattr(self, "stats_page") and hasattr(self.stats_page, "set_period"):
            self.stats_page.set_period(period)

    def _open_statistics_custom_period(self):
        if hasattr(self, "stats_page") and hasattr(self.stats_page, "open_custom_period_dialog"):
            if self.stats_page.open_custom_period_dialog(self):
                self._stats_period = "custom"

    def choose_excel_path(self):
        # показываем инструкцию с картинками перед выбором файла
        dlg = QDialog(self)
        dlg.setWindowTitle("Структура файла Excel")
        dlg.setMinimumSize(700, 500)
        dlg.resize(750, 600)

        dlg_layout = QVBoxLayout(dlg)
        dlg_layout.setContentsMargins(0, 0, 0, 12)
        dlg_layout.setSpacing(0)

        # скролл-область
        scroll = QScrollArea()
        QScroller.grabGesture(scroll.viewport(), QScroller.TouchGesture)
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("QScrollArea { border: none; background: #ffffff; }")

        content = QWidget()
        content.setStyleSheet("background: #ffffff;")
        lay = QVBoxLayout(content)
        lay.setContentsMargins(28, 24, 28, 24)
        lay.setSpacing(16)

        def add_title(text):
            lbl = QLabel(text)
            lbl.setStyleSheet("font-size: 20px; font-weight: 800; color: #111827; background: transparent;")
            lbl.setWordWrap(True)
            lay.addWidget(lbl)

        def add_text(text):
            lbl = QLabel(text)
            lbl.setStyleSheet("font-size: 14px; color: #374151; background: transparent;")
            lbl.setWordWrap(True)
            lay.addWidget(lbl)

        def add_image(path, max_width=640):
            if not os.path.isfile(path):
                return
            lbl = QLabel()
            lbl.setStyleSheet("background: transparent;")
            pix = QPixmap(path)
            if not pix.isNull():
                if pix.width() > max_width:
                    pix = pix.scaledToWidth(max_width, Qt.SmoothTransformation)
                lbl.setPixmap(pix)
            lay.addWidget(lbl)

        def add_separator():
            sep = QFrame()
            sep.setFrameShape(QFrame.HLine)
            sep.setStyleSheet("color: #e5e7eb; background: #e5e7eb; border: none; max-height: 1px;")
            lay.addWidget(sep)

        # --- содержимое инструкции ---
        hero_lbl = QLabel()
        hero_lbl.setStyleSheet("background: transparent;")
        hero_lbl.setAlignment(Qt.AlignHCenter)
        hero_pix = QPixmap(HAPPY_HERO_PATH)
        if not hero_pix.isNull():
            hero_pix = hero_pix.scaled(220, 220, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            hero_lbl.setPixmap(hero_pix)
            lay.addWidget(hero_lbl)

        add_title("Структура файла Excel")
        add_text(
            "Файл должен содержать 3 листа с точными названиями: «продукт», «изготовил», «цех».\n"
            "Первая строка каждого листа — заголовки колонок, данные начинаются со второй строки."
        )

        add_separator()

        add_text("Внизу файла должны быть видны 3 вкладки:")
        add_image(HELP_IMG_TABS)

        add_separator()

        add_title("1. Лист «продукт»")
        add_text(
            "Заголовки: Код | Наименование | Срок годности (ч) | Ед. измер. | Активен | Комментарий\n\n"
            "• «Ед. измер.» — через запятую: кг, шт или кг,шт\n"
            "• «Активен» — Да или 1 (показывать), Нет или 0 (скрыть)"
        )
        add_image(HELP_IMG_PRODUCT)

        add_separator()

        add_title("2. Лист «изготовил»")
        add_text("Заголовки: ФИО | Активен")
        add_image(HELP_IMG_MADE)

        add_separator()

        add_title("3. Лист «цех»")
        add_text("Заголовки: Цех | Активен")
        add_image(HELP_IMG_WORKSHOP)

        add_separator()

        add_text(
            "Если в листе «изготовил» или «цех» только одна запись — "
            "она подставится автоматически, и работнику не нужно будет её выбирать каждый раз."
        )

        scroll.setWidget(content)
        dlg_layout.addWidget(scroll, 1)

        # кнопки
        btn_row = QHBoxLayout()
        btn_row.setContentsMargins(20, 0, 20, 0)
        btn_row.setSpacing(12)
        btn_row.addStretch(1)

        btn_cancel = QPushButton("Отмена")
        btn_cancel.setMinimumHeight(40)
        btn_cancel.setCursor(Qt.PointingHandCursor)
        btn_cancel.clicked.connect(dlg.reject)

        btn_ok = QPushButton("Выбрать файл")
        btn_ok.setMinimumHeight(40)
        btn_ok.setCursor(Qt.PointingHandCursor)
        btn_ok.setStyleSheet(
            "QPushButton { background: #16a34a; color: #ffffff; font-weight: 700; "
            "border: 1px solid #15803d; border-radius: 12px; padding: 8px 24px; }"
            "QPushButton:hover { background: #15803d; }"
        )
        btn_ok.clicked.connect(dlg.accept)

        btn_row.addWidget(btn_cancel)
        btn_row.addWidget(btn_ok)
        dlg_layout.addLayout(btn_row)

        if dlg.exec_() != QDialog.Accepted:
            return

        current_dir = os.path.dirname(self.excel_path)

        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Выберите файл Excel с данными",
            current_dir,
            "Excel файлы (*.xlsx *.xls);;Все файлы (*)",
        )

        if not file_path:
            return

        self.excel_path = file_path
        settings = _load_settings()
        settings["excel_path"] = file_path
        _save_settings(settings)

        self.reload_excel(show_message=True)

    # ---------------- Helpers ----------------
    def get_product(self, name):
        name = (name or "").strip()
        return next((p for p in self.products if (p.get("name") or "").strip() == name), None)

    def _clear_layout(self, layout):
        while layout.count():
            item = layout.takeAt(0)
            w = item.widget()
            child_layout = item.layout()
            if child_layout is not None:
                self._clear_layout(child_layout)
            if w is not None:
                w.deleteLater()

    # ---------------- History helpers ----------------
    def _filtered_history_entries(self):
        if not self._history_filter_text:
            return list(self.history_entries)
        q = self._history_filter_text
        result = []
        for e in self.history_entries:
            text = " ".join(
                [
                    str(e.get("product", "")),
                    str(e.get("qty", "")),
                    str(e.get("made", "")),
                    str(e.get("checked", "")),
                    str(e.get("time", "")),
                    str(e.get("batch", "")),
                ]
            ).lower()
            if q in text:
                result.append(e)
        return result

    def _rebuild_history_view(self):
        if not hasattr(self, "history_list_layout"):
            return

        filtered_history = self._filtered_history_entries()
        total = len(filtered_history)
        page_size = max(1, self.history_page_size)
        pages = max(1, math.ceil(total / page_size))
        self.history_page = max(0, min(self.history_page, pages - 1))

        start = self.history_page * page_size
        end = start + page_size
        page_items = filtered_history[start:end]

        self._clear_layout(self.history_list_layout)

        for e in page_items:
            card = QFrame()
            card.setObjectName("HistoryCard")
            card.setCursor(Qt.PointingHandCursor)
            card.setProperty("selected", (e.get("id") == getattr(self, "_selected_history_id", None)))
            card_layout = QVBoxLayout(card)
            card_layout.setContentsMargins(10, 8, 10, 8)
            card_layout.setSpacing(4)

            top_row = QHBoxLayout()
            top_row.setSpacing(6)

            prod_label = QLabel(str(e.get("product", "")))
            prod_label.setStyleSheet("font-weight: 600; font-size: 15px; background: transparent; color: #0f172a;")
            prod_label.setAttribute(Qt.WA_TransparentForMouseEvents, True)
            qty_label = QLabel(str(e.get("qty", "")))
            qty_label.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            qty_label.setStyleSheet("font-weight: 600; color: #0f172a; font-size: 15px; background: transparent;")
            qty_label.setAttribute(Qt.WA_TransparentForMouseEvents, True)

            top_row.addWidget(prod_label, 1)
            top_row.addWidget(qty_label, 0)

            made = str(e.get("made", ""))
            checked = str(e.get("checked", ""))
            mid_parts = [p for p in [made, checked] if p]
            mid_text = " · ".join(mid_parts) if mid_parts else ""
            mid_row = QLabel(mid_text)
            mid_row.setStyleSheet("color: #374151; font-size: 14px; background: transparent;")
            mid_row.setAttribute(Qt.WA_TransparentForMouseEvents, True)

            bottom_row = QHBoxLayout()
            bottom_row.setSpacing(6)

            time_label = QLabel(str(e.get("time", "")))
            time_label.setStyleSheet("color: #4b5563; font-size: 13px; background: transparent;")
            time_label.setAttribute(Qt.WA_TransparentForMouseEvents, True)

            batch = str(e.get("batch", ""))
            batch_label = QLabel(f"№ {batch}" if batch else "")
            batch_label.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            batch_label.setStyleSheet("color: #4b5563; font-size: 13px; background: transparent;")
            batch_label.setAttribute(Qt.WA_TransparentForMouseEvents, True)

            bottom_row.addWidget(time_label, 1)
            bottom_row.addWidget(batch_label, 0)

            card_layout.addLayout(top_row)
            card_layout.addWidget(mid_row)
            card_layout.addLayout(bottom_row)

            card.mousePressEvent = (lambda ev, ent=e: self._on_history_clicked(ent))
            card.style().unpolish(card)
            card.style().polish(card)
            self.history_list_layout.addWidget(card)

        self.history_list_layout.addStretch(1)

        if hasattr(self, "history_page_label"):
            self.history_page_label.setText(f"Страница {self.history_page + 1} из {pages}")
        if hasattr(self, "history_prev_btn"):
            self.history_prev_btn.setEnabled(self.history_page > 0)
        if hasattr(self, "history_next_btn"):
            self.history_next_btn.setEnabled(self.history_page < pages - 1)

    def _on_history_search_text_changed(self, text):
        self._history_filter_text = (text or "").strip().lower()
        self.history_page = 0
        self._rebuild_history_view()

    def _change_history_page(self, delta):
        self.history_page += delta
        self._rebuild_history_view()

    def _build_history_entry_from_label(self, label, qty_display, unit_ui):
        produced_at = getattr(label, "produced_at", datetime.now())
        preview_text = (
            f"{getattr(label, 'weekday', '')}\n"
            f"Продукт: {getattr(label, 'product_name', '')}\n"
            f"Вес/шт: {getattr(label, 'qty_value', '')} {getattr(label, 'qty_unit_ru', '')}\n"
            f"Дата/время: {format_dt(produced_at)}\n"
            f"№ партии: {getattr(label, 'batch', '')}\n"
            f"Годен до: {format_dt(getattr(label, 'expires_at', produced_at))}\n"
            f"Изготовил: {getattr(label, 'made_by', '')}\n"
            f"Цех: {getattr(label, 'checked_by', '')}\n"
        )

        return {
            "id": time.time_ns(),
            "ts": time.time(),
            "product_name": getattr(label, "product_name", ""),
            "unit_ui": unit_ui,
            "qty_value": str(qty_display),
            "made_by": getattr(label, "made_by", ""),
            "made_manual": bool(self.made_manual.isChecked()),
            "checked_by": getattr(label, "checked_by", ""),
            "checked_manual": bool(self.checked_manual.isChecked()),
            "preview_text": preview_text,
            "product": getattr(label, "product_name", ""),
            "qty": f"{qty_display} {unit_ui}".strip(),
            "made": getattr(label, "made_by", ""),
            "checked": getattr(label, "checked_by", ""),
            "time": format_dt(produced_at),
            "batch": getattr(label, "batch", ""),
        }

    def _append_history_entry(self, entry):
        if not hasattr(self, "history_entries") or not isinstance(self.history_entries, list):
            self.history_entries = []
        self.history_entries.insert(0, entry)
        if not self._history_filter_text:
            self.history_page = 0
        self._rebuild_history_view()

    def _on_history_clicked(self, entry):
        self._selected_history_id = entry.get("id")
        self.apply_history_entry(entry)
        self._rebuild_history_view()

    def apply_history_entry(self, entry):
        self._loading_from_history = True
        try:
            to_block = (
                self.product_combo,
                self.qty_input,
                self.made_manual,
                self.made_combo,
                self.made_input,
                self.checked_manual,
                self.checked_combo,
                self.checked_input,
            )
            for w in to_block:
                w.blockSignals(True)

            product_name = str(entry.get("product_name") or entry.get("product") or "").strip()
            unit_ui = str(entry.get("unit_ui") or "").strip()
            qty_value = str(entry.get("qty_value") or "").strip()
            made_by = str(entry.get("made_by") or entry.get("made") or "").strip()
            made_manual = bool(entry.get("made_manual", False))
            checked_by = str(entry.get("checked_by") or entry.get("checked") or "").strip()
            checked_manual = bool(entry.get("checked_manual", False))

            idx = self.product_combo.findText(product_name)
            if idx >= 0:
                self.product_combo.setCurrentIndex(idx)
            else:
                self.product_combo.setEditText(product_name)

            self.on_product_changed(product_name)

            self.unit_combo.blockSignals(True)
            try:
                if unit_ui:
                    idxu = self.unit_combo.findText(unit_ui)
                    if idxu >= 0:
                        self.unit_combo.setCurrentIndex(idxu)
            finally:
                self.unit_combo.blockSignals(False)

            self.qty_input.setText(qty_value)

            self.made_manual.setChecked(made_manual)
            self.toggle_made_mode()
            if made_manual:
                self.made_input.setText(made_by)
            else:
                idxm = self.made_combo.findText(made_by)
                if idxm >= 0:
                    self.made_combo.setCurrentIndex(idxm)
                elif made_by:
                    self.made_manual.setChecked(True)
                    self.toggle_made_mode()
                    self.made_input.setText(made_by)

            self.checked_manual.setChecked(checked_manual)
            self.toggle_checked_mode()
            if checked_manual:
                self.checked_input.setText(checked_by)
            else:
                idxc = self.checked_combo.findText(checked_by)
                if idxc >= 0:
                    self.checked_combo.setCurrentIndex(idxc)
                elif checked_by:
                    self.checked_manual.setChecked(True)
                    self.toggle_checked_mode()
                    self.checked_input.setText(checked_by)

        finally:
            for w in (
                self.product_combo,
                self.qty_input,
                self.made_manual,
                self.made_combo,
                self.made_input,
                self.checked_manual,
                self.checked_combo,
                self.checked_input,
            ):
                w.blockSignals(False)
            self._loading_from_history = False

        preview_html = entry.get("preview_html")
        preview_text = entry.get("preview_text")
        if isinstance(preview_html, str) and preview_html.strip():
            self._set_preview_html_programmatically(preview_html)
            self._user_edited_preview = True
            _, can_print = self._build_label_plain_text()
            self._refresh_print_btn_state()
        elif isinstance(preview_text, str) and preview_text.strip():
            self._set_preview_text_programmatically(preview_text)
            self._user_edited_preview = True
            _, can_print = self._build_label_plain_text()
            self._refresh_print_btn_state()
        else:
            self._user_edited_preview = False
            self.refresh_preview()

    # ---------------- UI actions ----------------
    def toggle_made_mode(self):
        manual = self.made_manual.isChecked()
        self.made_combo.setVisible(not manual)
        self.made_input.setVisible(manual)
        if not getattr(self, "_loading_from_history", False):
            self.refresh_preview()

    def toggle_checked_mode(self):
        manual = self.checked_manual.isChecked()
        self.checked_combo.setVisible(not manual)
        self.checked_input.setVisible(manual)
        if not getattr(self, "_loading_from_history", False):
            self.refresh_preview()

    def on_product_changed(self, product_name):
        self.unit_combo.blockSignals(True)
        self.unit_combo.clear()

        product = self.get_product(product_name)
        if product:
            units = product.get("allowed_units", [])
            if isinstance(units, str):
                units = [u.strip() for u in units.split(",") if u.strip()]

            # добавляем только непустые уникальные значения
            seen = set()
            for u in units:
                u = (u or "").strip()
                if not u or u in seen:
                    continue
                seen.add(u)
                if u == "kg":
                    self.unit_combo.addItem("кг")
                elif u == "pcs":
                    self.unit_combo.addItem("шт")
                else:
                    self.unit_combo.addItem(u)

        self.unit_combo.blockSignals(False)
        if not getattr(self, "_loading_from_history", False):
            self.refresh_preview()

    def _step_for_unit(self):
        unit_ru = self.unit_combo.currentText()
        return 1.0 if unit_ru == "шт" else 0.1

    # ─── Весы (COM-порт, автоопределение) ───────────────────
    _SCALE_BAUD = 9600
    _scale_reader = None     # инстанс ScaleReader

    def _read_scale_weight(self):
        """Запускает фоновое автоопределение порта весов и считывание веса."""
        if not SERIAL_AVAILABLE:
            QMessageBox.warning(
                self, "Весы",
                "Модуль pyserial не установлен.\n"
                "Выполните: python -m pip install pyserial"
            )
            return
        if self._scale_reader is not None and self._scale_reader.isRunning():
            return  # уже идёт чтение — игнорируем повторный клик

        self.scale_btn.setEnabled(False)
        self.scale_btn.setText("Поиск весов...")

        # Подтянуть сохранённый порт (если уже находили ранее) — это ускорит запуск
        preferred = ""
        try:
            preferred = str(_load_settings().get("scale_port", "") or "")
        except Exception:
            preferred = ""

        self._scale_reader = ScaleReader(
            preferred_port=preferred,
            baud=self._SCALE_BAUD,
            per_port_timeout=1.5,
            saved_port_timeout=2.5,
            parent=self,
        )
        self._scale_reader.weight_received.connect(self._on_scale_weight)
        self._scale_reader.raw_received.connect(
            lambda msg: sys.stderr.write(f"[Scale] {msg}\n")
        )
        self._scale_reader.error_occurred.connect(self._on_scale_error)
        self._scale_reader.progress.connect(self._on_scale_progress)
        self._scale_reader.port_found.connect(self._on_scale_port_found)
        self._scale_reader.finished.connect(self._on_scale_done)
        self._scale_reader.start()

    def _on_scale_weight(self, value: float):
        """Подставляет считанный вес в qty_input."""
        try:
            idx_kg = self.unit_combo.findText("кг")
            if idx_kg >= 0 and self.unit_combo.currentIndex() != idx_kg:
                self.unit_combo.setCurrentIndex(idx_kg)
        except Exception:
            pass
        text = str(round(float(value), 3)).rstrip("0").rstrip(".")
        if not text:
            text = "0"
        self.qty_input.setText(text)
        try:
            self.refresh_preview()
        except Exception:
            pass

    def _on_scale_error(self, msg: str):
        QMessageBox.warning(self, "Весы — ошибка", msg)

    def _on_scale_progress(self, msg: str):
        """Показывает текущий проверяемый порт прямо на кнопке."""
        try:
            self.scale_btn.setText(msg)
        except Exception:
            pass

    def _on_scale_port_found(self, port: str):
        """Сохраняет найденный рабочий порт для быстрого запуска в будущем."""
        try:
            s = _load_settings()
            if s.get("scale_port") != port:
                s["scale_port"] = port
                _save_settings(s)
                sys.stderr.write(f"[Scale] Сохранён рабочий порт: {port}\n")
        except Exception as e:
            sys.stderr.write(f"[Scale] Не удалось сохранить порт: {e}\n")

    def _on_scale_done(self):
        """Восстанавливает кнопку после завершения чтения."""
        self.scale_btn.setEnabled(True)
        self.scale_btn.setText("Взвесить")
        # Восстановить иконку (Qt сбрасывает её при setText в некоторых стилях — на всякий случай)
        try:
            _pix = QPixmap(resource_path("assets/scales.png"))
            if not _pix.isNull():
                _pix = _pix.scaledToHeight(28, Qt.SmoothTransformation)
                self.scale_btn.setIcon(QIcon(_pix))
                self.scale_btn.setIconSize(QSize(28, 28))
        except Exception:
            pass

    def increase_qty(self):
        step = self._step_for_unit()
        try:
            value = float(self.qty_input.text().replace(",", "."))
        except Exception:
            value = 0.0
        value += step
        self.qty_input.setText(str(round(value, 3)).rstrip("0").rstrip("."))

    def decrease_qty(self):
        step = self._step_for_unit()
        try:
            value = float(self.qty_input.text().replace(",", "."))
        except Exception:
            value = 0.0
        value = max(0.0, value - step)
        self.qty_input.setText(str(round(value, 3)).rstrip("0").rstrip("."))

    def clear_fields(self):
        # Очистка полей означает: активной "выбранной из истории" этикетки больше нет.
        self._selected_history_id = None
        # Сразу снимаем подсветку в UI (иначе она останется до следующей перестройки истории).
        try:
            self._rebuild_history_view()
        except Exception:
            pass
        self.product_combo.setCurrentIndex(-1)
        self.product_combo.setEditText("")

        self.unit_combo.setCurrentIndex(0)
        self.qty_input.clear()
        if hasattr(self, "copies_input"):
            self.copies_input.clear()

        self.made_manual.setChecked(False)
        self.checked_manual.setChecked(False)

        # если в списке один вариант — оставляем автовыбор, иначе сбрасываем
        if self.made_combo.count() == 2:
            self.made_combo.setCurrentIndex(1)
        else:
            self.made_combo.setCurrentIndex(0)

        if self.checked_combo.count() == 2:
            self.checked_combo.setCurrentIndex(1)
        else:
            self.checked_combo.setCurrentIndex(0)

        self.made_input.clear()
        self.checked_input.clear()

        # Сброс текста и форматирования preview (иначе следующая автоэтикетка наследует ручной стиль).
        if hasattr(self, "preview"):
            self.preview.clearFocus()
            self._reset_preview_document_defaults()

        if hasattr(self, "btn_bold"):
            self.btn_bold.blockSignals(True)
            self.btn_bold.setChecked(False)
            self.btn_bold.blockSignals(False)

        if hasattr(self, "btn_italic"):
            self.btn_italic.blockSignals(True)
            self.btn_italic.setChecked(False)
            self.btn_italic.blockSignals(False)

        if hasattr(self, "btn_underline"):
            self.btn_underline.blockSignals(True)
            self.btn_underline.setChecked(False)
            self.btn_underline.blockSignals(False)

        if hasattr(self, "btn_align_left"):
            self.btn_align_left.blockSignals(True)
            self.btn_align_left.setChecked(False)
            self.btn_align_left.blockSignals(False)

        if hasattr(self, "btn_align_center"):
            self.btn_align_center.blockSignals(True)
            self.btn_align_center.setChecked(False)
            self.btn_align_center.blockSignals(False)

        if hasattr(self, "btn_align_right"):
            self.btn_align_right.blockSignals(True)
            self.btn_align_right.setChecked(False)
            self.btn_align_right.blockSignals(False)

        self.label_w_mm = 70.0
        self.label_h_mm = 70.0
        self._resize_label_preview()

        # reset toolbar font state (иначе auto preview наследует выбранный вручную шрифт/размер)
        try:
            if hasattr(self, "font_combo"):
                self.font_combo.blockSignals(True)
                self.font_combo.setCurrentFont(QFont("MS Shell Dlg 2"))
                self.font_combo.blockSignals(False)
        except Exception:
            pass

        self._base_font_size = 20

        try:
            if hasattr(self, "font_size_combo"):
                self.font_size_combo.blockSignals(True)
                self.font_size_combo.setCurrentText("20")
                self.font_size_combo.blockSignals(False)
        except Exception:
            pass

        self._user_edited_preview = False

        # Важно: после сброса пусть каждая этикетка заново
        # применит СВОИ дефолтные автонастройки.
        self.refresh_preview(force=True)

    def _reset_preview_document_defaults(self):
        """Жёсткий reset QTextDocument после «Очистить», чтобы автоэтикетка не наследовала ручной стиль."""
        if not hasattr(self, "preview"):
            return
        self._updating_preview = True
        try:
            self.preview.blockSignals(True)
            font_family = self.font_combo.currentFont().family()
            preview_scale = self._effective_preview_font_scale()
            effective_base = float(self._base_font_size) * float(preview_scale)

            default_font = QFont(font_family, max(1, int(round(effective_base))))

            # Важно: создаём новый документ, чтобы не наследовать состояние/форматы старого QTextDocument
            doc = QTextDocument(self.preview)
            doc.setDefaultFont(default_font)
            self.preview.setDocument(doc)

            # Сброс состояния редактора/курсорных форматов
            self.preview.setAlignment(Qt.AlignLeft)
            base_fmt = QTextCharFormat()
            base_fmt.setFontFamily(font_family)
            base_fmt.setFontPointSize(effective_base)
            base_fmt.setFontWeight(QFont.Normal)
            base_fmt.setFontItalic(False)
            base_fmt.setFontUnderline(False)
            self.preview.setCurrentCharFormat(base_fmt)

            cursor = self.preview.textCursor()
            cursor.movePosition(QTextCursor.Start)
            self.preview.setTextCursor(cursor)
        finally:
            self.preview.blockSignals(False)
            self._updating_preview = False
    # ---------------- Preview / validation ----------------
    def _unit_code_from_ui(self, unit_text):
        if unit_text == "кг":
            return "kg"
        if unit_text == "шт":
            return "pcs"
        return None

    def _made_value(self):
        if self.made_manual.isChecked():
            return self.made_input.text().strip()
        val = self.made_combo.currentText().strip()
        return "" if val.startswith("—") else val

    def _checked_value(self):
        if self.checked_manual.isChecked():
            return self.checked_input.text().strip()
        val = self.checked_combo.currentText().strip()
        return "" if val.startswith("—") else val

    _WEEKDAYS = {"ПОНЕДЕЛЬНИК", "ВТОРНИК", "СРЕДА", "ЧЕТВЕРГ", "ПЯТНИЦА", "СУББОТА", "ВОСКРЕСЕНЬЕ"}

    def _set_preview_text_programmatically(self, text):
        self._updating_preview = True
        try:
            self.preview.blockSignals(True)
            font_family = self.font_combo.currentFont().family()
            preview_scale = self._effective_preview_font_scale()
            effective_base_font = float(self._base_font_size) * float(preview_scale)
            fmt_base = QTextCharFormat()
            fmt_base.setFontFamily(font_family)
            fmt_base.setFontPointSize(effective_base_font)
            fmt_base.setFontWeight(QFont.Normal)
            fmt_base.setFontItalic(False)
            fmt_base.setFontUnderline(False)

            # Автопостроение должно быть независимым от editor state:
            # начинаем с чистого документа и работаем только через локальные форматы/cursor.
            self.preview.clear()
            self.preview.document().setDefaultFont(
                QFont(font_family, max(1, int(round(effective_base_font))))
            )
            self.preview.setPlainText(text)

            doc = self.preview.document()
            cursor = QTextCursor(doc)
            cursor.beginEditBlock()
            try:
                # База: жёстко применяем char+block формат ко ВСЕМУ документу
                cursor.select(QTextCursor.Document)
                cursor.setCharFormat(fmt_base)
                block_base = QTextBlockFormat()
                block_base.setAlignment(Qt.AlignLeft)
                cursor.mergeBlockFormat(block_base)
                cursor.clearSelection()
 
                is_colored = hasattr(self, "label_size_combo") and self.label_size_combo.currentText() == "Цветные"
                weekday_line_index = 1 if is_colored else 0
 
                lines = text.split("\n")
                weekday_line = ""
                if len(lines) > weekday_line_index:
                    weekday_line = lines[weekday_line_index].strip()
 
                if weekday_line in self._WEEKDAYS:
                    cursor.movePosition(QTextCursor.Start)
 
                    for _ in range(weekday_line_index):
                        cursor.movePosition(QTextCursor.NextBlock)
 
                    cursor.movePosition(QTextCursor.StartOfBlock)
                    cursor.movePosition(QTextCursor.EndOfBlock, QTextCursor.KeepAnchor)
 
                    fmt_weekday = QTextCharFormat()
                    fmt_weekday.setFontFamily(font_family)
 
                    if is_colored:
                        fmt_weekday.setFontPointSize(max(8.0, effective_base_font * 1.8))
                    else:
                        fmt_weekday.setFontPointSize(max(8.0, effective_base_font * 1.3))
 
                    fmt_weekday.setFontWeight(QFont.Bold)
                    cursor.setCharFormat(fmt_weekday)
 
                    block_fmt = QTextBlockFormat()
                    block_fmt.setAlignment(Qt.AlignHCenter)
                    cursor.mergeBlockFormat(block_fmt)
 
                cursor.clearSelection()
                cursor.movePosition(QTextCursor.Start)
                self.preview.setTextCursor(cursor)
            finally:
                cursor.endEditBlock()
        finally:
            self.preview.blockSignals(False)
            self._updating_preview = False

    def _set_preview_html_programmatically(self, html):
        self._updating_preview = True
        try:
            self.preview.blockSignals(True)
            self.preview.setHtml(html)
            cursor = self.preview.textCursor()
            cursor.movePosition(QTextCursor.Start)
            self.preview.setTextCursor(cursor)
        finally:
            self.preview.blockSignals(False)
            self._updating_preview = False

    def _build_label_plain_text(self):
        product_name = self.product_combo.currentText().strip()
        product = self.get_product(product_name)

        unit_ui = self.unit_combo.currentText()
        unit_code = self._unit_code_from_ui(unit_ui)
        qty = self.qty_input.text().strip().replace(",", ".")

        if not product:
            return ("Выберите продукт.", False)

        if unit_code is None:
            return ("Выберите единицу измерения (кг или шт).", False)

        if not qty:
            return ("Введите количество.", False)

        try:
            qty_float = float(qty)
        except Exception:
            return ("Количество должно быть числом (например 2 или 0.5).", False)

        if qty_float <= 0:
            return ("Количество должно быть больше 0.", False)

        if unit_code == "pcs" and abs(qty_float - round(qty_float)) > 1e-9:
            return ("Для 'шт' количество должно быть целым числом.", False)

        made_by = self._made_value()
        checked_by = self._checked_value()

        produced_at = None
        if getattr(self, "_preview_manual_mode", False):
            d = self.manual_datetime_picker.date()
            t = self.manual_datetime_picker.time_()
            produced_at = datetime(d.year(), d.month(), d.day(), t.hour(), t.minute())

        label = build_label(
            product_name=product["name"],
            shelf_life_hours=product["shelf_life_hours"],
            qty_value=str(qty_float).rstrip("0").rstrip("."),
            unit=unit_code,
            made_by=made_by,
            checked_by=checked_by,
            produced_at=produced_at,
        )

        text_parts = []
        is_colored = hasattr(self, "label_size_combo") and self.label_size_combo.currentText() == "Цветные"

        if is_colored:
            text_parts.append("")
            text_parts.append("")
        else:
            text_parts.append(f"{label.weekday}")

        text_parts.append(f"Продукт: {label.product_name}")
        text_parts.append(f"Вес/шт: {label.qty_value} {label.qty_unit_ru}")
        text_parts.append(f"Дата/время: {format_dt(label.produced_at)}")
        text_parts.append(f"№ партии: {label.batch}")
        text_parts.append(f"Годен до: {format_dt(label.expires_at)}")
        text_parts.append(f"Цех: {label.checked_by}")
        text_parts.append(f"Изготовил: {label.made_by}")

        text = "\n".join(text_parts) + "\n"
        return (text, True)

    def refresh_preview(self, force=False):
        if getattr(self, "_loading_from_history", False):
            return
        text, can_print = self._build_label_plain_text()
        self._refresh_print_btn_state()

        # Если текущая этикетка "пустая" (нет продукта / стандартное сообщение),
        # то история не должна держать подсветку выбранной карточки.
        # Важно: делаем это ДО ранних return (например, когда preview редактировали вручную).
        if getattr(self, "_selected_history_id", None) is not None:
            t = (text or "").strip()
            if (not t) or (t == "Выберите продукт."):
                self._selected_history_id = None
                self._rebuild_history_view()

        if self._user_edited_preview:
            return

        if self.preview.hasFocus() and not force:
            return

        self._set_preview_text_programmatically(text)

    def _on_preview_text_changed(self):
        if self._updating_preview:
            return
        self._user_edited_preview = True

    # ---------------- Editor formatting (selection-only) ----------------
    def _merge_format_on_selection(self, fmt):
        cursor = self.preview.textCursor()
        if not cursor.hasSelection():
            cursor.select(QTextCursor.WordUnderCursor)
        cursor.mergeCharFormat(fmt)
        self.preview.mergeCurrentCharFormat(fmt)

    def _toggle_bold_on_selection(self):
        fmt = QTextCharFormat()
        desired_bold = self.btn_bold.isChecked()
        fmt.setFontWeight(QFont.Bold if desired_bold else QFont.Normal)
        self._merge_format_on_selection(fmt)
        self._sync_format_toolbar_from_cursor()

    def _toggle_italic_on_selection(self):
        fmt = QTextCharFormat()
        desired = self.btn_italic.isChecked()
        fmt.setFontItalic(desired)
        self._merge_format_on_selection(fmt)
        self._sync_format_toolbar_from_cursor()

    def _toggle_underline_on_selection(self):
        fmt = QTextCharFormat()
        desired = self.btn_underline.isChecked()
        fmt.setFontUnderline(desired)
        self._merge_format_on_selection(fmt)
        self._sync_format_toolbar_from_cursor()

    def _sync_format_toolbar_from_cursor(self):
        cursor = self.preview.textCursor()
        fmt = cursor.charFormat() if cursor.charFormat().isValid() else self.preview.currentCharFormat()

        self.btn_bold.blockSignals(True)
        self.btn_italic.blockSignals(True)
        self.btn_underline.blockSignals(True)
        self.btn_align_left.blockSignals(True)
        self.btn_align_center.blockSignals(True)
        self.btn_align_right.blockSignals(True)
        try:
            self.btn_bold.setChecked(fmt.fontWeight() >= QFont.Bold)
            self.btn_italic.setChecked(fmt.fontItalic())
            self.btn_underline.setChecked(fmt.fontUnderline())

            align = cursor.blockFormat().alignment()
            self.btn_align_left.setChecked(align == Qt.AlignLeft)
            self.btn_align_center.setChecked(align == Qt.AlignHCenter)
            self.btn_align_right.setChecked(align == Qt.AlignRight)
        finally:
            self.btn_bold.blockSignals(False)
            self.btn_italic.blockSignals(False)
            self.btn_underline.blockSignals(False)
            self.btn_align_left.blockSignals(False)
            self.btn_align_center.blockSignals(False)
            self.btn_align_right.blockSignals(False)
    
        self._sync_font_size_from_cursor()

    def _sync_font_size_from_cursor(self):
        if not hasattr(self, "font_size_combo"):
            return

        cursor = self.preview.textCursor()
        fmt = cursor.charFormat()

        size = fmt.fontPointSize()
        if size <= 0:
            size = fmt.font().pointSizeF()
        if size <= 0:
            size = float(self._base_font_size)

        eff = self._effective_preview_font_scale()
        if eff > 1e-6:
            size = float(size) / eff

        size_int = int(round(size))
        if size_int <= 0:
            size_int = self._base_font_size

        self.font_size_combo.blockSignals(True)
        self.font_size_combo.lineEdit().setText(str(size_int))
        self.font_size_combo.blockSignals(False)

    def _set_alignment(self, align_flag):
        if align_flag == Qt.AlignLeft:
            self.preview.setAlignment(Qt.AlignLeft)
        elif align_flag == Qt.AlignHCenter:
            self.preview.setAlignment(Qt.AlignHCenter)
        elif align_flag == Qt.AlignRight:
            self.preview.setAlignment(Qt.AlignRight)
        elif align_flag == Qt.AlignJustify:
            self.preview.setAlignment(Qt.AlignJustify)
        self._sync_format_toolbar_from_cursor()

    def _set_font_family_on_selection(self, font):
        fmt = QTextCharFormat()
        fmt.setFontFamily(font.family())
        self._merge_format_on_selection(fmt)

    def _change_font_size(self, delta):
        self._base_font_size = max(8, min(72, self._base_font_size + delta))
        fmt = QTextCharFormat()
        fmt.setFontPointSize(float(self._base_font_size) * self._effective_preview_font_scale())
        self._merge_format_on_selection(fmt)
        if hasattr(self, "font_size_combo"):
            self.font_size_combo.blockSignals(True)
            self.font_size_combo.setCurrentText(str(self._base_font_size))
            self.font_size_combo.blockSignals(False)

    def on_font_size_combo_changed(self, text):
        t = (text or "").strip()
        if not t:
            return
        try:
            size = int(float(t))
        except Exception:
            return
        size = max(8, min(72, size))
        self._base_font_size = size
        fmt = QTextCharFormat()
        fmt.setFontPointSize(float(size) * self._effective_preview_font_scale())
        self._merge_format_on_selection(fmt)
        if hasattr(self, "font_size_combo"):
            self.font_size_combo.blockSignals(True)
            self.font_size_combo.setCurrentText(str(size))
            self.font_size_combo.blockSignals(False)

    def _get_copies(self):
        if not hasattr(self, "copies_input"):
            return 0
        txt = self.copies_input.text().strip()
        if not txt:
            return 0
        v = _safe_int(txt, 0)
        return max(0, min(9999, v))

    def _refresh_print_btn_state(self):
        """Обновляет доступность кнопки Печать с учётом и этикетки, и копий."""
        if not hasattr(self, "print_btn"):
            return
        _, can_label = self._build_label_plain_text()
        copies_ok = self._get_copies() > 0
        self.print_btn.setEnabled(can_label and copies_ok)

    def _sanitize_copies(self):
        if not hasattr(self, "copies_input"):
            return
        txt = self.copies_input.text().strip()
        # Разрешаем пустое поле — пользователь должен сам ввести количество.
        if not txt:
            self._refresh_print_btn_state()
            return
        # Убираем нецифровые символы, ограничиваем 1-999.
        digits_only = "".join(c for c in txt if c.isdigit())
        v = _safe_int(digits_only, 0)
        v = max(0, min(9999, v))
        corrected = str(v) if v > 0 else ""
        if self.copies_input.text().strip() != corrected:
            self.copies_input.blockSignals(True)
            self.copies_input.setText(corrected)
            self.copies_input.blockSignals(False)
        self._refresh_print_btn_state()

    def increase_copies(self):
        if not hasattr(self, "copies_input"):
            return
        self.copies_input.setText(str(self._get_copies() + 1))

    def decrease_copies(self):
        if not hasattr(self, "copies_input"):
            return
        self.copies_input.setText(str(max(1, self._get_copies() - 1)))

    def _apply_copies_to_tspl(self, tspl, copies):
        lines = tspl.strip().splitlines()
        for i in range(len(lines) - 1, -1, -1):
            if lines[i].strip().upper().startswith("PRINT"):
                lines[i] = f"PRINT {copies}"
                return "\n".join(lines).strip()
        return (tspl.strip() + f"\nPRINT {copies}").strip()

    # ---------------- Rendering preview → TSPL bitmap ----------------
    def _render_preview_to_tspl_bytes(
        self,
        label_w_mm=None,
        label_h_mm=None,
        dpi=203,
        density=10,
        speed=4,
        threshold=200,
        copies=1,
    ):
        w_mm = label_w_mm if label_w_mm is not None else getattr(self, "label_w_mm", 58.0)
        h_mm = label_h_mm if label_h_mm is not None else getattr(self, "label_h_mm", 80.0)
        w_px = int(w_mm / 25.4 * dpi)
        h_px = int(h_mm / 25.4 * dpi)

        doc = self.preview.document().clone()

        # Печать должна быть одинаковой на любом мониторе/диагонали/DPI.
        # Не используем self.preview.viewport(), потому что его размер зависит от окна
        # и от системного масштабирования Windows, из-за чего итоговая печать "плывёт".
        # Используем фиксированный виртуальный канвас; компенсация экрана уже в pt документа
        # (через _effective_preview_font_scale), здесь снимаем только «оконный» _preview_scale, если ≠1.
        PRINT_VIRTUAL_W = 450.0
        print_virtual_h = PRINT_VIRTUAL_W * (h_mm / w_mm) if w_mm > 0 else PRINT_VIRTUAL_W

        preview_scale = float(getattr(self, "_preview_scale", 1.0)) or 1.0
        if preview_scale > 0 and abs(preview_scale - 1.0) > 1e-6:
            blk = doc.begin()
            while blk.isValid():
                it = blk.begin()
                while not it.atEnd():
                    frag = it.fragment()
                    if frag.isValid():
                        old_pt = frag.charFormat().fontPointSize()
                        if old_pt > 0:
                            cur = QTextCursor(doc)
                            cur.setPosition(frag.position())
                            cur.setPosition(
                                frag.position() + frag.length(),
                                QTextCursor.KeepAnchor,
                            )
                            fmt = QTextCharFormat()
                            fmt.setFontPointSize(old_pt * (1.0 / preview_scale))
                            cur.mergeCharFormat(fmt)
                    it += 1
                blk = blk.next()
            df = doc.defaultFont()
            if df.pointSizeF() > 0:
                nf = QFont(df)
                nf.setPointSizeF(df.pointSizeF() * (1.0 / preview_scale))
                doc.setDefaultFont(nf)

        doc.setPageSize(QSizeF(PRINT_VIRTUAL_W, print_virtual_h))
        # Клон мог унаследовать textWidth от QTextEdit (viewport); фиксируем под виртуальный канвас.
        doc.setTextWidth(PRINT_VIRTUAL_W)

        scale_x = w_px / PRINT_VIRTUAL_W
        scale_y = h_px / print_virtual_h

        # Только QImage (не QPixmap): явно фиксированный DPI и DPR=1, без привязки к экрану.
        img = QImage(w_px, h_px, QImage.Format_RGB32)
        if hasattr(img, "setDevicePixelRatio"):
            img.setDevicePixelRatio(1.0)
        _dpm_96 = int(round(96 * 1000.0 / 25.4))
        img.setDotsPerMeterX(_dpm_96)
        img.setDotsPerMeterY(_dpm_96)
        img.fill(QColor(255, 255, 255))

        painter = QPainter(img)
        painter.setRenderHint(QPainter.Antialiasing, True)
        painter.setRenderHint(QPainter.TextAntialiasing, True)
        painter.scale(scale_x, scale_y)

        # Для «Цветных» этикеток (приходят с типографии с верхней цветной
        # полосой и днём недели) сдвигаем текст вниз на ~2 мм, чтобы первая
        # строка не задевала эту полосу. Единицы painter после scale() —
        # виртуальные пиксели PRINT_VIRTUAL_W; конвертация: 2 мм * (PRINT_VIRTUAL_W / w_mm).
        # Превью НЕ затрагивается (его рендерит другой путь через QGraphicsView).
        try:
            _is_colored_print = (
                hasattr(self, "label_size_combo")
                and self.label_size_combo.currentText() == "Цветные"
            )
            if _is_colored_print and w_mm > 0:
                _colored_top_offset_mm = 2.0   # ← одна цифра для подстройки
                _colored_offset_virtual_px = PRINT_VIRTUAL_W * _colored_top_offset_mm / float(w_mm)
                painter.translate(0.0, _colored_offset_virtual_px)
        except Exception:
            pass

        doc.drawContents(painter)
        painter.end()

        width_bytes = (w_px + 7) // 8
        raster = bytearray()

        for y in range(h_px):
            for xb in range(width_bytes):
                byte_val = 0
                for bit in range(8):
                    x = xb * 8 + bit
                    if x >= w_px:
                        byte_val |= (1 << (7 - bit))
                        continue
                    pixel = img.pixel(x, y)
                    gray = ((pixel >> 16) & 0xFF) + ((pixel >> 8) & 0xFF) + (pixel & 0xFF)
                    gray = gray // 3
                    if gray > threshold:
                        byte_val |= (1 << (7 - bit))
                raster.append(byte_val)

        gap_mm = 2.0
        header = (
            f"SIZE {w_mm} mm, {h_mm} mm\r\n"
            f"GAP {gap_mm} mm, 0 mm\r\n"
        ).encode("ascii")
        header += (
            f"SPEED {speed}\r\n"
            f"DENSITY {density}\r\n"
            f"DIRECTION 1\r\n"
            f"REFERENCE 0,0\r\n"
            f"CLS\r\n"
            f"BITMAP 0,0,{width_bytes},{h_px},0,"
        ).encode("ascii")

        tail = f"\r\nPRINT {copies}\r\n".encode("ascii")

        return header + bytes(raster) + tail

    # ---------------- Printing ----------------
    def print_label(self):
        preview_text = self.preview.toPlainText()
        preview_html = self.preview.toHtml()
        if not (preview_text or "").strip():
            QMessageBox.warning(self, "Печать", "Предпросмотр пуст. Заполните форму или введите текст в предпросмотр.")
            return

        printer_name = win32print.GetDefaultPrinter()

        try:
            copies = self._get_copies()
            tspl_bytes = self._render_preview_to_tspl_bytes(threshold=200, copies=copies)
            n_bytes = print_raw(printer_name, tspl_bytes)
            print(f"SENDING BITMAP... {n_bytes} bytes, copies={copies}")
        except Exception as e:
            QMessageBox.warning(self, "Печать", f"Не удалось отправить на печать:\n{e}")
            return

        self.last_printed_preview_text = preview_text
        self._last_printed_tspl_bytes = tspl_bytes

        product_name = self.product_combo.currentText().strip()
        product = self.get_product(product_name)
        unit_ui = self.unit_combo.currentText()
        unit_code = self._unit_code_from_ui(unit_ui)
        qty = self.qty_input.text().strip().replace(",", ".")

        if product and unit_code and qty:
            try:
                label = build_label(
                    product_name=product["name"],
                    shelf_life_hours=product["shelf_life_hours"],
                    qty_value=qty,
                    unit=unit_code,
                    made_by=self._made_value(),
                    checked_by=self._checked_value(),
                )
                entry = self._build_history_entry_from_label(
                    label,
                    qty_display=qty,
                    unit_ui=unit_ui,
                )
                entry["preview_text"] = preview_text
                entry["preview_html"] = preview_html
            except Exception:
                lines = preview_text.strip().splitlines()
                first_line = (lines[0] if lines else "").strip() or "Этикетка"
                entry = {
                    "id": time.time_ns(),
                    "ts": time.time(),
                    "preview_text": preview_text,
                    "preview_html": preview_html,
                    "product": first_line,
                    "qty": "",
                    "made": "",
                    "checked": "",
                    "time": format_dt(datetime.now()),
                    "batch": "",
                }
        else:
            lines = preview_text.strip().splitlines()
            first_line = (lines[0] if lines else "").strip() or "Этикетка"
            entry = {
                "id": time.time_ns(),
                "ts": time.time(),
                "preview_text": preview_text,
                "preview_html": preview_html,
                "product": first_line,
                "qty": "",
                "made": "",
                "checked": "",
                "time": format_dt(datetime.now()),
                "batch": "",
            }
        self.last_history_entry = entry
        try:
            self._archive_printed_label(preview_text=preview_text, entry=entry, copies=copies)
        except Exception as e:
            sys.stderr.write(f"[MirlisMark] Archive save error: {e}\n")
        # New structured stats journal (JSONL). Must never break printing.
        try:
            _append_stats_entry(
                product=str(entry.get("product_name") or entry.get("product") or ""),
                qty=str(entry.get("qty_value") or ""),
                unit=str(entry.get("unit_ui") or ""),
                made_by=str(entry.get("made_by") or entry.get("made") or ""),
                workshop=str(entry.get("checked_by") or entry.get("checked") or ""),
                batch=str(entry.get("batch") or ""),
                copies=int(copies),
                ts=float(entry.get("ts") or time.time()),
                record_id=str(entry.get("id") or ""),
            )
        except Exception as e:
            sys.stderr.write(f"[MirlisMark] stats_journal append error: {e}\n")
        self._append_history_entry(entry)

    def repeat_last_print(self):
        tspl_bytes = getattr(self, "_last_printed_tspl_bytes", None)
        if not tspl_bytes:
            return
        try:
            printer_name = win32print.GetDefaultPrinter()
            copies = self._get_copies()
            if copies != 1:
                tspl_bytes = tspl_bytes.rsplit(b"\r\nPRINT ", 1)[0] + f"\r\nPRINT {copies}\r\n".encode("ascii")
            print_raw(printer_name, tspl_bytes)
            if self.last_history_entry is not None:
                e = dict(self.last_history_entry)
                e["id"] = time.time_ns()
                e["ts"] = time.time()
                try:
                    self._archive_printed_label(preview_text=e.get("preview_text", ""), entry=e, copies=copies)
                except Exception as arch_e:
                    sys.stderr.write(f"[MirlisMark] Archive save error: {arch_e}\n")
                # New structured stats journal (JSONL). Must never break printing.
                try:
                    _append_stats_entry(
                        product=str(e.get("product_name") or e.get("product") or ""),
                        qty=str(e.get("qty_value") or ""),
                        unit=str(e.get("unit_ui") or ""),
                        made_by=str(e.get("made_by") or e.get("made") or ""),
                        workshop=str(e.get("checked_by") or e.get("checked") or ""),
                        batch=str(e.get("batch") or ""),
                        copies=int(copies),
                        ts=float(e.get("ts") or time.time()),
                        record_id=str(e.get("id") or ""),
                    )
                except Exception as stats_e:
                    sys.stderr.write(f"[MirlisMark] stats_journal append error: {stats_e}\n")
                self._append_history_entry(e)
        except Exception as e:
            QMessageBox.warning(self, "Повтор", f"Не удалось повторить печать:\n{e}")

    # ---------------- Printed labels archive ----------------
    def _labels_archive_root(self) -> str:
        return os.path.join(app_data_dir(), "Готовые этикетки")

    _MONTH_NAMES_RU_NOM = [
        "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
        "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь",
    ]

    def _labels_archive_day_dir(self) -> str:
        now = datetime.now()
        year = f"{now.year:04d}"
        month = f"{now.month:02d} — {self._MONTH_NAMES_RU_NOM[now.month - 1]}"
        day = now.strftime("%d.%m.%Y")
        return os.path.join(self._labels_archive_root(), year, month, day)

    def _sanitize_filename_part(self, s: str) -> str:
        s = (s or "").strip()
        if not s:
            return "—"
        for ch in ['<', '>', ':', '"', '/', '\\', '|', '?', '*']:
            s = s.replace(ch, "_")
        s = s.replace("\n", " ").replace("\r", " ").strip()
        while "  " in s:
            s = s.replace("  ", " ")
        return s[:120].strip() or "—"

    def _cleanup_old_label_archives(self, days: int = 365):
        """Удаляет папки дней старше N дней. Структура: YYYY/MM — Месяц/DD.MM.YYYY/."""
        root = self._labels_archive_root()
        try:
            if not os.path.isdir(root):
                return
            today = datetime.now().date()
            for year_name in list(os.listdir(root)):
                year_path = os.path.join(root, year_name)
                if not os.path.isdir(year_path):
                    continue
                for month_name in list(os.listdir(year_path)):
                    month_path = os.path.join(year_path, month_name)
                    if not os.path.isdir(month_path):
                        continue
                    for day_name in list(os.listdir(month_path)):
                        day_path = os.path.join(month_path, day_name)
                        if not os.path.isdir(day_path):
                            continue
                        try:
                            d = datetime.strptime(day_name.strip(), "%d.%m.%Y").date()
                        except Exception:
                            continue
                        age_days = (today - d).days
                        if age_days >= days:
                            shutil.rmtree(day_path, ignore_errors=True)
                    # Удаляем пустую папку месяца
                    try:
                        if os.path.isdir(month_path) and not os.listdir(month_path):
                            os.rmdir(month_path)
                    except Exception:
                        pass
                # Удаляем пустую папку года
                try:
                    if os.path.isdir(year_path) and not os.listdir(year_path):
                        os.rmdir(year_path)
                except Exception:
                    pass
        except Exception as e:
            sys.stderr.write(f"[MirlisMark] Archive cleanup error: {e}\n")

    def _migrate_labels_archive_layout(self):
        """Однократная миграция: переносит папки DD.MM.YYYY из корня архива
        в новую структуру YYYY/MM — Месяц/DD.MM.YYYY/. Безопасно при повторных запусках."""
        root = self._labels_archive_root()
        try:
            if not os.path.isdir(root):
                return
            for name in list(os.listdir(root)):
                src = os.path.join(root, name)
                if not os.path.isdir(src):
                    continue
                try:
                    d = datetime.strptime(name.strip(), "%d.%m.%Y").date()
                except Exception:
                    continue  # это не папка дня — пропускаем (вероятно YYYY)
                year_dir = os.path.join(root, f"{d.year:04d}")
                month_dir = os.path.join(
                    year_dir,
                    f"{d.month:02d} — {self._MONTH_NAMES_RU_NOM[d.month - 1]}",
                )
                target = os.path.join(month_dir, name)
                try:
                    os.makedirs(month_dir, exist_ok=True)
                except Exception:
                    continue
                if os.path.exists(target):
                    # Целевая папка уже есть — переносим только файлы
                    try:
                        for f in list(os.listdir(src)):
                            src_f = os.path.join(src, f)
                            dst_f = os.path.join(target, f)
                            if not os.path.exists(dst_f):
                                shutil.move(src_f, dst_f)
                        if not os.listdir(src):
                            os.rmdir(src)
                    except Exception:
                        pass
                else:
                    try:
                        shutil.move(src, target)
                    except Exception:
                        pass
        except Exception as e:
            sys.stderr.write(f"[MirlisMark] Archive migration error: {e}\n")

    def _archive_printed_label(self, preview_text: str, entry: dict | None, copies: int):
        # очистка старых папок — при каждом сохранении
        self._cleanup_old_label_archives(days=365)

        e = entry or {}
        product = self._sanitize_filename_part(e.get("product") or e.get("product_name") or "Этикетка")
        weight = self._sanitize_filename_part((e.get("qty") or "").replace(" ", "") or "—")
        batch = self._sanitize_filename_part(e.get("batch") or "—")

        base_name = f"{product}_{weight}_{batch}"
        day_dir = self._labels_archive_day_dir()
        os.makedirs(day_dir, exist_ok=True)

        content = (preview_text or "").rstrip()
        try:
            copies_int = int(copies)
        except Exception:
            copies_int = 1
        content = f"{content}\nКоличество: {copies_int}"

        n = 1
        while True:
            suffix = "" if n == 1 else f"_{n}"
            fname = f"{base_name}{suffix}.txt"
            fpath = os.path.join(day_dir, fname)
            if not os.path.exists(fpath):
                break
            n += 1

        with open(fpath, "w", encoding="utf-8") as f:
            f.write(content + "\n")


def main():
    try:
        import ctypes
        user32 = ctypes.windll.user32
        user32.SetProcessDPIAware()
        screen_w = user32.GetSystemMetrics(0)
        screen_h = user32.GetSystemMetrics(1)
        scale_w = screen_w / 1920
        scale_h = screen_h / 1080
        scale = min(scale_w, scale_h)
        if scale < 0.95:
            os.environ["QT_SCALE_FACTOR"] = str(round(scale, 3))
    except Exception:
        pass

    fmt = QSurfaceFormat()
    fmt.setRenderableType(QSurfaceFormat.OpenGL)
    QSurfaceFormat.setDefaultFormat(fmt)

    app = QApplication(sys.argv)
    main_window = None
    selected_mode = None

    dlg = ModeSelectDialog()
    if dlg.exec_() != QDialog.Accepted or not dlg.selected_mode:
        app.quit()
        return
    selected_mode = dlg.selected_mode

    def on_splash_finished():
        nonlocal main_window
        main_window = MirlisMarkApp(app_mode=selected_mode)
        # Отключаем системный квадрат Windows при долгом касании (press-and-hold)
        if sys.platform == "win32":
            try:
                import ctypes
                hwnd = int(main_window.winId())
                ATOM_NAME = "MicrosoftTabletPenServiceProperty"
                ctypes.windll.kernel32.GlobalAddAtomW(ATOM_NAME)
                DISABLE_FLAGS = (
                    0x00000001 |  # PRESSANDHOLD
                    0x00000008 |  # PENTAPFEEDBACK
                    0x00000010 |  # PENBARRELFEEDBACK
                    0x00010000 |  # FLICKS
                    0x00100000    # FLICKFALLBACKKEYS
                )
                ctypes.windll.user32.SetPropW(hwnd, ATOM_NAME, DISABLE_FLAGS)
            except Exception as e:
                sys.stderr.write(f"[MirlisMark] disable touch feedback: {e}\n")
        main_window.showMaximized()
        # Защита от ресайза окна сенсорной клавиатурой Windows (TabTip)
        # После того как окно развернулось — запоминаем размер и блокируем его.
        try:
            QTimer.singleShot(100, main_window._lock_window_size_against_osk)
        except Exception:
            pass

    video_path = SPLASH_VIDEO_PATH
    if os.path.isfile(video_path):
        splash = SplashVideo(video_path, on_finished_callback=on_splash_finished)
        splash.showFullScreen()
        splash.play()
    else:
        on_splash_finished()

    sys.exit(app.exec_())

if __name__ == "__main__":
    main()



















































