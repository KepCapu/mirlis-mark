from __future__ import annotations

import getpass
import json
import os
import platform
import shutil
import time
import uuid
import zipfile
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from statistics_data import normalize_stat_key


def _export_account_user_name() -> str:
    """
    Имя учётной записи для manifest (Windows / Unix).
    На Windows сначала переменные окружения, затем GetUserNameW — у части окружений
    (службы, некоторые лаунчеры) USERNAME пустой, хотя интерактивный пользователь есть.
    """
    for key in ("USERNAME", "USER", "LOGNAME"):
        v = os.environ.get(key)
        if v and str(v).strip():
            return str(v).strip()
    if os.name == "nt":
        try:
            import ctypes

            advapi32 = ctypes.windll.advapi32
            buf = ctypes.create_unicode_buffer(512)
            size = ctypes.c_uint(len(buf))
            if advapi32.GetUserNameW(buf, ctypes.byref(size)):
                s = (buf.value or "").strip()
                if s:
                    return s
        except Exception:
            pass
    try:
        s = (getpass.getuser() or "").strip()
        if s:
            return s
    except Exception:
        pass
    try:
        s = (os.getlogin() or "").strip()
        if s:
            return s
    except Exception:
        pass
    return ""


def export_package(
    entries: Iterable[dict[str, Any]],
    from_ts: float,
    to_ts: float,
    save_path: str,
    station_uuid: str,
    station_label: str,
    app_version: str,
) -> int:
    """
    Export a portable ZIP package with:
    - manifest.json (utf-8)
    - records.jsonl (utf-8, one JSON object per line)

    Returns number of exported records.
    Does not create a file if there are no records in the period.
    """
    from_ts_f = float(from_ts or 0.0)
    to_ts_f = float(to_ts or 0.0)
    if to_ts_f < from_ts_f:
        from_ts_f, to_ts_f = to_ts_f, from_ts_f

    save_path = str(save_path or "").strip()
    if not save_path:
        raise ValueError("save_path is empty")

    filtered: list[dict[str, Any]] = []
    for e in entries or []:
        if not isinstance(e, dict):
            continue
        try:
            ts = float(e.get("ts") or 0.0)
        except Exception:
            ts = 0.0
        if ts <= 0:
            continue
        if from_ts_f <= ts <= to_ts_f:
            filtered.append(e)

    if not filtered:
        return 0

    package_id = uuid.uuid4().hex
    exported_at = float(time.time())

    if os.name == "nt":
        _computer = str(os.environ.get("COMPUTERNAME") or "").strip() or (platform.node() or "")
    else:
        _computer = str(platform.node() or "").strip()
    _user = _export_account_user_name()

    manifest = {
        "format_version": 1,
        "package_id": package_id,
        "source_station_uuid": str(station_uuid or ""),
        "source_station_label": str(station_label or ""),
        "source_computer_name": _computer,
        "source_user_name": _user,
        "exported_at": exported_at,
        "period_from": from_ts_f,
        "period_to": to_ts_f,
        "record_count": int(len(filtered)),
        "app_version": str(app_version or ""),
    }

    manifest_bytes = json.dumps(manifest, ensure_ascii=False, indent=2).encode("utf-8")
    records_text = "\n".join(json.dumps(r, ensure_ascii=False) for r in filtered) + "\n"
    records_bytes = records_text.encode("utf-8")

    with zipfile.ZipFile(save_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("manifest.json", manifest_bytes)
        zf.writestr("records.jsonl", records_bytes)

    return int(len(filtered))


def export_to_excel(
    entries: Iterable[dict[str, Any]],
    from_ts: float,
    to_ts: float,
    save_path: str,
) -> int:
    """
    Export human-friendly Excel (.xlsx) for the selected period.
    Uses the same entry dict structure as stats_journal.jsonl.

    Returns number of exported records.
    Does not create a file if there are no records in the period.
    """
    from_ts_f = float(from_ts or 0.0)
    to_ts_f = float(to_ts or 0.0)
    if to_ts_f < from_ts_f:
        from_ts_f, to_ts_f = to_ts_f, from_ts_f

    save_path = str(save_path or "").strip()
    if not save_path:
        raise ValueError("save_path is empty")

    # filter by period
    filtered: list[dict[str, Any]] = []
    for e in entries or []:
        if not isinstance(e, dict):
            continue
        try:
            ts = float(e.get("ts") or 0.0)
        except Exception:
            ts = 0.0
        if ts <= 0:
            continue
        if from_ts_f <= ts <= to_ts_f:
            filtered.append(e)

    if not filtered:
        return 0

    # Derived values
    def _dt(ts: float):
        from datetime import datetime

        return datetime.fromtimestamp(float(ts))

    total_ops = len(filtered)
    total_labels = 0
    today = _dt(to_ts_f).date()
    day_shift_total = 0
    night_shift_total = 0

    # day aggregation
    by_day_labels: dict[str, int] = {}
    by_day_ops: dict[str, int] = {}

    # category aggregation: labels + ops
    prod_labels: dict[str, int] = {}
    prod_ops: dict[str, int] = {}
    staff_labels: dict[str, int] = {}
    staff_ops: dict[str, int] = {}
    ws_labels: dict[str, int] = {}
    ws_ops: dict[str, int] = {}

    uniq_products: set[str] = set()
    uniq_staff: set[str] = set()
    uniq_ws: set[str] = set()

    def _inc(d: dict[str, int], k: str, n: int = 1) -> None:
        d[k] = int(d.get(k, 0)) + int(n)

    for e in filtered:
        try:
            ts = float(e.get("ts") or 0.0)
        except Exception:
            continue
        if ts <= 0:
            continue
        dt = _dt(ts)
        date_key = dt.strftime("%d.%m.%Y")

        try:
            copies = int(e.get("copies") or 1)
        except Exception:
            copies = 1
        if copies <= 0:
            copies = 1
        total_labels += copies

        _inc(by_day_labels, date_key, copies)
        _inc(by_day_ops, date_key, 1)

        # shifts: only "today" (same as dashboard KPI logic)
        if dt.date() == today:
            if 8 <= int(dt.hour) < 20:
                day_shift_total += copies
            else:
                night_shift_total += copies

        p = normalize_stat_key(e.get("product"))
        s = normalize_stat_key(e.get("made_by"))
        w = normalize_stat_key(e.get("workshop"))

        if p:
            _inc(prod_labels, p, copies)
            _inc(prod_ops, p, 1)
            uniq_products.add(p)
        if s:
            _inc(staff_labels, s, copies)
            _inc(staff_ops, s, 1)
            uniq_staff.add(s)
        if w:
            _inc(ws_labels, w, copies)
            _inc(ws_ops, w, 1)
            uniq_ws.add(w)

    # Build workbook
    wb = Workbook()

    header_font = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    def _autosize(ws) -> None:
        for col in range(1, ws.max_column + 1):
            max_len = 0
            for row in range(1, ws.max_row + 1):
                v = ws.cell(row=row, column=col).value
                if v is None:
                    continue
                s = str(v)
                if len(s) > max_len:
                    max_len = len(s)
            ws.column_dimensions[get_column_letter(col)].width = min(60, max(10, max_len + 2))

    # Sheet 1: Summary
    ws = wb.active
    ws.title = "Сводка"
    ws.append(["Период отчёта", f"{_dt(from_ts_f).strftime('%d.%m.%Y %H:%M')} — {_dt(to_ts_f).strftime('%d.%m.%Y %H:%M')}"])
    ws.append(["Всего этикеток", total_labels])
    ws.append(["Всего операций", total_ops])
    ws.append(["Дневная смена (сегодня)", day_shift_total])
    ws.append(["Ночная смена (сегодня)", night_shift_total])
    ws.append(["Уникальных продуктов", len(uniq_products)])
    ws.append(["Уникальных сотрудников", len(uniq_staff)])
    ws.append(["Уникальных цехов", len(uniq_ws)])
    ws["A1"].font = header_font
    ws["B1"].alignment = wrap
    _autosize(ws)

    # Sheet 2: By days
    ws = wb.create_sheet("По дням")
    ws.append(["Дата", "Этикеток", "Операций"])
    for c in range(1, 4):
        ws.cell(row=1, column=c).font = header_font
    ws.freeze_panes = "A2"
    for day in sorted(by_day_labels.keys(), key=lambda s: _dt(time.mktime(time.strptime(s, "%d.%m.%Y")))):
        ws.append([day, int(by_day_labels.get(day, 0)), int(by_day_ops.get(day, 0))])
    _autosize(ws)

    def _top_sheet(title: str, labels_d: dict[str, int], ops_d: dict[str, int]) -> None:
        wsx = wb.create_sheet(title)
        wsx.append([title[:-0] if False else title.replace("Топ ", ""), "Этикеток", "Операций", "Доля %"])
        # Override header names per spec
        wsx["A1"].value = "Категория"
        if title == "Топ продуктов":
            wsx["A1"].value = "Продукт"
        elif title == "Топ сотрудников":
            wsx["A1"].value = "Сотрудник"
        elif title == "По цехам":
            wsx["A1"].value = "Цех"
        for c in range(1, 5):
            wsx.cell(row=1, column=c).font = header_font
        wsx.freeze_panes = "A2"
        items = sorted(labels_d.items(), key=lambda kv: int(kv[1]), reverse=True)
        for name, lbls in items:
            ops = int(ops_d.get(name, 0))
            share = (float(lbls) / float(total_labels) * 100.0) if total_labels > 0 else 0.0
            wsx.append([name, int(lbls), ops, round(share, 1)])
        _autosize(wsx)

    # Sheets 3-5
    _top_sheet("Топ продуктов", prod_labels, prod_ops)
    _top_sheet("Топ сотрудников", staff_labels, staff_ops)
    _top_sheet("По цехам", ws_labels, ws_ops)

    # Sheet 6: Raw records
    ws = wb.create_sheet("Сырые записи")
    ws.append([
        "Дата/время",
        "Продукт",
        "Количество",
        "Ед.",
        "Изготовил",
        "Цех / Проверил",
        "Партия",
        "Копий",
        "station_label",
        "station_uuid",
        "record_id",
    ])
    for c in range(1, 12):
        ws.cell(row=1, column=c).font = header_font
    ws.freeze_panes = "A2"
    for e in sorted(filtered, key=lambda x: float(x.get("ts") or 0.0)):
        try:
            dt = _dt(float(e.get("ts") or 0.0)).strftime("%d.%m.%Y %H:%M:%S")
        except Exception:
            dt = ""
        ws.append([
            dt,
            str(e.get("product") or ""),
            str(e.get("qty") or ""),
            str(e.get("unit") or ""),
            str(e.get("made_by") or ""),
            str(e.get("workshop") or ""),
            str(e.get("batch") or ""),
            int(e.get("copies") or 1),
            str(e.get("station_label") or ""),
            str(e.get("station_uuid") or ""),
            str(e.get("record_id") or ""),
        ])
    _autosize(ws)

    wb.save(save_path)
    return int(len(filtered))


def _app_data_dir() -> str:
    if os.name == "nt":
        root = os.environ.get("LOCALAPPDATA", os.path.expanduser("~"))
    else:
        root = os.path.expanduser("~")
    path = os.path.join(root, "MirlisMark")
    os.makedirs(path, exist_ok=True)
    return path


def _imported_dir() -> str:
    p = os.path.join(_app_data_dir(), "imported")
    os.makedirs(p, exist_ok=True)
    return p


def _sources_registry_path() -> str:
    return os.path.join(_app_data_dir(), "imported_sources.json")


def _load_sources_registry() -> list[dict[str, Any]]:
    p = _sources_registry_path()
    if not os.path.isfile(p):
        return []
    try:
        with open(p, "r", encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, list) else []
    except Exception:
        return []


def _save_sources_registry(items: list[dict[str, Any]]) -> None:
    try:
        with open(_sources_registry_path(), "w", encoding="utf-8") as f:
            json.dump(items or [], f, ensure_ascii=False, indent=2)
    except Exception:
        return


def list_imported_sources() -> list[dict[str, Any]]:
    """Return imported_sources.json as a list (may be empty)."""
    return list(_load_sources_registry() or [])


def resolve_imported_zip_path(record: dict[str, Any]) -> str:
    """
    Filesystem path to the imported package zip for a registry record.
    Uses local_path when present; otherwise the default layout under imported/.
    Path may point to a missing file (caller should check os.path.isfile).
    """
    p = str((record or {}).get("local_path") or "").strip()
    if p:
        return p
    pid = str((record or {}).get("package_id") or "").strip()
    if not pid:
        return ""
    return os.path.join(_imported_dir(), f"{pid}.zip")


def _read_zip_text(zf: zipfile.ZipFile, name: str) -> str:
    try:
        b = zf.read(name)
    except KeyError:
        raise FileNotFoundError(f"{name} not found in zip")
    try:
        return b.decode("utf-8-sig")
    except Exception:
        return b.decode("utf-8", errors="replace")


def inspect_package(zip_path: str) -> dict[str, Any]:
    """Read and return manifest.json without importing (raises on errors)."""
    zip_path = str(zip_path or "").strip()
    if not zip_path:
        raise ValueError("zip_path is empty")
    if not os.path.isfile(zip_path):
        raise FileNotFoundError(zip_path)
    with zipfile.ZipFile(zip_path, "r") as zf:
        manifest_text = _read_zip_text(zf, "manifest.json")
        try:
            m = json.loads(manifest_text)
        except Exception as e:
            raise ValueError(f"manifest.json is not valid JSON: {e}") from e
        if not isinstance(m, dict):
            raise ValueError("manifest.json must be an object")
        if not str(m.get("package_id") or "").strip():
            raise ValueError("manifest.json missing package_id")
        # also validate records exist
        try:
            zf.getinfo("records.jsonl")
        except KeyError as e:
            raise FileNotFoundError("records.jsonl not found in zip") from e
        return m


def read_package_manifest_dict(zip_path: str) -> dict[str, Any] | None:
    """
    Прочитать manifest.json из zip без полной валидации пакета (для подписей в UI).
    Возвращает dict или None при ошибке.
    """
    p = str(zip_path or "").strip()
    if not p or not os.path.isfile(p):
        return None
    try:
        with zipfile.ZipFile(p, "r") as zf:
            text = _read_zip_text(zf, "manifest.json")
        m = json.loads(text)
        return m if isinstance(m, dict) else None
    except Exception:
        return None


def import_package(zip_path: str) -> dict[str, Any]:
    """
    Import a previously exported stats package zip into:
      %LOCALAPPDATA%\\MirlisMark\\imported\\<package_id>.zip
    and register it in imported_sources.json.

    Returns the registry record that was added.
    Raises ValueError/FileNotFoundError on errors.
    """
    m = inspect_package(zip_path)
    package_id = str(m.get("package_id") or "").strip()
    source_station_uuid = str(m.get("source_station_uuid") or "").strip()

    reg = _load_sources_registry()
    for it in reg:
        if str(it.get("package_id") or "").strip() == package_id:
            raise ValueError("Этот пакет уже импортирован (package_id совпадает).")

    # Prevent self-import (same station) by default: it would create duplicates.
    try:
        from stats_store import ensure_station_identity as _ensure_station_identity

        local_station_uuid, _local_station_label = _ensure_station_identity()
        if source_station_uuid and str(local_station_uuid or "").strip() == source_station_uuid:
            raise ValueError(
                "Этот пакет был создан на текущей станции/этом ПК. "
                "Импорт приведёт к дублям и не требуется."
            )
    except ValueError:
        raise
    except Exception:
        # If we cannot determine station identity, don't block import.
        pass

    # copy zip locally
    local_zip = os.path.join(_imported_dir(), f"{package_id}.zip")
    os.makedirs(os.path.dirname(local_zip), exist_ok=True)
    shutil.copy2(zip_path, local_zip)

    archive_basename = os.path.basename(str(zip_path).replace("\\", "/"))

    record = {
        "package_id": package_id,
        "source_station_uuid": str(m.get("source_station_uuid") or ""),
        "source_station_label": str(m.get("source_station_label") or ""),
        "source_archive_basename": archive_basename,
        "source_computer_name": str(m.get("source_computer_name") or "").strip(),
        "source_user_name": str(m.get("source_user_name") or "").strip(),
        "imported_at": float(time.time()),
        "period_from": float(m.get("period_from") or 0.0),
        "period_to": float(m.get("period_to") or 0.0),
        "record_count": int(m.get("record_count") or 0),
        "local_path": local_zip,
    }
    reg.append(record)
    _save_sources_registry(reg)
    return record


def read_imported_entries() -> list[dict[str, Any]]:
    """Read records.jsonl from all imported packages. Bad lines are skipped."""
    out: list[dict[str, Any]] = []
    for src in _load_sources_registry():
        p = str(src.get("local_path") or "").strip()
        if not p or not os.path.isfile(p):
            continue
        try:
            with zipfile.ZipFile(p, "r") as zf:
                text = _read_zip_text(zf, "records.jsonl")
            for line in (text or "").splitlines():
                s = (line or "").strip()
                if not s:
                    continue
                try:
                    obj = json.loads(s)
                    if isinstance(obj, dict):
                        out.append(obj)
                except Exception:
                    continue
        except Exception:
            continue
    return out


def remove_imported_source(package_id: str) -> bool:
    """
    Remove an imported source by package_id:
    - delete the local zip file (if present)
    - remove the registry entry from imported_sources.json
    Local stats_journal.jsonl is never touched.
    Returns True if an entry was removed, False otherwise.
    Never raises.
    """
    try:
        pid = str(package_id or "").strip()
        if not pid:
            return False
        reg = _load_sources_registry()
        kept: list[dict[str, Any]] = []
        removed: dict[str, Any] | None = None
        for it in reg:
            if str(it.get("package_id") or "").strip() == pid and removed is None:
                removed = it
                continue
            kept.append(it)
        if removed is None:
            return False

        # Remove local file if it exists.
        try:
            p = str(removed.get("local_path") or "").strip()
            if p and os.path.isfile(p):
                os.remove(p)
        except Exception:
            pass

        _save_sources_registry(kept)
        return True
    except Exception:
        return False

